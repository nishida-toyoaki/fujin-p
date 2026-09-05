"""エクセルとのやり取り。

方眼（原セルの列）と行の重なりを，そのままシートの列と行に落とす。
合併したところは結合セルにし，列幅は表示幅から出す。
読み戻せるように「対応」シートを添え，どのセルがどの倉庫部品かを書いておく。
"""

import io
import re

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from . import store

SHEET_MAIN = '表'
SHEET_MAP = '対応'
CELL_PX = 64.0            # 原セル1個の見かけの幅（設計 px）
CHAR_PER_PX = 7.0         # エクセルの列幅1文字 ≒ 7px
ROW_PT = 15.0             # 1行の高さ（pt）
THIN = Side(style='thin', color='B0B8C4')
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _plain(html):
    """HTML から文字だけ取り出す（改行は残す）。"""
    t = re.sub(r'(?i)<br\s*/?>', '\n', html or '')
    t = re.sub(r'(?i)</(p|div|li|tr|h[1-6])>', '\n', t)
    t = re.sub(r'<[^>]+>', '', t)
    t = t.replace('&nbsp;', ' ').replace('&amp;', '&').replace('&lt;', '<').replace('&gt;', '>')
    return re.sub(r'\n{3,}', '\n\n', t).strip()


def _text_of(box):
    """倉庫部品の中身を，エクセルに置く文字にする。"""
    if not box:
        return ''
    if box.get('recipe') == 'text':
        return _plain(box.get('body') or '')
    lines = [(k.get('body') or '').strip() for k in store.children(box['id']) if k['kind'] == 'item']
    return '\n'.join('・' + x for x in lines if x)


def _grid_w(arena):
    z = (store._layout(arena).get('size') or {})
    if z.get('u') == 'cell' and z.get('w'):
        return int(z['w'])
    cells = []
    _walk(arena, arena, [], 0, 10 ** 6, 0, cells, {})
    return max([c['c1'] for c in cells if c['c1'] < 10 ** 5] or [1])


def _walk(root, arena, path, a, b, row, cells, seen):
    """方眼への割り当て。戻り値は使った行数。埋め込みは親の原セルに1対1で写す。"""
    return _walk_node(arena, store._layout(arena).get('root'), a, b, row, cells, 0)


def _walk_node(arena, node, a, b, row, cells, depth):
    if not isinstance(node, dict) or depth > 6:
        return 0
    t = node.get('t')
    if t == 'row':
        kids = node.get('items') or []
        ws = []
        for k in kids:
            z = (k.get('size') or {})
            ws.append(int(z['w']) if z.get('u') == 'cell' and z.get('w') else None)
        fixed = sum(x for x in ws if x)
        miss = len([x for x in ws if x is None])
        each = max(1, (b - a - fixed) // miss) if miss else 0
        c, h = a, 0
        for k, w in zip(kids, ws):
            w = w if w is not None else each
            h = max(h, _walk_node(arena, k, c, c + w, row, cells, depth))
            c += w
        return max(1, h)
    if t == 'col':
        r, h = row, 0
        for k in (node.get('items') or []):
            n = _walk_node(arena, k, a, b, r, cells, depth)
            r += n
            h += n
        return max(1, h)
    if t == 'text':
        q = store.resolve(node.get('ref'))
        if q and q.get('recipe') == 'arena':
            z = (store._layout(q).get('size') or {})
            wid = int(z['w']) if z.get('u') == 'cell' and z.get('w') else (b - a)
            return _walk_node(q, store._layout(q).get('root'), a, a + wid, row, cells, depth + 1)
        cells.append({'r0': row, 'r1': row + 1, 'c0': a, 'c1': b,
                      'label': node.get('label') or '',
                      'ref': str(node.get('ref') or ''),
                      'key': (q.get('key_path') or '') if q else '',
                      'qid': q['id'] if q else None,
                      'recipe': (q.get('recipe') or '') if q else '',
                      'text': _text_of(q) if q else ''})
        return 1
    return 0


def layout_cells(arena):
    """作品を方眼に割り当てた結果（部品ごとの行と列の範囲，中身）。"""
    cells = []
    _walk_node(arena, store._layout(arena).get('root'), 0, _grid_w(arena), 0, cells, 0)
    return cells


def build_xlsx(arena_id):
    """作品を xlsx にする。"""
    a = store.get_quantum(arena_id)
    if not a or a.get('recipe') != 'arena':
        return None
    cells = layout_cells(a)
    W = _grid_w(a)
    lay = store._layout(a)
    cellw = (lay.get('view') or {}).get('cellw') or []

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_MAIN
    ws.sheet_view.showGridLines = False

    for i in range(W):
        px = float(cellw[i]) if i < len(cellw) else CELL_PX
        ws.column_dimensions[get_column_letter(i + 1)].width = max(1.5, px / CHAR_PER_PX)

    head = Font(bold=True)
    fill = PatternFill('solid', fgColor='F1F4F8')
    for c in cells:
        r, col = c['r0'] + 1, c['c0'] + 1
        cell = ws.cell(row=r, column=col, value=c['text'])
        cell.alignment = Alignment(wrap_text=True, vertical='top')
        cell.border = BOX
        if not c['text'] and c['label']:
            cell.value = c['label']
            cell.font = head
            cell.fill = fill
        if c['c1'] - c['c0'] > 1 or c['r1'] - c['r0'] > 1:
            ws.merge_cells(start_row=r, start_column=col,
                           end_row=c['r1'], end_column=min(W, c['c1']))
    # 行の高さ：文字数と幅から見当をつける
    for r in range(1, max([c['r1'] for c in cells] or [1]) + 1):
        n = 1
        for c in cells:
            if c['r0'] + 1 != r:
                continue
            w = sum(float(cellw[i]) if i < len(cellw) else CELL_PX for i in range(c['c0'], min(W, c['c1'])))
            per = max(4, int(w / 14))
            n = max(n, sum(max(1, (len(ln) // per) + 1) for ln in (c['text'] or ' ').split('\n')))
        ws.row_dimensions[r].height = min(400, ROW_PT * n)

    mp = wb.create_sheet(SHEET_MAP)
    mp.append(['セル', '行', '列', '見出し', '倉庫のkey_path', '倉庫のID', '型', '指し先'])
    for c in mp['A1:H1'][0]:
        c.font = head
    for c in cells:
        mp.append(['%s%d' % (get_column_letter(c['c0'] + 1), c['r0'] + 1), c['r0'] + 1, c['c0'] + 1,
                   c['label'], c['key'], c['qid'], c['recipe'], c['ref']])
    for col, w in zip('ABCDEFGH', (10, 6, 6, 24, 26, 10, 10, 16)):
        mp.column_dimensions[col].width = w
    mp.append([])
    mp.append(['※ この「対応」シートは，読み戻すときに使います。列や行を並べ替えないでください。'])
    mp.append(['※ 「表」シートのセルを直して取り込むと，その倉庫部品の中身が置き換わります。'])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def read_xlsx(data):
    """xlsx を読んで {倉庫のID: 文字} を返す（対応シートを手掛かりにする）。"""
    wb = load_workbook(io.BytesIO(data), data_only=True)
    if SHEET_MAP not in wb.sheetnames or SHEET_MAIN not in wb.sheetnames:
        return None, 'こんかが書き出した xlsx ではありません（「表」と「対応」のシートが要ります）'
    mp, ws = wb[SHEET_MAP], wb[SHEET_MAIN]
    out = {}
    for row in mp.iter_rows(min_row=2, values_only=True):
        if not row or not row[0] or not isinstance(row[0], str):
            continue
        addr, key, qid, recipe = row[0], row[4], row[5], row[6]
        if not qid:
            continue
        try:
            v = ws[addr].value
        except (KeyError, ValueError):
            continue
        out[int(qid)] = {'text': '' if v is None else str(v), 'key': key or '', 'recipe': recipe or ''}
    return out, None


def apply_values(values, me, wperm, wacl, actor_id):
    """読んだ中身を倉庫部品へ書き戻す。権限の無いものは飛ばす。"""
    rep = {'updated': 0, 'skipped': [], 'same': 0}
    for qid, v in (values or {}).items():
        q = store.get_quantum(qid)
        if not q or q.get('kind') != 'box':
            rep['skipped'].append({'id': qid, 'why': '倉庫にありません'})
            continue
        if not store.box_perm(q, me, wperm, wacl).get('write'):
            rep['skipped'].append({'id': qid, 'title': q.get('title') or '', 'why': '書く権限がありません'})
            continue
        text = (v.get('text') or '').replace('\r\n', '\n').strip()
        if _plain_of(q).strip() == text:            # 変わっていなければ触らない
            rep['same'] += 1
            continue
        if q.get('recipe') == 'text':
            html = '\n'.join('<p>%s</p>' % _esc(ln) for ln in text.split('\n') if ln.strip()) or ''
            store.demand(q, 'distribute', {'value': html, 'actor_id': actor_id, 'origin': 'xlsx'})
        else:
            store.demand(q, 'distribute', {'value': text, 'actor_id': actor_id, 'origin': 'xlsx'})
        rep['updated'] += 1
    return rep


def _plain_of(q):
    return _plain(q.get('body') or '') if q.get('recipe') == 'text' else _text_of(q)


def _esc(t):
    return (t or '').replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
