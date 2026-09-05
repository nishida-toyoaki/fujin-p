# SPDX-FileCopyrightText: 2024-2026 Toyoaki Nishida
# SPDX-License-Identifier: AGPL-3.0-or-later
#
# This file is part of FUJIN-P.
# Copyright (C) 2024-2026 Toyoaki Nishida
#
# FUJIN-P is free software: you can redistribute it and/or modify
# it under the terms of the GNU Affero General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# FUJIN-P is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU Affero General Public License for more details.
#
# You should have received a copy of the GNU Affero General Public License
# along with FUJIN-P.  If not, see <https://www.gnu.org/licenses/>.
#
# Source: https://github.com/nishida-toyoaki/fujin-p

"""方眼（Grid）—— 絵エクセルの読み書きと，行の署名。

エクセルのうち，テーコンが見るのは次の四つだけである。

  1. どのセルに何が書いてあるか
  2. どのセルが縦横いくつ分に合併されているか
  3. 列の幅と行の高さ
  4. 罫線・塗り・寄せ（見かけ。分解では見ず，生成のときに様式から与える）

1と2が構造で，3と4は見かけである。対応式が扱うのは 1 と 2 で，
3 と 4 は様式の側（対応式の styles）に置く。
"""

import io
import re

RE_WS = re.compile(r'[\s\u3000]+')


def col_letter(c):
    s = ''
    while c > 0:
        c, r = divmod(c - 1, 26)
        s = chr(65 + r) + s
    return s


def col_index(s):
    n = 0
    for ch in (s or '').upper():
        if 'A' <= ch <= 'Z':
            n = n * 26 + (ord(ch) - 64)
    return n


def norm(v):
    """比べるための文字。全角空白も畳み，前後を落とす。"""
    if v is None:
        return ''
    t = str(v).replace('\r\n', '\n').replace('\r', '\n')
    t = '\n'.join(RE_WS.sub(' ', ln).strip() for ln in t.split('\n'))
    return t.strip()


class Grid(object):
    """1枚のシート。番地は 1 始まり（エクセルと同じ）。"""

    def __init__(self, name=''):
        self.name = name
        self.rows = 0
        self.cols = 0
        self.values = {}        # (r,c) -> 文字（合併の左上だけ）
        self.spans = {}         # (r,c) -> (rs,cs)  合併の左上だけ
        self.covered = set()    # 合併に呑まれたところ
        self.col_widths = {}    # c -> 幅（エクセルの文字数）
        self.row_heights = {}   # r -> 高さ（pt）
        self.styles = {}        # (r,c) -> 見かけの名前（対応式の styles を引く）

    # ---------------------------------------------------------------- 参照
    def val(self, r, c):
        return self.values.get((r, c), '')

    def span(self, r, c):
        return self.spans.get((r, c), (1, 1))

    def anchor(self, r, c):
        """そこに枠があるか（中身があるか，合併の左上か）。"""
        if (r, c) in self.covered:
            return False
        return bool(self.values.get((r, c))) or (r, c) in self.spans

    def sig(self, r):
        """行の署名。B2x1,C1x5 のような形。行型の見分けに使う。"""
        out = []
        for c in range(1, self.cols + 1):
            if not self.anchor(r, c):
                continue
            rs, cs = self.span(r, c)
            out.append(col_letter(c) + ('' if (rs, cs) == (1, 1) else '%dx%d' % (rs, cs)))
        return ','.join(out) or '-'

    def is_blank(self, r):
        for c in range(1, self.cols + 1):
            if self.anchor(r, c):
                return False
        return True

    def put(self, r, c, v, rs=1, cs=1, style=None):
        self.values[(r, c)] = '' if v is None else str(v)
        if style:
            self.styles[(r, c)] = style
        if (rs, cs) != (1, 1):
            self.spans[(r, c)] = (int(rs), int(cs))
            for rr in range(r, r + int(rs)):
                for cc in range(c, c + int(cs)):
                    if (rr, cc) != (r, c):
                        self.covered.add((rr, cc))
        self.rows = max(self.rows, r + int(rs) - 1)
        self.cols = max(self.cols, c + int(cs) - 1)


# ================================================================ 読み込み

def read_xlsx(data, sheet=None):
    """xlsx のバイト列を読んで {シート名: Grid} を返す。"""
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data), data_only=True)
    out = {}
    for ws in wb.worksheets:
        if sheet and ws.title != sheet:
            continue
        out[ws.title] = _grid_of(ws)
    return out


def _grid_of(ws):
    g = Grid(ws.title)
    for rng in ws.merged_cells.ranges:
        rs = rng.max_row - rng.min_row + 1
        cs = rng.max_col - rng.min_col + 1
        g.spans[(rng.min_row, rng.min_col)] = (rs, cs)
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                if (r, c) != (rng.min_row, rng.min_col):
                    g.covered.add((r, c))
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            if (cell.row, cell.column) in g.covered:
                continue
            g.values[(cell.row, cell.column)] = str(cell.value)
    g.rows = ws.max_row or 0
    g.cols = ws.max_column or 0
    for key, dim in (ws.column_dimensions or {}).items():
        if dim.width:
            g.col_widths[col_index(key)] = float(dim.width)
    for key, dim in (ws.row_dimensions or {}).items():
        if dim.height:
            try:
                g.row_heights[int(key)] = float(dim.height)
            except (TypeError, ValueError):
                pass
    return g


# ================================================================ 書き出し

def write_xlsx(grids, styles=None):
    """{シート名: Grid} を xlsx のバイト列にする。"""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    styles = styles or {}
    thin = Side(style='thin', color='808080')
    boxed = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = Workbook()
    wb.remove(wb.active)
    for name, g in grids.items():
        ws = wb.create_sheet(name[:31] or 'Sheet1')
        ws.sheet_view.showGridLines = False
        for (r, c), v in sorted(g.values.items()):
            cell = ws.cell(row=r, column=c, value=v)
            st = styles.get(g.styles.get((r, c))) or {}
            cell.alignment = Alignment(wrap_text=True,
                                       vertical=st.get('valign', 'top'),
                                       horizontal=st.get('align'))
            if st.get('bold'):
                cell.font = Font(bold=True)
            if st.get('fill'):
                cell.fill = PatternFill('solid', fgColor=st['fill'])
            if st.get('border', True):
                cell.border = boxed
        for (r, c), (rs, cs) in sorted(g.spans.items()):
            if (rs, cs) != (1, 1):
                ws.merge_cells(start_row=r, start_column=c,
                               end_row=r + rs - 1, end_column=c + cs - 1)
        for c, w in g.col_widths.items():
            ws.column_dimensions[get_column_letter(c)].width = w
        for r, h in g.row_heights.items():
            ws.row_dimensions[r].height = h
    if not wb.sheetnames:
        wb.create_sheet('Sheet1')
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ================================================================ 画面用

def to_html(g, r0=1, r1=None, mark=None, max_chars=60):
    """方眼をそのまま HTML の表にする（観察・プレビュー用）。

    mark は (r,c) -> 印の文字列。帯に取られたところに色を付けるのに使う。
    """
    r1 = min(g.rows, r1 or g.rows)
    mark = mark or {}
    out = ['<table class="grid">']
    out.append('<tr><th class="rn"></th>')
    for c in range(1, g.cols + 1):
        out.append('<th>%s</th>' % col_letter(c))
    out.append('</tr>')
    for r in range(r0, r1 + 1):
        out.append('<tr><th class="rn">%d</th>' % r)
        for c in range(1, g.cols + 1):
            if (r, c) in g.covered:
                continue
            rs, cs = g.span(r, c)
            v = g.val(r, c)
            short = v if len(v) <= max_chars else v[:max_chars] + '…'
            klass = mark.get((r, c), '')
            attr = ''
            if rs > 1:
                attr += ' rowspan="%d"' % rs
            if cs > 1:
                attr += ' colspan="%d"' % cs
            out.append('<td%s class="%s" title="%s">%s</td>'
                       % (attr, klass, _esc(v)[:400], _esc(short).replace('\n', '<br>')))
        out.append('</tr>')
    out.append('</table>')
    return ''.join(out)


def _esc(t):
    return (t or '').replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;').replace('"', '&quot;')


def outline(g, limit=200):
    """行の署名と先頭の文字を並べた，生成AIに見せるための素描。"""
    lines = []
    for r in range(1, min(g.rows, limit) + 1):
        head = ''
        for c in range(1, g.cols + 1):
            if g.anchor(r, c) and g.val(r, c):
                head = norm(g.val(r, c))[:48]
                break
        lines.append('%4d  %-42s  %s' % (r, g.sig(r), head))
    if g.rows > limit:
        lines.append('... （以下 %d 行省略）' % (g.rows - limit))
    return '\n'.join(lines)


def sig_census(g):
    """署名の出現回数。少ないパターンで出来ていることを見せる。"""
    from collections import Counter
    cnt = Counter(g.sig(r) for r in range(1, g.rows + 1))
    return cnt.most_common()
