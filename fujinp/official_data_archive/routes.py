# SPDX-FileCopyrightText: 2024-2026 Toyoaki Nishida
# SPDX-License-Identifier: AGPL-3.0-or-later
#
# This file is part of FUJIN-P.
# Copyright (C) 2024-2026 Toyoaki Nishida
#
# FUJIN-P is free software: you can redistribute it and/or modify
# it under the terms of the GNU Affero General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# FUJIN-P is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU Affero General Public License for more details.
#
# You should have received a copy of the GNU Affero General Public License
# along with FUJIN-P.  If not, see <https://www.gnu.org/licenses/>.
#
# Source: https://github.com/nishida-toyoaki/fujin-p

"""
official_data_archive routes.py
公式データ集 — 公式テーブルの登録簿管理と読み取り専用ビューア

設計方針:
  - 登録簿テーブル official_data_archive_tables（<owner>$default）に、
    公開対象のSQLテーブル名とその所在DB（default / fujinp / public）を登録する。
  - データへのアクセスは「登録簿に載っているテーブル」のみ許可
    （テーブル名の正規表現検査ではなく、DB照合による許可リスト方式）。
  - 削除は登録簿からの除外のみ。SQLテーブル自体は絶対に DROP しない。
  - 閲覧: users.category が regular / admin
  - 管理（追加・削除）: admin、または まいぐるグループ「公式データ集_管理者」
"""
import os
import re
import math
import datetime
import logging

from pytz import timezone
from flask import render_template, request, jsonify, session
import mysql.connector

from auth import redirect_to_dashboard
from config import Config
from db import DatabaseConfig
from decorators import login_required
from ..user_groups.utils import user_is_in_group

from . import official_data_archive_bp

JST = timezone('Asia/Tokyo')

# ────────────────────────────────────────────
# 定数
# ────────────────────────────────────────────

# 管理グループの命名規約：
#   まいぐる上で「公式データ集_」で始まる名前のグループが管理グループになる。
#   例：公式データ集_財務、公式データ集_教務 …（小刻みに定義できる）
#   各アイテムには登録時に管理グループを1つ紐づけ、その所属者が管理できる。
GROUP_PREFIX = '公式データ集_'

# 全体管理者グループ（admin カテゴリと同等：全アイテムを管理できる）
GROUP_GLOBAL = '公式データ集_管理者'

# 閲覧を許可する users.category
VIEW_CATEGORIES = ('admin', 'regular')

# アイテムの所在DBの選択肢。キーを登録簿の database_name に保存する。
# 接続情報そのものは db.py に集約されている（実値をここに書かない）。
DB_CHOICES = {
    'default': DatabaseConfig.default,
    'fujinp':  DatabaseConfig.fujinp,
    'public':  DatabaseConfig.public,
}

# テーブル名として許可する文字（追加時の検査。アクセス時は登録簿照合のみ）
# 日本語名・中点（・）を許可し、SQL組み立て上危険な文字だけを禁止する
# （空白・バッククォート・引用符・% ・; ・\ ・. ・/）。長さはMySQL上限の64文字。
TABLE_NAME_RE = re.compile(r'^[^\s`\'\"%;\\./]{1,64}$')

PAGE_LIMIT_MAX = 5000


def get_jst_now():
    """現在の日時をJSTで取得（naive datetime）。INSERT/UPDATEに使う。"""
    return datetime.datetime.now(JST).replace(tzinfo=None)


# ────────────────────────────────────────────
# 権限ヘルパー
# ────────────────────────────────────────────

def _get_user_category(user_id):
    """users.category を返す。"""
    try:
        conn = mysql.connector.connect(**DatabaseConfig.default())
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT category FROM users WHERE id = %s", (user_id,))
        row = cursor.fetchone()
        return row['category'] if row else None
    except Exception as e:
        logging.error("official_data_archive _get_user_category error: %s", e)
        return None
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close(); conn.close()


def can_view(user_id):
    """閲覧権限：regular 以上。"""
    return _get_user_category(user_id) in VIEW_CATEGORIES


def _user_effective_group_ids(user_id):
    """
    ユーザの有効な所属グループID集合。
    まいぐるが公開する get_user_effective_group_ids() を使い、
    取得できない場合のみ有効期間つき所属を直接照合する。
    """
    try:
        from ..user_groups.utils import get_user_effective_group_ids
        return set(get_user_effective_group_ids(user_id))
    except Exception:
        pass
    try:
        now = get_jst_now()
        conn = mysql.connector.connect(**DatabaseConfig.default())
        cursor = conn.cursor()
        cursor.execute("""
            SELECT group_id FROM user_group_memberships
            WHERE user_id = %s
              AND (valid_from  IS NULL OR valid_from  <= %s)
              AND (valid_until IS NULL OR valid_until >= %s)
        """, (user_id, now, now))
        return {r[0] for r in cursor.fetchall()}
    except Exception as e:
        logging.error("official_data_archive _user_effective_group_ids: %s", e)
        return set()
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close(); conn.close()


def _prefixed_manager_groups():
    """
    名前が GROUP_PREFIX で始まる まいぐるグループ一覧 [{id, name}] を返す。
    これが「管理グループ」の全体集合（全体管理者グループも含む）。
    """
    try:
        conn = mysql.connector.connect(**DatabaseConfig.default())
        cursor = conn.cursor(dictionary=True)
        # アンダースコアはLIKEのワイルドカードなのでエスケープして前方一致
        pattern = GROUP_PREFIX.replace('\\', '\\\\').replace('_', '\\_') + '%'
        cursor.execute("""
            SELECT id, name FROM user_groups
            WHERE name LIKE %s
            ORDER BY name
        """, (pattern,))
        return cursor.fetchall()
    except Exception as e:
        logging.error("official_data_archive _prefixed_manager_groups: %s", e)
        return []
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close(); conn.close()


def is_global_manager(user_id):
    """全体管理者：admin カテゴリ、または GROUP_GLOBAL の所属者。"""
    if _get_user_category(user_id) == 'admin':
        return True
    try:
        return user_is_in_group(user_id, GROUP_GLOBAL)
    except Exception as e:
        logging.warning("official_data_archive is_global_manager: %s", e)
        return False


def user_manage_groups(user_id):
    """
    ユーザが担当できる管理グループ [{id, name}] を返す。
    全体管理者は全管理グループ、そうでなければ所属する管理グループのみ。
    """
    groups = _prefixed_manager_groups()
    if is_global_manager(user_id):
        return groups
    eff = _user_effective_group_ids(user_id)
    return [g for g in groups if g['id'] in eff]


def can_add(user_id):
    """追加権限：全体管理者、またはいずれかの管理グループの所属者。"""
    return is_global_manager(user_id) or bool(user_manage_groups(user_id))


# ────────────────────────────────────────────
# 登録簿ヘルパー
# ────────────────────────────────────────────

def _registry_lookup(table_name, database_name):
    """
    登録簿に (database_name, table_name) のアイテムがあるか照合する。
    あれば登録行（dict）、なければ None。
    データアクセス可否は必ずこの照合で判定する（許可リスト方式）。
    """
    try:
        conn = mysql.connector.connect(**DatabaseConfig.default())
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT id, table_name, database_name, display_name, note
            FROM official_data_archive_tables
            WHERE table_name = %s AND database_name = %s
            LIMIT 1
        """, (table_name, database_name))
        return cursor.fetchone()
    except Exception as e:
        logging.error("official_data_archive _registry_lookup error: %s", e)
        return None
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close(); conn.close()


def _table_exists(table_name, database_name):
    """
    指定DBに実テーブルが存在するか information_schema で確認する。
    追加（登録）時の検査に使う。
    """
    try:
        conn = mysql.connector.connect(**DB_CHOICES[database_name]())
        cursor = conn.cursor()
        cursor.execute("""
            SELECT COUNT(*) FROM information_schema.tables
            WHERE table_schema = DATABASE() AND table_name = %s
        """, (table_name,))
        return cursor.fetchone()[0] > 0
    except Exception as e:
        logging.error("official_data_archive _table_exists error: %s", e)
        return False
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close(); conn.close()


def safe_val(v):
    """JSONで返せる値に変換する（日時3層ルール：バックエンドで文字列化）。"""
    if v is None:
        return None
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.isoformat()
    if isinstance(v, datetime.timedelta):
        s = int(v.total_seconds())
        return f"{s//3600:02d}:{(s%3600)//60:02d}"
    if isinstance(v, bytes):
        return v.decode('utf-8', errors='replace')
    return v


# ────────────────────────────────────────────
# 画面
# ────────────────────────────────────────────

@official_data_archive_bp.route('/')
@login_required
def index():
    """公式データ集メイン画面（2ペインビューア）。"""
    user_id = session.get('user_id')
    if not can_view(user_id):
        return redirect_to_dashboard()
    return render_template('official_data_archive/index.html')


# ────────────────────────────────────────────
# API: 公式テーブル一覧
# ────────────────────────────────────────────

@official_data_archive_bp.route('/api/tables', methods=['GET'])
@login_required
def api_tables():
    """
    登録簿の一覧を返す。各アイテムに行数（row_count）と
    呼び出しユーザの管理可否（can_manage：全体管理者、または
    そのアイテムの管理グループ所属者）を付ける。
    あわせて追加可否（can_add）と、ユーザが指定できる管理グループ
    一覧（manage_groups）も返し、フロントは追加・削除UIの表示判定に使う。
    """
    user_id = session.get('user_id')
    if not can_view(user_id):
        return jsonify({'success': False, 'error': '権限がありません'}), 403
    try:
        conn = mysql.connector.connect(**DatabaseConfig.default())
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT t.id, t.table_name, t.database_name,
                   t.display_name, t.note, t.manager_group_id,
                   g.name AS manager_group_name
            FROM official_data_archive_tables t
            LEFT JOIN user_groups g ON t.manager_group_id = g.id
            ORDER BY COALESCE(t.display_name, t.table_name)
        """)
        tables = cursor.fetchall()
        cursor.close(); conn.close()

        # ユーザの管理範囲（アイテムごとの can_manage 判定用）
        is_global = is_global_manager(user_id)
        my_groups = user_manage_groups(user_id)
        my_group_ids = {g['id'] for g in my_groups}
        for t in tables:
            t['can_manage'] = bool(
                is_global or (t['manager_group_id'] in my_group_ids
                              if t['manager_group_id'] else False))

        # 行数はDBごとにまとめて取得する（接続を最小限に）
        by_db = {}
        for t in tables:
            by_db.setdefault(t['database_name'], []).append(t)
        for db_key, items in by_db.items():
            if db_key not in DB_CHOICES:
                for t in items:
                    t['row_count'] = None
                continue
            try:
                db_conn = mysql.connector.connect(**DB_CHOICES[db_key]())
                db_cur = db_conn.cursor()
                for t in items:
                    try:
                        db_cur.execute(
                            "SELECT COUNT(*) FROM `%s`" % t['table_name'])
                        t['row_count'] = db_cur.fetchone()[0]
                    except Exception as e:
                        logging.warning(
                            "official_data_archive row_count error %s.%s: %s",
                            db_key, t['table_name'], e)
                        t['row_count'] = None
                db_cur.close(); db_conn.close()
            except Exception as e:
                logging.error("official_data_archive db connect (%s): %s",
                              db_key, e)
                for t in items:
                    t['row_count'] = None

        return jsonify({
            'success': True,
            'tables': tables,
            'can_add': can_add(user_id),
            'is_global': is_global,
            'manage_groups': my_groups,
            'db_choices': list(DB_CHOICES.keys()),
        })
    except Exception as e:
        logging.error("api_tables error: %s", e)
        return jsonify({'success': False, 'error': str(e)}), 500


# ────────────────────────────────────────────
# API: アイテムの追加・削除（管理者）
# ────────────────────────────────────────────

@official_data_archive_bp.route('/api/tables/add', methods=['POST'])
@login_required
def api_table_add():
    """
    公式テーブルへアイテム（SQLテーブル）を追加する。
    リクエスト: {
        "table_name":       "T_10_01",    # 必須（英数字とアンダースコア）
        "database_name":    "fujinp",     # 必須（default / fujinp / public）
        "manager_group_id": 12,           # 管理グループ（まいぐる user_groups.id）
                                          #   全体管理者のみ null（＝全体管理者専管）可
        "display_name":     "（任意）表示名",
        "note":             "（任意）説明"
    }
    管理グループは「公式データ集_」で始まる まいぐるグループから選ぶ。
    全体管理者以外は、自分が所属する管理グループしか指定できない。
    実テーブルの存在を確認してから登録する。
    """
    user_id = session.get('user_id')
    if not can_add(user_id):
        return jsonify({'success': False, 'error': '管理権限がありません'}), 403
    try:
        data          = request.json or {}
        table_name    = (data.get('table_name') or '').strip()
        database_name = (data.get('database_name') or '').strip()
        display_name  = (data.get('display_name') or '').strip() or None
        note          = (data.get('note') or '').strip() or None
        manager_group_id = data.get('manager_group_id') or None

        if not TABLE_NAME_RE.match(table_name):
            return jsonify({'success': False,
                            'error': 'テーブル名は英数字とアンダースコア'
                                     '（64文字以内）で指定してください'}), 400
        if database_name not in DB_CHOICES:
            return jsonify({'success': False,
                            'error': 'データベースの指定が不正です'}), 400

        # ── 管理グループの検証 ──
        is_global = is_global_manager(user_id)
        allowed_ids = {g['id'] for g in user_manage_groups(user_id)}
        if manager_group_id is None:
            # 管理グループなし＝全体管理者専管。全体管理者のみ許可。
            if not is_global:
                return jsonify({'success': False,
                                'error': '管理グループを指定してください'}), 400
        else:
            try:
                manager_group_id = int(manager_group_id)
            except (TypeError, ValueError):
                return jsonify({'success': False,
                                'error': '管理グループの指定が不正です'}), 400
            if manager_group_id not in allowed_ids:
                return jsonify({'success': False,
                                'error': '指定できない管理グループです'
                                         '（所属していないか、'
                                         f'「{GROUP_PREFIX}」で始まる'
                                         'グループではありません）'}), 403

        if _registry_lookup(table_name, database_name):
            return jsonify({'success': False,
                            'error': f'{database_name} の {table_name} は'
                                     'すでに登録されています'}), 400
        if not _table_exists(table_name, database_name):
            return jsonify({'success': False,
                            'error': f'{database_name} に {table_name} という'
                                     'テーブルが見つかりません'}), 404

        conn = mysql.connector.connect(**DatabaseConfig.default())
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO official_data_archive_tables
                (table_name, database_name, display_name, note,
                 manager_group_id, created_by, created_at)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """, (table_name, database_name, display_name, note,
              manager_group_id, user_id, get_jst_now()))
        conn.commit()
        new_id = cursor.lastrowid
        return jsonify({'success': True, 'id': new_id})
    except Exception as e:
        logging.error("api_table_add error: %s", e)
        if 'conn' in locals():
            conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close(); conn.close()


@official_data_archive_bp.route('/api/tables/delete', methods=['POST'])
@login_required
def api_table_delete():
    """
    公式テーブルからアイテムを削除する。
    リクエスト: { "id": 登録簿のid }
    削除できるのは全体管理者、またはそのアイテムの管理グループ所属者。
    ※ 登録簿からの除外のみ。SQLテーブル自体は絶対に DROP しない。
    """
    user_id = session.get('user_id')
    try:
        data    = request.json or {}
        item_id = data.get('id')
        if not item_id:
            return jsonify({'success': False, 'error': 'idが必要です'}), 400

        conn = mysql.connector.connect(**DatabaseConfig.default())
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT id, manager_group_id FROM official_data_archive_tables
            WHERE id = %s
        """, (item_id,))
        item = cursor.fetchone()
        if not item:
            return jsonify({'success': False,
                            'error': '対象のアイテムが見つかりません'}), 404

        # ── アイテム単位の権限判定 ──
        if not is_global_manager(user_id):
            my_ids = {g['id'] for g in user_manage_groups(user_id)}
            if not (item['manager_group_id']
                    and item['manager_group_id'] in my_ids):
                return jsonify({'success': False,
                                'error': 'このアイテムの管理権限が'
                                         'ありません'}), 403

        cursor.execute(
            "DELETE FROM official_data_archive_tables WHERE id = %s",
            (item_id,))
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        logging.error("api_table_delete error: %s", e)
        if 'conn' in locals():
            conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close(); conn.close()


# ────────────────────────────────────────────
# API: テーブルデータ（読み取り専用）
# ────────────────────────────────────────────

@official_data_archive_bp.route('/api/data', methods=['GET'])
@login_required
def api_data():
    """
    登録済みテーブルのデータをページ単位で返す。
    パラメータ: table, db, limit（最大5000）, offset
    アクセス可否は登録簿照合のみで判定する（許可リスト方式）。
    """
    user_id = session.get('user_id')
    if not can_view(user_id):
        return jsonify({'success': False, 'error': '権限がありません'}), 403

    table_name    = request.args.get('table', '').strip()
    database_name = request.args.get('db', '').strip()
    limit  = min(int(request.args.get('limit', 500)), PAGE_LIMIT_MAX)
    offset = max(int(request.args.get('offset', 0)), 0)

    if not table_name or database_name not in DB_CHOICES:
        return jsonify({'success': False,
                        'error': 'テーブル名とデータベースが必要です'}), 400
    if not _registry_lookup(table_name, database_name):
        return jsonify({'success': False,
                        'error': 'アクセス不可: ' + table_name}), 403

    try:
        conn = mysql.connector.connect(**DB_CHOICES[database_name]())
        cursor = conn.cursor(dictionary=True)

        cursor.execute("SHOW COLUMNS FROM `%s`" % table_name)
        columns = [c['Field'] for c in cursor.fetchall()]

        cursor.execute("SELECT COUNT(*) AS cnt FROM `%s`" % table_name)
        total_count = cursor.fetchone()['cnt']

        cursor.execute(
            "SELECT * FROM `%s` LIMIT %%s OFFSET %%s" % table_name,
            (limit, offset))
        rows = cursor.fetchall()

        serialized = [{k: safe_val(v) for k, v in row.items()}
                      for row in rows]

        return jsonify({
            'success'      : True,
            'table'        : table_name,
            'columns'      : columns,
            'rows'         : serialized,
            'total_count'  : total_count,
            'fetched_count': len(rows),
            'offset'       : offset,
            'limit'        : limit,
        })
    except Exception as e:
        logging.error("api_data error: %s", e)
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close(); conn.close()


# ────────────────────────────────────────────
# API: XLSXダウンロード
# ────────────────────────────────────────────

@official_data_archive_bp.route('/api/download', methods=['GET'])
@login_required
def api_download():
    """登録済みテーブルを全件XLSXでダウンロードする。"""
    import io
    import openpyxl
    from openpyxl.styles import PatternFill, Font
    from flask import send_file

    user_id = session.get('user_id')
    if not can_view(user_id):
        return jsonify({'success': False, 'error': '権限がありません'}), 403

    table_name    = request.args.get('table', '').strip()
    database_name = request.args.get('db', '').strip()
    if not table_name or database_name not in DB_CHOICES:
        return jsonify({'success': False,
                        'error': 'テーブル名とデータベースが必要です'}), 400
    if not _registry_lookup(table_name, database_name):
        return jsonify({'success': False, 'error': 'アクセス不可'}), 403

    try:
        conn = mysql.connector.connect(**DB_CHOICES[database_name]())
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT * FROM `%s`" % table_name)
        rows = cursor.fetchall()
        columns = ([d[0] for d in cursor.description]
                   if cursor.description else [])
        cursor.close(); conn.close()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = table_name[:31]  # シート名は31文字まで

        ws.append(columns)
        for row in rows:
            ws.append([
                (v.isoformat() if isinstance(v, (datetime.datetime,
                                                 datetime.date))
                 else (f"{int(v.total_seconds())//3600:02d}:"
                       f"{(int(v.total_seconds())%3600)//60:02d}"
                       if isinstance(v, datetime.timedelta)
                       else (v.decode('utf-8', errors='replace')
                             if isinstance(v, bytes) else v)))
                for v in [row[c] for c in columns]
            ])

        header_fill = PatternFill('solid', fgColor='1A3A5C')
        header_font = Font(color='FFFFFF', bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(
            buf,
            mimetype='application/vnd.openxmlformats-officedocument'
                     '.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f"{table_name}.xlsx")
    except Exception as e:
        logging.error("api_download error: %s", e)
        return jsonify({'success': False, 'error': str(e)}), 500


# ────────────────────────────────────────────
# 指標ビュー
#   保存されたSELECT文による動的ビュー。
#   1行＝（指標名, SQLクエリ, 表示順位数）のタプル。
#   クエリは fujinp DB（公式テーブルの所在）に対する
#   単文のSELECTのみ実行を許可する（読み取り専用の建付けを守る）。
# ────────────────────────────────────────────

INDICATOR_DB        = 'fujinp'   # クエリ実行先（DB_CHOICES のキー）
INDICATOR_MAX_ROWS  = 1000       # 実行結果の最大行数
INDICATOR_PREVIEW_ROWS = 200     # 試し実行の最大行数


def _validate_select(query):
    """
    指標ビューのSQLクエリを検査する。
    許可：単文のSELECTのみ。
    戻り値 (正規化済みクエリ, None) または (None, エラーメッセージ)。
    """
    q = (query or '').strip()
    # 末尾のセミコロンは1つだけ許して落とす
    if q.endswith(';'):
        q = q[:-1].rstrip()
    if not q:
        return None, 'SQLクエリが空です'
    if ';' in q:
        return None, 'セミコロンを含む複文は実行できません（単文のSELECTのみ）'
    if not re.match(r'^select\b', q, flags=re.IGNORECASE):
        return None, 'SELECT文のみ実行できます'
    low = q.lower()
    for bad in ('into outfile', 'into dumpfile'):
        if bad in low:
            return None, f'許可されない構文が含まれています：{bad}'
    return q, None


def _run_indicator_query(query, max_rows):
    """
    検査済みクエリを INDICATOR_DB で実行し、
    (columns, rows(serialized), truncated) を返す。例外は呼び出し側で処理。
    """
    conn = mysql.connector.connect(**DB_CHOICES[INDICATOR_DB]())
    try:
        cursor = conn.cursor(dictionary=True)
        cursor.execute(query)
        columns = ([d[0] for d in cursor.description]
                   if cursor.description else [])
        rows = cursor.fetchmany(max_rows + 1)
        truncated = len(rows) > max_rows
        rows = rows[:max_rows]
        serialized = [{k: safe_val(v) for k, v in row.items()}
                      for row in rows]
        return columns, serialized, truncated
    finally:
        if conn.is_connected():
            cursor.close(); conn.close()


def _get_indicator(iv_id):
    """指標ビュー1件を取得（管理グループ名つき）。なければ None。"""
    try:
        conn = mysql.connector.connect(**DatabaseConfig.default())
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT v.id, v.name, v.query, v.sort_order, v.chart_type,
                   v.manager_group_id, g.name AS manager_group_name
            FROM official_data_archive_indicator_views v
            LEFT JOIN user_groups g ON v.manager_group_id = g.id
            WHERE v.id = %s
        """, (iv_id,))
        return cursor.fetchone()
    except Exception as e:
        logging.error("official_data_archive _get_indicator: %s", e)
        return None
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close(); conn.close()


def _can_manage_item(user_id, manager_group_id):
    """アイテム単位の管理可否（公式テーブル・指標ビュー共通の判定）。"""
    if is_global_manager(user_id):
        return True
    if not manager_group_id:
        return False
    my_ids = {g['id'] for g in user_manage_groups(user_id)}
    return manager_group_id in my_ids


@official_data_archive_bp.route('/api/indicators', methods=['GET'])
@login_required
def api_indicators():
    """
    指標ビュー一覧を表示順位（sort_order 昇順、同順位は名前順）で返す。
    各アイテムに can_manage を付け、追加可否（can_add）と
    指定できる管理グループ一覧（manage_groups）も返す。
    """
    user_id = session.get('user_id')
    if not can_view(user_id):
        return jsonify({'success': False, 'error': '権限がありません'}), 403
    try:
        conn = mysql.connector.connect(**DatabaseConfig.default())
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT v.id, v.name, v.query, v.sort_order, v.chart_type,
                   v.manager_group_id, g.name AS manager_group_name
            FROM official_data_archive_indicator_views v
            LEFT JOIN user_groups g ON v.manager_group_id = g.id
            ORDER BY v.sort_order, v.name
        """)
        views = cursor.fetchall()
        cursor.close(); conn.close()

        is_global = is_global_manager(user_id)
        my_groups = user_manage_groups(user_id)
        my_group_ids = {g['id'] for g in my_groups}
        for v in views:
            v['can_manage'] = bool(
                is_global or (v['manager_group_id'] in my_group_ids
                              if v['manager_group_id'] else False))
            # DECIMAL は Decimal で返るため JSON 化できるよう float に正規化
            if v.get('sort_order') is not None:
                v['sort_order'] = float(v['sort_order'])

        return jsonify({
            'success': True,
            'views': views,
            'can_add': can_add(user_id),
            'is_global': is_global,
            'manage_groups': my_groups,
        })
    except Exception as e:
        logging.error("api_indicators error: %s", e)
        return jsonify({'success': False, 'error': str(e)}), 500


@official_data_archive_bp.route('/api/indicators/run', methods=['GET'])
@login_required
def api_indicator_run():
    """
    登録済み指標ビューのクエリを実行し、結果を返す。
    パラメータ: id
    """
    user_id = session.get('user_id')
    if not can_view(user_id):
        return jsonify({'success': False, 'error': '権限がありません'}), 403
    iv = _get_indicator(request.args.get('id', type=int))
    if not iv:
        return jsonify({'success': False,
                        'error': '指標ビューが見つかりません'}), 404
    query, err = _validate_select(iv['query'])
    if err:
        return jsonify({'success': False, 'error': err}), 400
    try:
        columns, rows, truncated = _run_indicator_query(
            query, INDICATOR_MAX_ROWS)
        return jsonify({
            'success': True,
            'id': iv['id'],
            'name': iv['name'],
            'query': iv['query'],
            'chart_type': iv.get('chart_type') or 'bar',
            'columns': columns,
            'rows': rows,
            'truncated': truncated,
        })
    except Exception as e:
        logging.error("api_indicator_run error (id=%s): %s",
                      iv.get('id'), e)
        return jsonify({'success': False,
                        'error': f'クエリ実行エラー：{e}'}), 400


@official_data_archive_bp.route('/api/indicators/preview', methods=['POST'])
@login_required
def api_indicator_preview():
    """
    試し実行（保存前のプレビュー）。管理者のみ。
    リクエスト: { "query": "SELECT ..." }
    """
    user_id = session.get('user_id')
    if not can_add(user_id):
        return jsonify({'success': False, 'error': '管理権限がありません'}), 403
    data = request.json or {}
    query, err = _validate_select(data.get('query'))
    if err:
        return jsonify({'success': False, 'error': err}), 400
    try:
        columns, rows, truncated = _run_indicator_query(
            query, INDICATOR_PREVIEW_ROWS)
        return jsonify({'success': True, 'columns': columns,
                        'rows': rows, 'truncated': truncated})
    except Exception as e:
        return jsonify({'success': False,
                        'error': f'クエリ実行エラー：{e}'}), 400


@official_data_archive_bp.route('/api/indicators/save', methods=['POST'])
@login_required
def api_indicator_save():
    """
    指標ビューの追加・編集。
    リクエスト: {
        "id":               null=新規 / 既存id=編集,
        "name":             "指標名",          # 必須
        "query":            "SELECT ...",      # 必須（単文SELECT）
        "sort_order":       10.5,              # 表示順位数（小さいほど上位）
                                               #   8桁整数部＋3桁小数部
                                               #   範囲 0〜99999999.999、小数第3位まで
        "chart_type":       "bar" / "line",    # グラフ種別（省略時 bar）
        "manager_group_id": 12                 # 公式テーブルと同じ規約
    }
    """
    user_id = session.get('user_id')
    if not can_add(user_id):
        return jsonify({'success': False, 'error': '管理権限がありません'}), 403
    try:
        data  = request.json or {}
        iv_id = data.get('id') or None
        name  = (data.get('name') or '').strip()
        manager_group_id = data.get('manager_group_id') or None
        # 表示順位数：8桁整数部＋3桁小数部の小数。
        # DECIMAL(11,3) に格納するため、桁あふれ・小数桁超過を検査する。
        try:
            sort_order = float(data.get('sort_order', 100))
        except (TypeError, ValueError):
            return jsonify({'success': False,
                            'error': '表示順位数は数値で指定してください'}), 400
        if not math.isfinite(sort_order):
            return jsonify({'success': False,
                            'error': '表示順位数は数値で指定してください'}), 400
        # 小数第3位までに丸めてから範囲・桁を検査
        sort_order = round(sort_order, 3)
        if not (0 <= sort_order <= 99999999.999):
            return jsonify({'success': False,
                            'error': '表示順位数は0〜99999999.999の範囲で'
                                     '指定してください'}), 400

        chart_type = (data.get('chart_type') or 'bar').strip().lower()
        if chart_type not in ('bar', 'line'):
            chart_type = 'bar'

        if not name:
            return jsonify({'success': False,
                            'error': '指標名を入力してください'}), 400
        query, err = _validate_select(data.get('query'))
        if err:
            return jsonify({'success': False, 'error': err}), 400

        # ── 管理グループの検証（公式テーブルの追加と同じ規則） ──
        is_global = is_global_manager(user_id)
        allowed_ids = {g['id'] for g in user_manage_groups(user_id)}
        if manager_group_id is None:
            if not is_global:
                return jsonify({'success': False,
                                'error': '管理グループを指定してください'}), 400
        else:
            try:
                manager_group_id = int(manager_group_id)
            except (TypeError, ValueError):
                return jsonify({'success': False,
                                'error': '管理グループの指定が不正です'}), 400
            if manager_group_id not in allowed_ids:
                return jsonify({'success': False,
                                'error': '指定できない管理グループです'}), 403

        now  = get_jst_now()
        conn = mysql.connector.connect(**DatabaseConfig.default())
        cursor = conn.cursor(dictionary=True)

        if iv_id:
            # ── 編集：対象アイテムの管理権限を確認 ──
            cursor.execute("""
                SELECT manager_group_id
                FROM official_data_archive_indicator_views WHERE id = %s
            """, (iv_id,))
            row = cursor.fetchone()
            if not row:
                return jsonify({'success': False,
                                'error': '指標ビューが見つかりません'}), 404
            if not _can_manage_item(user_id, row['manager_group_id']):
                return jsonify({'success': False,
                                'error': 'この指標ビューの管理権限が'
                                         'ありません'}), 403
            cursor.execute("""
                UPDATE official_data_archive_indicator_views
                SET name=%s, query=%s, sort_order=%s, chart_type=%s,
                    manager_group_id=%s, updated_by=%s, updated_at=%s
                WHERE id=%s
            """, (name, query, sort_order, chart_type, manager_group_id,
                  user_id, now, iv_id))
            conn.commit()
            return jsonify({'success': True, 'id': iv_id})
        else:
            cursor.execute("""
                INSERT INTO official_data_archive_indicator_views
                    (name, query, sort_order, chart_type, manager_group_id,
                     created_by, created_at)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            """, (name, query, sort_order, chart_type, manager_group_id,
                  user_id, now))
            conn.commit()
            return jsonify({'success': True, 'id': cursor.lastrowid})
    except Exception as e:
        logging.error("api_indicator_save error: %s", e)
        if 'conn' in locals():
            conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close(); conn.close()


@official_data_archive_bp.route('/api/indicators/chart_type', methods=['POST'])
@login_required
def api_indicator_chart_type():
    """
    グラフ種別（棒／折れ線）だけを更新する軽量エンドポイント。
    リクエスト: { "id": 指標ビューのid, "chart_type": "bar" / "line" }
    管理権限のある利用者の選択のみ保存する（権限がなければ 403）。
    """
    user_id = session.get('user_id')
    if not can_view(user_id):
        return jsonify({'success': False, 'error': '権限がありません'}), 403
    data = request.json or {}
    iv = _get_indicator(data.get('id'))
    if not iv:
        return jsonify({'success': False,
                        'error': '指標ビューが見つかりません'}), 404
    if not _can_manage_item(user_id, iv.get('manager_group_id')):
        return jsonify({'success': False,
                        'error': 'この指標ビューの管理権限がありません'}), 403
    chart_type = (data.get('chart_type') or '').strip().lower()
    if chart_type not in ('bar', 'line'):
        return jsonify({'success': False,
                        'error': 'グラフ種別の指定が不正です'}), 400
    try:
        conn = mysql.connector.connect(**DatabaseConfig.default())
        cursor = conn.cursor()
        cursor.execute("""
            UPDATE official_data_archive_indicator_views
            SET chart_type=%s, updated_by=%s, updated_at=%s
            WHERE id=%s
        """, (chart_type, user_id, get_jst_now(), iv['id']))
        conn.commit()
        return jsonify({'success': True, 'id': iv['id'],
                        'chart_type': chart_type})
    except Exception as e:
        logging.error("api_indicator_chart_type error: %s", e)
        if 'conn' in locals():
            conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close(); conn.close()


@official_data_archive_bp.route('/api/indicators/delete', methods=['POST'])
@login_required
def api_indicator_delete():
    """
    指標ビューの削除。全体管理者またはアイテムの管理グループ所属者のみ。
    リクエスト: { "id": 指標ビューのid }
    """
    user_id = session.get('user_id')
    try:
        data  = request.json or {}
        iv_id = data.get('id')
        if not iv_id:
            return jsonify({'success': False, 'error': 'idが必要です'}), 400

        conn = mysql.connector.connect(**DatabaseConfig.default())
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT manager_group_id
            FROM official_data_archive_indicator_views WHERE id = %s
        """, (iv_id,))
        row = cursor.fetchone()
        if not row:
            return jsonify({'success': False,
                            'error': '指標ビューが見つかりません'}), 404
        if not _can_manage_item(user_id, row['manager_group_id']):
            return jsonify({'success': False,
                            'error': 'この指標ビューの管理権限がありません'}), 403
        cursor.execute(
            "DELETE FROM official_data_archive_indicator_views WHERE id=%s",
            (iv_id,))
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        logging.error("api_indicator_delete error: %s", e)
        if 'conn' in locals():
            conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close(); conn.close()


@official_data_archive_bp.route('/api/indicators/download', methods=['GET'])
@login_required
def api_indicator_download():
    """指標ビューの実行結果をXLSXでダウンロードする。パラメータ: id"""
    import io
    import openpyxl
    from openpyxl.styles import PatternFill, Font
    from flask import send_file

    user_id = session.get('user_id')
    if not can_view(user_id):
        return jsonify({'success': False, 'error': '権限がありません'}), 403
    iv = _get_indicator(request.args.get('id', type=int))
    if not iv:
        return jsonify({'success': False,
                        'error': '指標ビューが見つかりません'}), 404
    query, err = _validate_select(iv['query'])
    if err:
        return jsonify({'success': False, 'error': err}), 400
    try:
        columns, rows, truncated = _run_indicator_query(
            query, INDICATOR_MAX_ROWS)

        wb = openpyxl.Workbook()
        ws = wb.active
        # シート名に使えない文字を除き31文字に収める
        sheet = re.sub(r'[\[\]:*?/\\]', '_', iv['name'])[:31] or 'indicator'
        ws.title = sheet

        ws.append(columns)
        for row in rows:
            ws.append([row.get(c) for c in columns])

        header_fill = PatternFill('solid', fgColor='1A3A5C')
        header_font = Font(color='FFFFFF', bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        fname = re.sub(r'[\\/:*?"<>|]', '_', iv['name']) or 'indicator'
        return send_file(
            buf,
            mimetype='application/vnd.openxmlformats-officedocument'
                     '.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f"{fname}.xlsx")
    except Exception as e:
        logging.error("api_indicator_download error (id=%s): %s",
                      iv.get('id'), e)
        return jsonify({'success': False,
                        'error': f'クエリ実行エラー：{e}'}), 400


# ────────────────────────────────────────────
# 複合指標
#   1つ以上（高々 COMPOSITE_MAX 個）の単一指標と色のペアから構成。
#   表示は折れ線固定・縦軸0〜全系列の最大値（フロント側）。
#   ここでは構成の管理と、各構成指標のクエリの一括実行を提供する。
# ────────────────────────────────────────────

COMPOSITE_MAX = 5
COLOR_RE = re.compile(r'^#[0-9A-Fa-f]{6}$')


@official_data_archive_bp.route('/api/composites', methods=['GET'])
@login_required
def api_composites():
    """
    複合指標一覧を表示順位（sort_order 昇順、同順位は名前順）で返す。
    各アイテムに構成（単一指標と色のペア、seq順）と can_manage を付ける。
    モーダル用に単一指標の選択肢（indicator_options）も返す。
    """
    user_id = session.get('user_id')
    if not can_view(user_id):
        return jsonify({'success': False, 'error': '権限がありません'}), 403
    try:
        conn = mysql.connector.connect(**DatabaseConfig.default())
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT c.id, c.name, c.sort_order,
                   c.manager_group_id, g.name AS manager_group_name
            FROM official_data_archive_composites c
            LEFT JOIN user_groups g ON c.manager_group_id = g.id
            ORDER BY c.sort_order, c.name
        """)
        composites = cursor.fetchall()

        cursor.execute("""
            SELECT cc.composite_id, cc.indicator_view_id, cc.color, cc.seq,
                   v.name AS indicator_name
            FROM official_data_archive_composite_components cc
            LEFT JOIN official_data_archive_indicator_views v
                   ON cc.indicator_view_id = v.id
            ORDER BY cc.composite_id, cc.seq
        """)
        comp_rows = cursor.fetchall()

        cursor.execute("""
            SELECT id, name FROM official_data_archive_indicator_views
            ORDER BY sort_order, name
        """)
        indicator_options = cursor.fetchall()
        cursor.close(); conn.close()

        by_comp = {}
        for r in comp_rows:
            by_comp.setdefault(r['composite_id'], []).append({
                'indicator_view_id': r['indicator_view_id'],
                'name':  r['indicator_name'] or '（削除された単一指標）',
                'color': r['color'],
                'seq':   r['seq'],
            })

        is_global = is_global_manager(user_id)
        my_groups = user_manage_groups(user_id)
        my_group_ids = {g['id'] for g in my_groups}
        for c in composites:
            c['components'] = by_comp.get(c['id'], [])
            c['can_manage'] = bool(
                is_global or (c['manager_group_id'] in my_group_ids
                              if c['manager_group_id'] else False))

        return jsonify({
            'success': True,
            'composites': composites,
            'indicator_options': indicator_options,
            'can_add': can_add(user_id),
            'is_global': is_global,
            'manage_groups': my_groups,
            'max_components': COMPOSITE_MAX,
        })
    except Exception as e:
        logging.error("api_composites error: %s", e)
        return jsonify({'success': False, 'error': str(e)}), 500


@official_data_archive_bp.route('/api/composites/run', methods=['GET'])
@login_required
def api_composite_run():
    """
    複合指標の全構成クエリを実行し、系列の配列を返す。
    一部の構成が失敗しても他は返す（失敗分は error を持つ）。
    パラメータ: id
    """
    user_id = session.get('user_id')
    if not can_view(user_id):
        return jsonify({'success': False, 'error': '権限がありません'}), 403
    comp_id = request.args.get('id', type=int)
    try:
        conn = mysql.connector.connect(**DatabaseConfig.default())
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT id, name FROM official_data_archive_composites
            WHERE id = %s
        """, (comp_id,))
        comp = cursor.fetchone()
        if not comp:
            return jsonify({'success': False,
                            'error': '複合指標が見つかりません'}), 404
        cursor.execute("""
            SELECT cc.color, cc.seq, v.id AS iv_id, v.name, v.query
            FROM official_data_archive_composite_components cc
            LEFT JOIN official_data_archive_indicator_views v
                   ON cc.indicator_view_id = v.id
            WHERE cc.composite_id = %s
            ORDER BY cc.seq
        """, (comp_id,))
        comps = cursor.fetchall()
        cursor.close(); conn.close()

        series = []
        for c in comps:
            entry = {'name': c['name'] or '（削除された単一指標）',
                     'color': c['color']}
            if not c['iv_id']:
                entry['error'] = '構成する単一指標が削除されています'
                series.append(entry); continue
            query, err = _validate_select(c['query'])
            if err:
                entry['error'] = err
                series.append(entry); continue
            try:
                columns, rows, truncated = _run_indicator_query(
                    query, INDICATOR_MAX_ROWS)
                entry['columns'] = columns
                entry['rows'] = rows
                entry['truncated'] = truncated
            except Exception as qe:
                entry['error'] = f'クエリ実行エラー：{qe}'
            series.append(entry)

        return jsonify({'success': True, 'id': comp['id'],
                        'name': comp['name'], 'series': series})
    except Exception as e:
        logging.error("api_composite_run error (id=%s): %s", comp_id, e)
        return jsonify({'success': False, 'error': str(e)}), 500


@official_data_archive_bp.route('/api/composites/save', methods=['POST'])
@login_required
def api_composite_save():
    """
    複合指標の追加・編集。
    リクエスト: {
        "id":               null=新規 / 既存id=編集,
        "name":             "複合指標名",     # 必須
        "sort_order":       10,
        "manager_group_id": 12,
        "components": [                       # 1〜COMPOSITE_MAX 個
            {"indicator_view_id": 3, "color": "#2e6da4"}, ...
        ]
    }
    構成は全削除→再挿入で更新する。
    """
    user_id = session.get('user_id')
    if not can_add(user_id):
        return jsonify({'success': False, 'error': '管理権限がありません'}), 403
    try:
        data    = request.json or {}
        comp_id = data.get('id') or None
        name    = (data.get('name') or '').strip()
        manager_group_id = data.get('manager_group_id') or None
        components = data.get('components') or []
        try:
            sort_order = int(data.get('sort_order', 100))
        except (TypeError, ValueError):
            return jsonify({'success': False,
                            'error': '表示順位数は整数で指定してください'}), 400

        if not name:
            return jsonify({'success': False,
                            'error': '複合指標名を入力してください'}), 400
        if not (1 <= len(components) <= COMPOSITE_MAX):
            return jsonify({'success': False,
                            'error': f'構成は1〜{COMPOSITE_MAX}個で'
                                     '指定してください'}), 400

        # ── 管理グループの検証（既存と同じ規則） ──
        is_global = is_global_manager(user_id)
        allowed_ids = {g['id'] for g in user_manage_groups(user_id)}
        if manager_group_id is None:
            if not is_global:
                return jsonify({'success': False,
                                'error': '管理グループを指定してください'}), 400
        else:
            try:
                manager_group_id = int(manager_group_id)
            except (TypeError, ValueError):
                return jsonify({'success': False,
                                'error': '管理グループの指定が不正です'}), 400
            if manager_group_id not in allowed_ids:
                return jsonify({'success': False,
                                'error': '指定できない管理グループです'}), 403

        conn = mysql.connector.connect(**DatabaseConfig.default())
        cursor = conn.cursor(dictionary=True)

        # ── 構成の検証（単一指標の存在・色の形式） ──
        cursor.execute(
            "SELECT id FROM official_data_archive_indicator_views")
        valid_iv_ids = {r['id'] for r in cursor.fetchall()}
        cleaned = []
        for i, comp in enumerate(components, 1):
            try:
                iv_id = int(comp.get('indicator_view_id'))
            except (TypeError, ValueError):
                return jsonify({'success': False,
                                'error': f'構成{i}：単一指標を選んで'
                                         'ください'}), 400
            if iv_id not in valid_iv_ids:
                return jsonify({'success': False,
                                'error': f'構成{i}：指定された単一指標が'
                                         '存在しません'}), 400
            color = (comp.get('color') or '').strip()
            if not COLOR_RE.match(color):
                return jsonify({'success': False,
                                'error': f'構成{i}：色は #RRGGBB 形式で'
                                         '指定してください'}), 400
            cleaned.append((iv_id, color))

        now = get_jst_now()
        if comp_id:
            # ── 編集：対象の管理権限を確認 ──
            cursor.execute("""
                SELECT manager_group_id
                FROM official_data_archive_composites WHERE id = %s
            """, (comp_id,))
            row = cursor.fetchone()
            if not row:
                return jsonify({'success': False,
                                'error': '複合指標が見つかりません'}), 404
            if not _can_manage_item(user_id, row['manager_group_id']):
                return jsonify({'success': False,
                                'error': 'この複合指標の管理権限が'
                                         'ありません'}), 403
            cursor.execute("""
                UPDATE official_data_archive_composites
                SET name=%s, sort_order=%s, manager_group_id=%s,
                    updated_by=%s, updated_at=%s
                WHERE id=%s
            """, (name, sort_order, manager_group_id, user_id, now, comp_id))
            cursor.execute("""
                DELETE FROM official_data_archive_composite_components
                WHERE composite_id = %s
            """, (comp_id,))
        else:
            cursor.execute("""
                INSERT INTO official_data_archive_composites
                    (name, sort_order, manager_group_id,
                     created_by, created_at)
                VALUES (%s, %s, %s, %s, %s)
            """, (name, sort_order, manager_group_id, user_id, now))
            comp_id = cursor.lastrowid

        for seq, (iv_id, color) in enumerate(cleaned, 1):
            cursor.execute("""
                INSERT INTO official_data_archive_composite_components
                    (composite_id, indicator_view_id, color, seq)
                VALUES (%s, %s, %s, %s)
            """, (comp_id, iv_id, color, seq))

        conn.commit()
        return jsonify({'success': True, 'id': comp_id})
    except Exception as e:
        logging.error("api_composite_save error: %s", e)
        if 'conn' in locals():
            conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close(); conn.close()


@official_data_archive_bp.route('/api/composites/delete', methods=['POST'])
@login_required
def api_composite_delete():
    """
    複合指標の削除（構成も併せて削除）。
    全体管理者またはアイテムの管理グループ所属者のみ。
    リクエスト: { "id": 複合指標のid }
    ※ 構成する単一指標そのものには影響しない。
    """
    user_id = session.get('user_id')
    try:
        data    = request.json or {}
        comp_id = data.get('id')
        if not comp_id:
            return jsonify({'success': False, 'error': 'idが必要です'}), 400

        conn = mysql.connector.connect(**DatabaseConfig.default())
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT manager_group_id
            FROM official_data_archive_composites WHERE id = %s
        """, (comp_id,))
        row = cursor.fetchone()
        if not row:
            return jsonify({'success': False,
                            'error': '複合指標が見つかりません'}), 404
        if not _can_manage_item(user_id, row['manager_group_id']):
            return jsonify({'success': False,
                            'error': 'この複合指標の管理権限がありません'}), 403
        cursor.execute("""
            DELETE FROM official_data_archive_composite_components
            WHERE composite_id = %s
        """, (comp_id,))
        cursor.execute(
            "DELETE FROM official_data_archive_composites WHERE id = %s",
            (comp_id,))
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        logging.error("api_composite_delete error: %s", e)
        if 'conn' in locals():
            conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close(); conn.close()


# ────────────────────────────────────────────
# 各種データ一括ダウンロード
#   登録簿に載っている公式テーブル・指標の定義情報を、まとめて
#   テキスト（表示・クリップボード用）またはファイルで取り出す。
#   いずれも閲覧権限（regular以上）で利用できる読み取り専用機能。
#   - テーブル名一括：登録簿のテーブル名一覧
#   - スキーマ一括：各テーブルの SHOW CREATE TABLE
#   - 単一指標一括：指標ビューの名称・SQL等
#   - 複合指標一括：複合指標の構成（単一指標と色）
# ────────────────────────────────────────────

def _registered_tables():
    """登録簿の全アイテムを表示名順で返す（dictのlist）。"""
    conn = mysql.connector.connect(**DatabaseConfig.default())
    try:
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT t.id, t.table_name, t.database_name,
                   t.display_name, t.note, g.name AS manager_group_name
            FROM official_data_archive_tables t
            LEFT JOIN user_groups g ON t.manager_group_id = g.id
            ORDER BY COALESCE(t.display_name, t.table_name)
        """)
        return cursor.fetchall()
    finally:
        if conn.is_connected():
            cursor.close(); conn.close()


def _bulk_tablenames_text():
    """テーブル名一括：DBごとに表示名つきのテーブル名一覧を組み立てる。"""
    rows = _registered_tables()
    lines = ['# 公式テーブル名一覧（{0} 件）'.format(len(rows)),
             '# 生成日時: {0} JST'.format(
                 get_jst_now().strftime('%Y-%m-%d %H:%M:%S')),
             '']
    by_db = {}
    for r in rows:
        by_db.setdefault(r['database_name'], []).append(r)
    for db_key in sorted(by_db.keys()):
        lines.append('## [{0}]'.format(db_key))
        for r in by_db[db_key]:
            disp = r['display_name'] or ''
            note = ('  — ' + disp) if disp else ''
            lines.append('{0}{1}'.format(r['table_name'], note))
        lines.append('')
    return '\n'.join(lines).rstrip() + '\n'


def _bulk_schema_text():
    """スキーマ一括：各テーブルの SHOW CREATE TABLE をDBごとにまとめる。"""
    rows = _registered_tables()
    out = ['-- 公式テーブル スキーマ一括（{0} 件）'.format(len(rows)),
           '-- 生成日時: {0} JST'.format(
               get_jst_now().strftime('%Y-%m-%d %H:%M:%S')),
           '']
    by_db = {}
    for r in rows:
        by_db.setdefault(r['database_name'], []).append(r)
    for db_key in sorted(by_db.keys()):
        out.append('-- ========================================')
        out.append('-- データベース: {0}'.format(db_key))
        out.append('-- ========================================')
        if db_key not in DB_CHOICES:
            out.append('-- （未対応のデータベース指定のため取得できません）')
            out.append('')
            continue
        try:
            conn = mysql.connector.connect(**DB_CHOICES[db_key]())
            cur = conn.cursor()
            for r in by_db[db_key]:
                tbl = r['table_name']
                disp = r['display_name'] or ''
                out.append('-- ----------------------------------------')
                out.append('-- {0}{1}'.format(
                    tbl, ('  ' + disp) if disp else ''))
                out.append('-- ----------------------------------------')
                try:
                    cur.execute('SHOW CREATE TABLE `%s`' % tbl)
                    row = cur.fetchone()
                    ddl = row[1] if row and len(row) > 1 else None
                    out.append((ddl or '-- （取得できませんでした）') + ';')
                except Exception as te:
                    out.append('-- 取得エラー: {0}'.format(te))
                out.append('')
            cur.close(); conn.close()
        except Exception as ce:
            out.append('-- データベース接続エラー: {0}'.format(ce))
            out.append('')
    return '\n'.join(out).rstrip() + '\n'


def _bulk_indicators_text():
    """単一指標一括：指標ビューの定義（名称・順位・グラフ種別・SQL）。"""
    conn = mysql.connector.connect(**DatabaseConfig.default())
    try:
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT v.id, v.name, v.query, v.sort_order, v.chart_type,
                   g.name AS manager_group_name
            FROM official_data_archive_indicator_views v
            LEFT JOIN user_groups g ON v.manager_group_id = g.id
            ORDER BY v.sort_order, v.name
        """)
        views = cursor.fetchall()
    finally:
        if conn.is_connected():
            cursor.close(); conn.close()

    out = ['-- 単一指標一括（{0} 件）'.format(len(views)),
           '-- 生成日時: {0} JST'.format(
               get_jst_now().strftime('%Y-%m-%d %H:%M:%S')),
           '']
    for v in views:
        so = v['sort_order']
        so = float(so) if so is not None else None
        out.append('-- ----------------------------------------')
        out.append('-- [{0}] {1}'.format(v['id'], v['name']))
        out.append('--   表示順位: {0} / グラフ: {1} / 管理: {2}'.format(
            so, v.get('chart_type') or 'bar',
            v.get('manager_group_name') or '（全体管理者専管）'))
        out.append('-- ----------------------------------------')
        q = (v['query'] or '').strip()
        out.append(q + (';' if not q.endswith(';') else ''))
        out.append('')
    return '\n'.join(out).rstrip() + '\n'


def _bulk_composites_text():
    """複合指標一括：複合指標と、その構成（単一指標と色）の一覧。"""
    conn = mysql.connector.connect(**DatabaseConfig.default())
    try:
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT c.id, c.name, c.sort_order, g.name AS manager_group_name
            FROM official_data_archive_composites c
            LEFT JOIN user_groups g ON c.manager_group_id = g.id
            ORDER BY c.sort_order, c.name
        """)
        composites = cursor.fetchall()
        cursor.execute("""
            SELECT cc.composite_id, cc.color, cc.seq,
                   v.id AS iv_id, v.name AS indicator_name
            FROM official_data_archive_composite_components cc
            LEFT JOIN official_data_archive_indicator_views v
                   ON cc.indicator_view_id = v.id
            ORDER BY cc.composite_id, cc.seq
        """)
        comp_rows = cursor.fetchall()
    finally:
        if conn.is_connected():
            cursor.close(); conn.close()

    by_comp = {}
    for r in comp_rows:
        by_comp.setdefault(r['composite_id'], []).append(r)

    out = ['# 複合指標一括（{0} 件）'.format(len(composites)),
           '# 生成日時: {0} JST'.format(
               get_jst_now().strftime('%Y-%m-%d %H:%M:%S')),
           '']
    for c in composites:
        so = c['sort_order']
        so = float(so) if so is not None else None
        out.append('## [{0}] {1}'.format(c['id'], c['name']))
        out.append('   表示順位: {0} / 管理: {1}'.format(
            so, c.get('manager_group_name') or '（全体管理者専管）'))
        comps = by_comp.get(c['id'], [])
        if not comps:
            out.append('   （構成なし）')
        for cc in comps:
            out.append('   {0}. {1}  [{2}]  (単一指標id={3})'.format(
                cc['seq'],
                cc['indicator_name'] or '（削除された単一指標）',
                cc['color'], cc['iv_id']))
        out.append('')
    return '\n'.join(out).rstrip() + '\n'


# 種別キー → (生成関数, 表示名, ダウンロードファイル名)
BULK_KINDS = {
    'tablenames':  (_bulk_tablenames_text, 'テーブル名一括',
                    'official_tablenames.txt'),
    'schema':      (_bulk_schema_text, 'スキーマ一括',
                    'official_schema.sql'),
    'indicators':  (_bulk_indicators_text, '単一指標一括',
                    'official_indicators.sql'),
    'composites':  (_bulk_composites_text, '複合指標一括',
                    'official_composites.txt'),
}


@official_data_archive_bp.route('/api/bulk', methods=['GET'])
@login_required
def api_bulk():
    """
    各種データ一括の本文を返す（表示・クリップボード用）。
    パラメータ: kind（tablenames / schema / indicators / composites）
    レスポンス: { success, kind, label, filename, text }
    """
    user_id = session.get('user_id')
    if not can_view(user_id):
        return jsonify({'success': False, 'error': '権限がありません'}), 403
    kind = (request.args.get('kind') or '').strip()
    if kind not in BULK_KINDS:
        return jsonify({'success': False,
                        'error': '種別の指定が不正です'}), 400
    fn, label, filename = BULK_KINDS[kind]
    try:
        text = fn()
        return jsonify({'success': True, 'kind': kind, 'label': label,
                        'filename': filename, 'text': text})
    except Exception as e:
        logging.error("api_bulk error (kind=%s): %s", kind, e)
        return jsonify({'success': False,
                        'error': f'生成に失敗しました：{e}'}), 500


@official_data_archive_bp.route('/api/bulk/download', methods=['GET'])
@login_required
def api_bulk_download():
    """各種データ一括をテキストファイルでダウンロードする。パラメータ: kind"""
    import io
    from flask import send_file

    user_id = session.get('user_id')
    if not can_view(user_id):
        return jsonify({'success': False, 'error': '権限がありません'}), 403
    kind = (request.args.get('kind') or '').strip()
    if kind not in BULK_KINDS:
        return jsonify({'success': False,
                        'error': '種別の指定が不正です'}), 400
    fn, label, filename = BULK_KINDS[kind]
    try:
        text = fn()
        buf = io.BytesIO(text.encode('utf-8-sig'))  # ExcelでもUTF-8と認識
        buf.seek(0)
        return send_file(buf, mimetype='text/plain; charset=utf-8',
                         as_attachment=True, download_name=filename)
    except Exception as e:
        logging.error("api_bulk_download error (kind=%s): %s", kind, e)
        return jsonify({'success': False,
                        'error': f'生成に失敗しました：{e}'}), 500


# ────────────────────────────────────────────
# データ更新（公式テーブルのExcel更新ワークフロー）
#   担当者（アイテムの管理グループ所属者）がExcelを提供し、
#   全体管理者がDBバックアップを取りながら正式テーブルへ転記する。
#   data_post の思想を公式テーブル固定・登録簿直結で単純化したもの。
# ────────────────────────────────────────────

UPDATE_STATUS_LABELS = {
    'pending':  '提出済（転記待ち）',
    'applied':  '転記済',
    'rejected': '却下',
}
UPLOAD_MAX_BYTES = 10 * 1024 * 1024   # 10MB
UPLOAD_MAX_ROWS  = 50000              # 転記の最大行数（暴走防止）


def _upload_dir():
    """提供ファイルの保存先（プラットフォーム独立：Config経由）。"""
    d = os.path.join(Config.UPLOAD_BASE_DIR, 'fujinp', 'static',
                     'official_data_archive_uploads')
    os.makedirs(d, exist_ok=True)
    return d


def _get_table_item(item_id):
    """公式テーブル登録簿の1件を返す。なければ None。"""
    try:
        conn = mysql.connector.connect(**DatabaseConfig.default())
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT id, table_name, database_name, display_name,
                   manager_group_id
            FROM official_data_archive_tables WHERE id = %s
        """, (item_id,))
        return cursor.fetchone()
    except Exception as e:
        logging.error("official_data_archive _get_table_item: %s", e)
        return None
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close(); conn.close()


def _table_columns(table_name, database_name):
    """対象テーブルの列名リスト。失敗時は None。"""
    try:
        conn = mysql.connector.connect(**DB_CHOICES[database_name]())
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SHOW COLUMNS FROM `%s`" % table_name)
        return [c['Field'] for c in cursor.fetchall()]
    except Exception as e:
        logging.error("official_data_archive _table_columns %s: %s",
                      table_name, e)
        return None
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close(); conn.close()


def _read_xlsx(path):
    """
    提供されたxlsxを読み、(header(list), rows(list of list)) を返す。
    先頭シートの1行目をヘッダーとみなす。
    """
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    header = []
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            header = [str(c).strip() if c is not None else '' for c in row]
            # 末尾の空ヘッダーを落とす
            while header and not header[-1]:
                header.pop()
            continue
        vals = list(row)[:len(header)]
        # 完全に空の行はスキップ
        if all(v is None or (isinstance(v, str) and not v.strip())
               for v in vals):
            continue
        # 不足分は None で埋める
        vals += [None] * (len(header) - len(vals))
        rows.append(vals)
        if len(rows) > UPLOAD_MAX_ROWS:
            raise ValueError(f'行数が上限（{UPLOAD_MAX_ROWS}行）を'
                             '超えています')
    wb.close()
    return header, rows


def _check_header(header, columns):
    """ヘッダーとテーブル列の照合。問題なければ None、あればエラー文字列。"""
    if not header:
        return 'Excelの1行目（ヘッダー行）が読み取れません'
    hset, cset = set(header), set(columns)
    if len(hset) != len(header):
        return 'Excelのヘッダーに重複する列名があります'
    missing = cset - hset
    extra   = hset - cset
    if missing or extra:
        msgs = []
        if missing:
            msgs.append('不足列：' + '、'.join(sorted(missing)))
        if extra:
            msgs.append('余分な列：' + '、'.join(sorted(extra)))
        return ('列構成がテーブルと一致しません（' + '　'.join(msgs) + '）。'
                'まず公式テーブルタブから現在のXLSXをダウンロードし、'
                'それを編集してください')
    return None


@official_data_archive_bp.route('/api/updates', methods=['GET'])
@login_required
def api_updates():
    """
    データ更新の履歴一覧（提供日時の降順）。
    あわせて、呼び出しユーザが提供できる公式テーブルアイテム一覧
    （uploadable_items）と、転記・却下の可否（can_apply）を返す。
    """
    user_id = session.get('user_id')
    if not can_view(user_id):
        return jsonify({'success': False, 'error': '権限がありません'}), 403
    try:
        conn = mysql.connector.connect(**DatabaseConfig.default())
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT u.id, u.table_item_id, u.table_name, u.database_name,
                   u.original_filename, u.stored_filename,
                   u.note, u.status, u.source_ref,
                   u.uploaded_by, u.uploaded_at,
                   u.applied_by, u.applied_at,
                   u.backup_table, u.reject_reason,
                   t.display_name, t.manager_group_id,
                   COALESCE(up.full_name, u.uploaded_by_name)
                       AS uploaded_by_name,
                   ap.full_name AS applied_by_name
            FROM official_data_archive_updates u
            LEFT JOIN official_data_archive_tables t
                   ON u.table_item_id = t.id
            LEFT JOIN users up ON u.uploaded_by = up.id
            LEFT JOIN users ap ON u.applied_by  = ap.id
            ORDER BY u.uploaded_at DESC
            LIMIT 500
        """)
        updates = cursor.fetchall()

        cursor.execute("""
            SELECT id, table_name, database_name, display_name,
                   manager_group_id
            FROM official_data_archive_tables
            ORDER BY COALESCE(display_name, table_name)
        """)
        items = cursor.fetchall()
        cursor.close(); conn.close()

        is_global = is_global_manager(user_id)
        my_ids = {g['id'] for g in user_manage_groups(user_id)}

        def manageable(mgid):
            return bool(is_global or (mgid in my_ids if mgid else False))

        for u in updates:
            u['uploaded_at'] = fmt_minutes(u['uploaded_at'])
            u['applied_at']  = fmt_minutes(u['applied_at'])
            u['status_label'] = UPDATE_STATUS_LABELS.get(
                u['status'], u['status'])
            # 提供ファイルのダウンロード可否（全体管理者・提供者・担当者）。
            # 他アプリから移行した記録（stored_filename が 'migrated:'）は
            # 実ファイルが無いのでダウンロード対象外。
            u['can_file'] = bool(
                (is_global or u['uploaded_by'] == user_id
                 or manageable(u.get('manager_group_id')))
                and u.get('stored_filename')
                and not str(u['stored_filename']).startswith('migrated:'))

        uploadable = [
            {'id': it['id'],
             'label': (it['display_name'] or it['table_name'])
                      + '（' + it['database_name'] + ' / '
                      + it['table_name'] + '）'}
            for it in items if manageable(it.get('manager_group_id'))
        ]

        return jsonify({
            'success': True,
            'updates': updates,
            'uploadable_items': uploadable,
            'can_apply': is_global,
        })
    except Exception as e:
        logging.error("api_updates error: %s", e)
        return jsonify({'success': False, 'error': str(e)}), 500


def fmt_minutes(d):
    """datetime → 'YYYY-MM-DD HH:MM' 文字列。None は ''。"""
    if d is None:
        return ''
    if isinstance(d, (datetime.datetime, datetime.date)):
        return d.strftime('%Y-%m-%d %H:%M')
    return str(d)


@official_data_archive_bp.route('/api/updates/upload', methods=['POST'])
@login_required
def api_update_upload():
    """
    更新データ（Excel）の提供。multipart/form-data：
        item_id    : 公式テーブル登録簿のid
        note       : メモ（任意）
        excel_file : .xlsx ファイル
    対象アイテムの管理グループ所属者（または全体管理者）のみ。
    アップロード時に列構成をテーブルと照合し、不一致は受け付けない。
    """
    user_id = session.get('user_id')
    try:
        item_id = request.form.get('item_id', type=int)
        note    = (request.form.get('note') or '').strip() or None
        f       = request.files.get('excel_file')

        item = _get_table_item(item_id)
        if not item:
            return jsonify({'success': False,
                            'error': '対象の公式テーブルが見つかりません'}), 404
        if not _can_manage_item(user_id, item.get('manager_group_id')):
            return jsonify({'success': False,
                            'error': 'このテーブルの更新データを提供する'
                                     '権限がありません'}), 403
        if not f or not f.filename:
            return jsonify({'success': False,
                            'error': 'Excelファイルを選択してください'}), 400
        if not f.filename.lower().endswith('.xlsx'):
            return jsonify({'success': False,
                            'error': '.xlsx ファイルのみ受け付けます'}), 400
        if request.content_length and request.content_length > UPLOAD_MAX_BYTES:
            return jsonify({'success': False,
                            'error': 'ファイルが大きすぎます（10MBまで）'}), 400

        columns = _table_columns(item['table_name'], item['database_name'])
        if columns is None:
            return jsonify({'success': False,
                            'error': '対象テーブルの列構成を取得できません'}), 500

        now = get_jst_now()
        stored = 'upd_{0}_u{1}_i{2}.xlsx'.format(
            now.strftime('%Y%m%d%H%M%S'), user_id, item_id)
        path = os.path.join(_upload_dir(), stored)
        f.save(path)

        # ── 列構成の照合（不一致なら保存を取り消して受付拒否） ──
        try:
            header, rows = _read_xlsx(path)
            err = _check_header(header, columns)
            if err:
                os.remove(path)
                return jsonify({'success': False, 'error': err}), 400
        except Exception as xe:
            try:
                os.remove(path)
            except OSError:
                pass
            return jsonify({'success': False,
                            'error': f'Excelの読み取りに失敗しました：{xe}'}), 400

        conn = mysql.connector.connect(**DatabaseConfig.default())
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO official_data_archive_updates
                (table_item_id, table_name, database_name,
                 original_filename, stored_filename, note,
                 uploaded_by, uploaded_at, status)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, 'pending')
        """, (item_id, item['table_name'], item['database_name'],
              f.filename[:255], stored, note, user_id, now))
        conn.commit()
        return jsonify({'success': True, 'id': cursor.lastrowid,
                        'rows': len(rows)})
    except Exception as e:
        logging.error("api_update_upload error: %s", e)
        if 'conn' in locals():
            conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close(); conn.close()


def _get_update(update_id):
    """更新レコード1件。なければ None。"""
    try:
        conn = mysql.connector.connect(**DatabaseConfig.default())
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT u.*, t.manager_group_id
            FROM official_data_archive_updates u
            LEFT JOIN official_data_archive_tables t
                   ON u.table_item_id = t.id
            WHERE u.id = %s
        """, (update_id,))
        return cursor.fetchone()
    except Exception as e:
        logging.error("official_data_archive _get_update: %s", e)
        return None
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close(); conn.close()


@official_data_archive_bp.route('/api/updates/file', methods=['GET'])
@login_required
def api_update_file():
    """提供ファイル（xlsx）のダウンロード。パラメータ: id"""
    from flask import send_file
    user_id = session.get('user_id')
    u = _get_update(request.args.get('id', type=int))
    if not u:
        return jsonify({'success': False,
                        'error': '更新レコードが見つかりません'}), 404
    if not (is_global_manager(user_id) or u['uploaded_by'] == user_id
            or _can_manage_item(user_id, u.get('manager_group_id'))):
        return jsonify({'success': False, 'error': '権限がありません'}), 403
    path = os.path.join(_upload_dir(), os.path.basename(u['stored_filename']))
    if not os.path.exists(path):
        return jsonify({'success': False,
                        'error': '提供ファイルが見つかりません'}), 404
    return send_file(path, as_attachment=True,
                     download_name=u['original_filename'] or u['stored_filename'])


@official_data_archive_bp.route('/api/updates/apply', methods=['POST'])
@login_required
def api_update_apply():
    """
    転記（全体管理者のみ）。リクエスト: { "id": 更新レコードのid }
    手順：
      1. 正式テーブルのDBバックアップを作成
         （CREATE TABLE `<name>_bak_<日時>` LIKE → INSERT SELECT）
      2. 正式テーブルを全件削除し、Excelの内容を挿入（同一トランザクション。
         失敗時はロールバック＝正式テーブルは転記前の状態のまま）
      3. レコードを applied に更新（転記者・日時・バックアップ名を記録）
    """
    user_id = session.get('user_id')
    if not is_global_manager(user_id):
        return jsonify({'success': False,
                        'error': '転記は全体管理者のみ実行できます'}), 403
    data = request.json or {}
    u = _get_update(data.get('id'))
    if not u:
        return jsonify({'success': False,
                        'error': '更新レコードが見つかりません'}), 404
    if u['status'] != 'pending':
        return jsonify({'success': False,
                        'error': 'このレコードはすでに処理済みです'
                                 f"（{UPDATE_STATUS_LABELS.get(u['status'])}）"}), 400
    if u['database_name'] not in DB_CHOICES:
        return jsonify({'success': False,
                        'error': '対象データベースの指定が不正です'}), 400

    table = u['table_name']
    path = os.path.join(_upload_dir(), os.path.basename(u['stored_filename']))
    if not os.path.exists(path):
        return jsonify({'success': False,
                        'error': '提供ファイルが見つかりません'}), 404

    try:
        # ── Excel再読込と列構成の最終確認 ──
        header, rows = _read_xlsx(path)
        columns = _table_columns(table, u['database_name'])
        if columns is None:
            return jsonify({'success': False,
                            'error': '対象テーブルの列構成を取得できません'}), 500
        err = _check_header(header, columns)
        if err:
            return jsonify({'success': False, 'error': err}), 400

        now = get_jst_now()
        backup = '{0}_bak_{1}'.format(table[:43], now.strftime('%Y%m%d%H%M%S'))

        db_conn = mysql.connector.connect(**DB_CHOICES[u['database_name']]())
        db_cur  = db_conn.cursor()

        # ── 1. バックアップ（DDLは即時コミットされる） ──
        db_cur.execute("CREATE TABLE `%s` LIKE `%s`" % (backup, table))
        db_cur.execute("INSERT INTO `%s` SELECT * FROM `%s`" % (backup, table))
        db_conn.commit()

        # ── 2. 全件削除→挿入（単一トランザクション） ──
        try:
            db_cur.execute("SET FOREIGN_KEY_CHECKS=0")
            db_cur.execute("DELETE FROM `%s`" % table)
            col_list = ', '.join('`%s`' % c for c in header)
            ph       = ', '.join(['%s'] * len(header))
            ins_sql  = ("INSERT INTO `%s` (%s) VALUES (%s)"
                        % (table, col_list, ph))
            cleaned = [
                tuple((v.strip() if isinstance(v, str) else v)
                      if not (isinstance(v, str) and not v.strip()) else None
                      for v in row)
                for row in rows
            ]
            if cleaned:
                db_cur.executemany(ins_sql, cleaned)
            db_cur.execute("SET FOREIGN_KEY_CHECKS=1")
            db_conn.commit()
        except Exception:
            db_conn.rollback()
            try:
                db_cur.execute("SET FOREIGN_KEY_CHECKS=1")
            except Exception:
                pass
            raise
        finally:
            db_cur.close(); db_conn.close()

        # ── 3. レコード更新 ──
        conn = mysql.connector.connect(**DatabaseConfig.default())
        cursor = conn.cursor()
        cursor.execute("""
            UPDATE official_data_archive_updates
            SET status='applied', applied_by=%s, applied_at=%s,
                backup_table=%s
            WHERE id=%s
        """, (user_id, now, backup, u['id']))
        conn.commit()
        return jsonify({'success': True, 'rows': len(rows),
                        'backup_table': backup})
    except Exception as e:
        logging.error("api_update_apply error (id=%s): %s",
                      data.get('id'), e)
        if 'conn' in locals():
            conn.rollback()
        return jsonify({'success': False,
                        'error': f'転記に失敗しました：{e}'}), 500
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close(); conn.close()


@official_data_archive_bp.route('/api/updates/reject', methods=['POST'])
@login_required
def api_update_reject():
    """
    却下（全体管理者のみ）。リクエスト: { "id": ..., "reason": "（任意）" }
    提供ファイルは記録として残す。
    """
    user_id = session.get('user_id')
    if not is_global_manager(user_id):
        return jsonify({'success': False,
                        'error': '却下は全体管理者のみ実行できます'}), 403
    try:
        data   = request.json or {}
        u      = _get_update(data.get('id'))
        reason = (data.get('reason') or '').strip()[:255] or None
        if not u:
            return jsonify({'success': False,
                            'error': '更新レコードが見つかりません'}), 404
        if u['status'] != 'pending':
            return jsonify({'success': False,
                            'error': 'このレコードはすでに処理済みです'}), 400
        conn = mysql.connector.connect(**DatabaseConfig.default())
        cursor = conn.cursor()
        cursor.execute("""
            UPDATE official_data_archive_updates
            SET status='rejected', applied_by=%s, applied_at=%s,
                reject_reason=%s
            WHERE id=%s
        """, (user_id, get_jst_now(), reason, u['id']))
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        logging.error("api_update_reject error: %s", e)
        if 'conn' in locals():
            conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close(); conn.close()


# ────────────────────────────────────────────
# ダッシュボードへ戻る
# ────────────────────────────────────────────

@official_data_archive_bp.route('/return_to_fujin')
@login_required
def return_to_fujin():
    """FUJIN-Pダッシュボードに戻る。"""
    return redirect_to_dashboard()