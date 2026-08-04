# SPDX-FileCopyrightText: 2024-2026 Toyoaki Nishida
# SPDX-License-Identifier: AGPL-3.0-or-later
#
# This file is part of FUJIN-P.
# Copyright (C) 2024-2026 Toyoaki Nishida
#
# FUJIN-P is free software: you can redistribute it and/or modify
# it under the terms of the GNU Affero General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# FUJIN-P is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU Affero General Public License for more details.
#
# You should have received a copy of the GNU Affero General Public License
# along with FUJIN-P.  If not, see <https://www.gnu.org/licenses/>.
#
# Source: https://github.com/nishida-toyoaki/fujin-p

"""
data_center routes.py
ストーリー管理・ノードツリー編集・閲覧（プレゼンテーション）
"""
import datetime
import logging

from pytz import timezone
from flask import render_template, request, jsonify, session
import mysql.connector
from auth import redirect_to_dashboard

from db import DatabaseConfig
from decorators import login_required
from . import data_center_bp

JST = timezone('Asia/Tokyo')

SOURCE_PROJECT_ID = 1
SOURCE_PROJECT_NAME = '大学公式データ_2025'
STORY_CREATOR_GROUP_NAME = 'データセンター_クリエイター'


def get_jst_now():
    return datetime.datetime.now(JST).replace(tzinfo=None)

# ---- アクセス制御ヘルパー ----------------------------------------

def _get_user_info(user_id):
    """users テーブルから category を取得"""
    try:
        conn = mysql.connector.connect(**DatabaseConfig.default())
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT category FROM users WHERE id = %s", (user_id,))
        return cursor.fetchone()
    except Exception:
        return None
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close(); conn.close()

def _user_group_ids(user_id):
    """現在有効な user_groups の id リストを返す。NULL の期間端は無期限。"""
    try:
        conn = mysql.connector.connect(**DatabaseConfig.default())
        cursor = conn.cursor()
        now = get_jst_now()
        cursor.execute("""
            SELECT group_id
            FROM user_group_memberships
            WHERE user_id = %s
              AND (valid_from IS NULL OR valid_from <= %s)
              AND (valid_until IS NULL OR valid_until >= %s)
        """, (user_id, now, now))
        return [r[0] for r in cursor.fetchall()]
    except Exception:
        return []
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close(); conn.close()

def _can_view(story, user_id):
    """閲覧権限チェック。True なら閲覧可"""
    if story['owner_user_id'] == user_id:
        return True
    policy = story['access_policy']
    if policy == 'public':
        return True
    if policy == 'private':
        info = _get_user_info(user_id)
        return info and info['category'] == 'admin'
    if policy == 'domestic':
        info = _get_user_info(user_id)
        return info and info['category'] in ('regular', 'admin')
    if policy == 'group':
        if not story['group_id']:
            return False
        return story['group_id'] in _user_group_ids(user_id)
    return False

def _can_edit(story, user_id):
    """編集権限チェック（オーナー or admin）"""
    if story['owner_user_id'] == user_id:
        return True
    info = _get_user_info(user_id)
    return info and info['category'] == 'admin'

def _can_create_story(user_id):
    """admin または指定クリエイターグループの有効な所属者なら作成可。"""
    info = _get_user_info(user_id)
    if info and info['category'] == 'admin':
        return True
    try:
        conn = mysql.connector.connect(**DatabaseConfig.default())
        cursor = conn.cursor()
        now = get_jst_now()
        cursor.execute("""
            SELECT 1
            FROM user_group_memberships m
            INNER JOIN user_groups g ON g.id = m.group_id
            WHERE m.user_id = %s
              AND g.name = %s
              AND (m.valid_from IS NULL OR m.valid_from <= %s)
              AND (m.valid_until IS NULL OR m.valid_until >= %s)
            LIMIT 1
        """, (user_id, STORY_CREATOR_GROUP_NAME, now, now))
        return cursor.fetchone() is not None
    except Exception:
        logging.exception("story creation permission check failed")
        return False
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close(); conn.close()

# ---- ツリー構築ヘルパー ------------------------------------------

def _build_tree(nodes):
    """
    フラットなノードリストを parent_id で入れ子ツリーに変換。
    各ノードに 'children' キーを付与して返す。
    ルート（parent_id IS NULL）のリストを返す。
    """
    by_id = {n['id']: dict(n, children=[]) for n in nodes}
    roots = []
    for n in by_id.values():
        pid = n['parent_id']
        if pid is None:
            roots.append(n)
        elif pid in by_id:
            by_id[pid]['children'].append(n)
    # 各階層を sort_order でソート
    def sort_children(node_list):
        node_list.sort(key=lambda x: x['sort_order'])
        for n in node_list:
            sort_children(n['children'])
    sort_children(roots)
    return roots

def _assign_numbers(nodes, prefix=''):
    """ツリーに表示番号を付与（再帰）。nodes は sort_order 済みリスト。"""
    for i, node in enumerate(nodes, 1):
        num = f"{prefix}{i}" if prefix else str(i)
        node['node_number'] = num
        _assign_numbers(node['children'], num + '.')

# ============================================================
# トップ画面（ストーリー集）
# ============================================================

@data_center_bp.route('/')
@login_required
def index():
    return render_template(
        'data_center/index.html',
        can_create_story=_can_create_story(session.get('user_id'))
    )

# ============================================================
# ストーリー一覧 API
# ============================================================

@data_center_bp.route('/api/stories', methods=['GET'])
@login_required
def api_stories():
    user_id = session.get('user_id')
    try:
        conn = mysql.connector.connect(**DatabaseConfig.default())
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT s.*, u.full_name AS owner_name,
                   g.name AS group_name,
                   (SELECT COUNT(*) FROM dc_nodes n WHERE n.story_id = s.id) AS node_count
            FROM dc_stories s
            LEFT JOIN users u ON s.owner_user_id = u.id
            LEFT JOIN user_groups g ON s.group_id = g.id
            ORDER BY s.sort_order, s.updated_at DESC
        """)
        rows = cursor.fetchall()
        result = []
        for r in rows:
            for k in ('created_at', 'updated_at'):
                if r.get(k): r[k] = r[k].isoformat()
            if _can_view(r, user_id):
                r['can_edit'] = _can_edit(r, user_id)
                result.append(r)
        return jsonify({'success': True, 'stories': result})
    except Exception as e:
        logging.error("api_stories error: %s", e)
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close(); conn.close()

# ============================================================
# ストーリー保存（新規 / 更新）
# ============================================================

@data_center_bp.route('/api/story/save', methods=['POST'])
@login_required
def api_story_save():
    user_id = session.get('user_id')
    data = request.json
    story_id    = data.get('id')
    title       = (data.get('title') or '').strip()
    description = (data.get('description') or '').strip()
    if not title:
        return jsonify({'success': False, 'error': 'タイトルは必須です'}), 400
    now = get_jst_now()
    try:
        conn = mysql.connector.connect(**DatabaseConfig.default())
        cursor = conn.cursor(dictionary=True)
        if story_id:
            cursor.execute("SELECT * FROM dc_stories WHERE id = %s", (story_id,))
            story = cursor.fetchone()
            if not story or not _can_edit(story, user_id):
                return jsonify({'success': False, 'error': '編集権限がありません'}), 403
            cursor.execute("""
                UPDATE dc_stories
                SET title=%s, description=%s, updated_at=%s
                WHERE id=%s
            """, (title, description, now, story_id))
            conn.commit()
            return jsonify({'success': True, 'message': 'ストーリーを更新しました', 'id': story_id})
        else:
            if not _can_create_story(user_id):
                return jsonify({
                    'success': False,
                    'error': (
                        'ストーリーを作成できるのは管理者または'
                        f'「{STORY_CREATOR_GROUP_NAME}」グループの'
                        '有効な所属者だけです'
                    )
                }), 403
            cursor.execute("""
                INSERT INTO dc_stories
                    (title, description, owner_user_id, access_policy, sort_order, created_at, updated_at)
                VALUES (%s, %s, %s, 'private', 0, %s, %s)
            """, (title, description, user_id, now, now))
            conn.commit()
            new_id = cursor.lastrowid
            return jsonify({'success': True, 'message': 'ストーリーを作成しました', 'id': new_id})
    except Exception as e:
        logging.error("api_story_save error: %s", e)
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close(); conn.close()

# ============================================================
# ストーリー削除
# ============================================================

@data_center_bp.route('/api/story/delete', methods=['POST'])
@login_required
def api_story_delete():
    user_id  = session.get('user_id')
    story_id = request.json.get('id')
    try:
        conn = mysql.connector.connect(**DatabaseConfig.default())
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT * FROM dc_stories WHERE id = %s", (story_id,))
        story = cursor.fetchone()
        if not story or not _can_edit(story, user_id):
            return jsonify({'success': False, 'error': '削除権限がありません'}), 403
        cursor.execute("DELETE FROM dc_stories WHERE id = %s", (story_id,))
        conn.commit()
        return jsonify({'success': True, 'message': 'ストーリーを削除しました'})
    except Exception as e:
        logging.error("api_story_delete error: %s", e)
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close(); conn.close()

# ============================================================
# アクセスポリシー設定
# ============================================================

@data_center_bp.route('/api/story/<int:story_id>/access', methods=['GET', 'POST'])
@login_required
def api_story_access(story_id):
    user_id = session.get('user_id')
    try:
        conn = mysql.connector.connect(**DatabaseConfig.default())
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT * FROM dc_stories WHERE id = %s", (story_id,))
        story = cursor.fetchone()
        if not story or not _can_edit(story, user_id):
            return jsonify({'success': False, 'error': '権限がありません'}), 403

        if request.method == 'GET':
            cursor.execute("SELECT id, name FROM user_groups ORDER BY name")
            groups = cursor.fetchall()
            return jsonify({
                'success'   : True,
                'policy'    : story['access_policy'],
                'group_id'  : story['group_id'],
                'all_groups': groups
            })

        data     = request.json
        policy   = data.get('policy')
        group_id = data.get('group_id') or None
        if policy not in ('public', 'domestic', 'group', 'private'):
            return jsonify({'success': False, 'error': '不正なポリシーです'}), 400
        if policy == 'group' and not group_id:
            return jsonify({'success': False, 'error': 'グループを指定してください'}), 400

        cursor.execute("""
            UPDATE dc_stories SET access_policy=%s, group_id=%s, updated_at=%s WHERE id=%s
        """, (policy, group_id if policy == 'group' else None, get_jst_now(), story_id))
        conn.commit()
        return jsonify({'success': True, 'message': 'アクセス設定を保存しました'})
    except Exception as e:
        logging.error("api_story_access error: %s", e)
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close(); conn.close()

# ============================================================
# ノードツリー取得
# ============================================================

@data_center_bp.route('/api/story/<int:story_id>/tree', methods=['GET'])
@login_required
def api_node_tree(story_id):
    user_id = session.get('user_id')
    try:
        conn = mysql.connector.connect(**DatabaseConfig.default())
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT * FROM dc_stories WHERE id = %s", (story_id,))
        story = cursor.fetchone()
        if not story or not _can_view(story, user_id):
            return jsonify({'success': False, 'error': 'アクセス権がありません'}), 403

        cursor.execute("""
            SELECT id, story_id, parent_id, sort_order,
                   LEFT(body_md, 80) AS body_preview,
                   CASE WHEN part_html IS NOT NULL THEN 1 ELSE 0 END AS has_part,
                   created_at, updated_at
            FROM dc_nodes
            WHERE story_id = %s
            ORDER BY sort_order
        """, (story_id,))
        nodes = cursor.fetchall()
        for n in nodes:
            for k in ('created_at', 'updated_at'):
                if n.get(k): n[k] = n[k].isoformat()

        tree = _build_tree(nodes)
        _assign_numbers(tree)
        return jsonify({
            'success'   : True,
            'story'     : {k: (v.isoformat() if isinstance(v, datetime.datetime) else v)
                           for k, v in story.items()},
            'tree'      : tree,
            'can_edit'  : _can_edit(story, user_id)
        })
    except Exception as e:
        logging.error("api_node_tree error: %s", e)
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close(); conn.close()

# ============================================================
# ノード1件取得（編集用・完全データ）
# ============================================================

@data_center_bp.route('/api/node/<int:node_id>', methods=['GET'])
@login_required
def api_node_get(node_id):
    user_id = session.get('user_id')
    try:
        conn = mysql.connector.connect(**DatabaseConfig.default())
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT * FROM dc_nodes WHERE id = %s", (node_id,))
        node = cursor.fetchone()
        if not node:
            return jsonify({'success': False, 'error': 'ノードが見つかりません'}), 404
        cursor.execute("SELECT * FROM dc_stories WHERE id = %s", (node['story_id'],))
        story = cursor.fetchone()
        if not story or not _can_edit(story, user_id):
            return jsonify({'success': False, 'error': '権限がありません'}), 403
        for k in ('created_at', 'updated_at'):
            if node.get(k): node[k] = node[k].isoformat()
        return jsonify({'success': True, 'node': node})
    except Exception as e:
        logging.error("api_node_get error: %s", e)
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close(); conn.close()

# ============================================================
# ノード保存（新規 / 更新）
# ============================================================

@data_center_bp.route('/api/node/save', methods=['POST'])
@login_required
def api_node_save():
    user_id = session.get('user_id')
    data    = request.json
    node_id   = data.get('id')
    story_id  = data.get('story_id')
    parent_id = data.get('parent_id')   # None = 第1世代
    body_md   = data.get('body_md', '')
    part_html = data.get('part_html') or None
    meta_note = data.get('meta_note', '')
    now = get_jst_now()
    try:
        conn = mysql.connector.connect(**DatabaseConfig.default())
        cursor = conn.cursor(dictionary=True)

        # 権限確認（story 経由）
        sid = story_id if not node_id else None
        if node_id:
            cursor.execute("SELECT story_id FROM dc_nodes WHERE id = %s", (node_id,))
            row = cursor.fetchone()
            sid = row['story_id'] if row else None
        cursor.execute("SELECT * FROM dc_stories WHERE id = %s", (sid,))
        story = cursor.fetchone()
        if not story or not _can_edit(story, user_id):
            return jsonify({'success': False, 'error': '編集権限がありません'}), 403

        if node_id:
            cursor.execute("""
                UPDATE dc_nodes
                SET body_md=%s, part_html=%s, meta_note=%s, updated_at=%s
                WHERE id=%s
            """, (body_md, part_html, meta_note, now, node_id))
            conn.commit()
            return jsonify({'success': True, 'message': 'ノードを更新しました', 'id': node_id})
        else:
            if parent_id is not None:
                cursor.execute(
                    "SELECT story_id FROM dc_nodes WHERE id = %s",
                    (parent_id,))
                parent = cursor.fetchone()
                if not parent:
                    return jsonify({
                        'success': False,
                        'error': '親ノードが見つかりません'
                    }), 400
                if parent['story_id'] != sid:
                    return jsonify({
                        'success': False,
                        'error': '別のストーリーのノードは親に指定できません'
                    }), 400

            # 兄弟の最大 sort_order を取得して末尾に追加
            cursor.execute("""
                SELECT COALESCE(MAX(sort_order), -1) + 1 AS next_order
                FROM dc_nodes
                WHERE story_id = %s AND parent_id <=> %s
            """, (sid, parent_id))
            next_order = cursor.fetchone()['next_order']
            cursor.execute("""
                INSERT INTO dc_nodes
                    (story_id, parent_id, sort_order, part_html, body_md, meta_note, created_at, updated_at)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            """, (sid, parent_id, next_order, part_html, body_md, meta_note, now, now))
            conn.commit()
            new_id = cursor.lastrowid
            return jsonify({'success': True, 'message': 'ノードを作成しました', 'id': new_id})
    except Exception as e:
        logging.error("api_node_save error: %s", e)
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close(); conn.close()

# ============================================================
# ノード削除
# ============================================================

@data_center_bp.route('/api/node/delete', methods=['POST'])
@login_required
def api_node_delete():
    user_id = session.get('user_id')
    node_id = request.json.get('id')
    try:
        conn = mysql.connector.connect(**DatabaseConfig.default())
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT * FROM dc_nodes WHERE id = %s", (node_id,))
        node = cursor.fetchone()
        if not node:
            return jsonify({'success': False, 'error': 'ノードが見つかりません'}), 404
        cursor.execute("SELECT * FROM dc_stories WHERE id = %s", (node['story_id'],))
        story = cursor.fetchone()
        if not story or not _can_edit(story, user_id):
            return jsonify({'success': False, 'error': '削除権限がありません'}), 403
        # CASCADE により子孫も自動削除される
        cursor.execute("DELETE FROM dc_nodes WHERE id = %s", (node_id,))
        conn.commit()
        return jsonify({'success': True, 'message': 'ノード（および子孫）を削除しました'})
    except Exception as e:
        logging.error("api_node_delete error: %s", e)
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close(); conn.close()

# ============================================================
# ノード移動（parent_id・sort_order の変更）
# ============================================================

@data_center_bp.route('/api/node/move', methods=['POST'])
@login_required
def api_node_move():
    """
    ノードを別の親または別の位置に移動する。
    payload: { id, new_parent_id (null可), new_sort_order }
    """
    user_id = session.get('user_id')
    data    = request.json
    node_id       = data.get('id')
    new_parent_id = data.get('new_parent_id')   # None = ルート
    try:
        new_sort_order = max(int(data.get('new_sort_order', 0)), 0)
    except (TypeError, ValueError):
        return jsonify({'success': False, 'error': '移動位置が不正です'}), 400
    try:
        conn = mysql.connector.connect(**DatabaseConfig.default())
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT * FROM dc_nodes WHERE id = %s", (node_id,))
        node = cursor.fetchone()
        if not node:
            return jsonify({'success': False, 'error': 'ノードが見つかりません'}), 404
        cursor.execute("SELECT * FROM dc_stories WHERE id = %s", (node['story_id'],))
        story = cursor.fetchone()
        if not story or not _can_edit(story, user_id):
            return jsonify({'success': False, 'error': '編集権限がありません'}), 403

        if new_parent_id is not None:
            if new_parent_id == node_id:
                return jsonify({
                    'success': False,
                    'error': 'ノードを自分自身の下へ移動できません'
                }), 400

            cursor.execute(
                "SELECT id, story_id, parent_id FROM dc_nodes WHERE id = %s",
                (new_parent_id,))
            parent = cursor.fetchone()
            if not parent:
                return jsonify({
                    'success': False,
                    'error': '移動先ノードが見つかりません'
                }), 400
            if parent['story_id'] != node['story_id']:
                return jsonify({
                    'success': False,
                    'error': '別のストーリーへノードを移動できません'
                }), 400

            # 移動先から親をたどり、移動元の子孫でないことを確認する。
            ancestor_id = new_parent_id
            visited = set()
            while ancestor_id is not None:
                if ancestor_id == node_id:
                    return jsonify({
                        'success': False,
                        'error': 'ノードを自分の子孫の下へ移動できません'
                    }), 400
                if ancestor_id in visited:
                    return jsonify({
                        'success': False,
                        'error': '既存のツリーに循環参照があります'
                    }), 409
                visited.add(ancestor_id)
                cursor.execute(
                    "SELECT parent_id FROM dc_nodes WHERE id = %s "
                    "AND story_id = %s",
                    (ancestor_id, node['story_id']))
                ancestor = cursor.fetchone()
                if not ancestor:
                    return jsonify({
                        'success': False,
                        'error': '移動先の親子関係が不正です'
                    }), 409
                ancestor_id = ancestor['parent_id']

        # 挿入先の兄弟を後ろにずらす
        cursor.execute("""
            UPDATE dc_nodes
            SET sort_order = sort_order + 1
            WHERE story_id = %s AND parent_id <=> %s AND sort_order >= %s AND id != %s
        """, (node['story_id'], new_parent_id, new_sort_order, node_id))

        cursor.execute("""
            UPDATE dc_nodes
            SET parent_id=%s, sort_order=%s, updated_at=%s
            WHERE id=%s
        """, (new_parent_id, new_sort_order, get_jst_now(), node_id))
        conn.commit()
        return jsonify({'success': True, 'message': 'ノードを移動しました'})
    except Exception as e:
        logging.error("api_node_move error: %s", e)
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close(); conn.close()

# ============================================================
# ノード並び替え（同一親内の sort_order 一括更新）
# ============================================================

@data_center_bp.route('/api/node/reorder', methods=['POST'])
@login_required
def api_node_reorder():
    """
    payload: { story_id, ordered_ids: [id, id, ...] }
    ordered_ids の順番を 0,1,2,… の sort_order に反映する。
    """
    user_id    = session.get('user_id')
    data       = request.json
    story_id   = data.get('story_id')
    ordered_ids = data.get('ordered_ids', [])
    try:
        conn = mysql.connector.connect(**DatabaseConfig.default())
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT * FROM dc_stories WHERE id = %s", (story_id,))
        story = cursor.fetchone()
        if not story or not _can_edit(story, user_id):
            return jsonify({'success': False, 'error': '編集権限がありません'}), 403
        if not isinstance(ordered_ids, list) or not ordered_ids:
            return jsonify({'success': False, 'error': '並び順が不正です'}), 400
        if len(ordered_ids) != len(set(ordered_ids)):
            return jsonify({'success': False, 'error': 'ノードIDが重複しています'}), 400

        placeholders = ','.join(['%s'] * len(ordered_ids))
        cursor.execute(
            "SELECT id, parent_id FROM dc_nodes "
            f"WHERE story_id = %s AND id IN ({placeholders})",
            (story_id, *ordered_ids))
        reorder_nodes = cursor.fetchall()
        if len(reorder_nodes) != len(ordered_ids):
            return jsonify({
                'success': False,
                'error': '別ストーリーまたは存在しないノードが含まれています'
            }), 400
        parent_ids = {row['parent_id'] for row in reorder_nodes}
        if len(parent_ids) != 1:
            return jsonify({
                'success': False,
                'error': '同じ階層のノードだけを並び替えられます'
            }), 400

        parent_id = next(iter(parent_ids))
        cursor.execute("""
            SELECT id FROM dc_nodes
            WHERE story_id = %s AND parent_id <=> %s
        """, (story_id, parent_id))
        sibling_ids = {row['id'] for row in cursor.fetchall()}
        if sibling_ids != set(ordered_ids):
            return jsonify({
                'success': False,
                'error': '兄弟ノードをすべて指定してください'
            }), 400
        now = get_jst_now()
        for i, nid in enumerate(ordered_ids):
            cursor.execute("""
                UPDATE dc_nodes SET sort_order=%s, updated_at=%s
                WHERE id=%s AND story_id=%s
            """, (i, now, nid, story_id))
        conn.commit()
        return jsonify({'success': True, 'message': '並び順を更新しました'})
    except Exception as e:
        logging.error("api_node_reorder error: %s", e)
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close(); conn.close()

# ============================================================
# 編集画面（クラフトマン用）
# ============================================================

@data_center_bp.route('/edit/<int:story_id>')
@login_required
def edit(story_id):
    return render_template('data_center/edit.html', story_id=story_id)

# ============================================================
# 閲覧画面（聴衆用・アクセス制御あり）
# ============================================================

@data_center_bp.route('/view/<int:story_id>')
@login_required
def view(story_id):
    return render_template('data_center/view.html', story_id=story_id)

# ============================================================
# 閲覧用ツリー取得（part_html・body_md。meta_note は非公開）
# ============================================================

@data_center_bp.route('/api/view/<int:story_id>/tree', methods=['GET'])
@login_required
def api_view_tree(story_id):
    user_id = session.get('user_id')
    try:
        conn = mysql.connector.connect(**DatabaseConfig.default())
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT * FROM dc_stories WHERE id = %s", (story_id,))
        story = cursor.fetchone()
        if not story or not _can_view(story, user_id):
            return jsonify({'success': False, 'error': 'アクセス権がありません'}), 403

        cursor.execute("""
            SELECT id, story_id, parent_id, sort_order,
                   part_html, body_md,
                   CASE WHEN part_html IS NOT NULL THEN 1 ELSE 0 END AS has_part
                   -- meta_note は閲覧者には返さない
            FROM dc_nodes
            WHERE story_id = %s
            ORDER BY sort_order
        """, (story_id,))
        nodes = cursor.fetchall()

        tree = _build_tree(nodes)
        _assign_numbers(tree)
        return jsonify({
            'success'  : True,
            'story'    : {k: (v.isoformat() if isinstance(v, datetime.datetime) else v)
                          for k, v in story.items()},
            'tree'     : tree,
            'can_edit' : _can_edit(story, user_id)
        })
    except Exception as e:
        logging.error("api_view_tree error: %s", e)
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close(); conn.close()

# ============================================================
# 元データ閲覧（guest用・読み取り専用）
# ============================================================
@data_center_bp.route('/source/')
@login_required
def source_index():
    return render_template('data_center/source.html')


@data_center_bp.route('/api/source/tables', methods=['GET'])
@login_required
def api_source_tables():
    try:
        conn   = mysql.connector.connect(**DatabaseConfig.default())
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT database_name, table_name
            FROM table_master_project_tables
            WHERE project_id = %s
            ORDER BY table_name
        """, (SOURCE_PROJECT_ID,))
        tables = cursor.fetchall()
        cursor.close(); conn.close()

        for t in tables:
            try:
                db_conn = mysql.connector.connect(
                    **DatabaseConfig.get_config(t['database_name']))
                db_cursor = db_conn.cursor()
                db_cursor.execute(
                    "SELECT COUNT(*) FROM `%s`" % t['table_name'])
                t['row_count'] = db_cursor.fetchone()[0]
            except Exception as e:
                logging.warning(
                    "row_count error %s.%s: %s",
                    t['database_name'], t['table_name'], e)
                t['row_count'] = None
            finally:
                if 'db_conn' in locals() and db_conn.is_connected():
                    db_cursor.close(); db_conn.close()

        return jsonify({
            'success': True,
            'project': SOURCE_PROJECT_NAME,
            'tables': tables
        })

    except Exception as e:
        logging.error("api_source_tables error: %s", e)
        return jsonify({'success': False, 'error': str(e)}), 500


@data_center_bp.route('/api/source/data', methods=['GET'])
@login_required
def api_source_data():
    table_name = request.args.get('table', '').strip()
    database_name = request.args.get('database', '').strip()
    try:
        limit = min(max(int(request.args.get('limit', 500)), 1), 5000)
        offset = max(int(request.args.get('offset', 0)), 0)
    except (TypeError, ValueError):
        return jsonify({'success': False, 'error': 'limit または offset が不正です'}), 400

    if not table_name or not database_name:
        return jsonify({
            'success': False,
            'error': 'データベース名とテーブル名が必要です'
        }), 400

    # プロジェクト所属チェック（正規表現なし・DB照合のみ）
    try:
        conn   = mysql.connector.connect(**DatabaseConfig.default())
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT id, database_name FROM table_master_project_tables
            WHERE project_id = %s AND database_name = %s AND table_name = %s
            LIMIT 1
        """, (SOURCE_PROJECT_ID, database_name, table_name))
        row = cursor.fetchone()
        cursor.close(); conn.close()
        if not row:
            return jsonify({'success': False,
                            'error': 'アクセス不可: ' + table_name}), 403
    except Exception as e:
        logging.error("api_source_data check error: %s", e)
        return jsonify({'success': False, 'error': str(e)}), 500

    # データ取得
    try:
        conn   = mysql.connector.connect(
                     **DatabaseConfig.get_config(row['database_name']))
        cursor = conn.cursor(dictionary=True)

        cursor.execute("SHOW COLUMNS FROM `%s`" % table_name)
        columns = [c['Field'] for c in cursor.fetchall()]

        cursor.execute("SELECT COUNT(*) as cnt FROM `%s`" % table_name)
        total_count = cursor.fetchone()['cnt']

        cursor.execute(
            "SELECT * FROM `%s` LIMIT %%s OFFSET %%s" % table_name,
            (limit, offset))
        rows = cursor.fetchall()

        def safe_val(v):
            if v is None: return None
            if isinstance(v, (datetime.datetime, datetime.date)):
                return v.isoformat()
            if isinstance(v, datetime.timedelta):
                s = int(v.total_seconds())
                return f"{s//3600:02d}:{(s%3600)//60:02d}"
            if isinstance(v, bytes):
                return v.decode('utf-8', errors='replace')
            return v

        serialized = [{k: safe_val(v) for k, v in row.items()}
                      for row in rows]

        return jsonify({
            'success'      : True,
            'database'     : row['database_name'],
            'table'        : table_name,
            'columns'      : columns,
            'rows'         : serialized,
            'total_count'  : total_count,
            'fetched_count': len(rows),
            'offset'       : offset,
            'limit'        : limit
        })

    except Exception as e:
        logging.error("api_source_data error: %s", e)
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close(); conn.close()

@data_center_bp.route('/api/source/download', methods=['GET'])
@login_required
def api_source_download():
    """テーブルをXLSXでダウンロード"""
    import io
    import openpyxl
    from flask import send_file

    table_name = request.args.get('table', '').strip()
    database_name = request.args.get('database', '').strip()
    if not table_name or not database_name:
        return jsonify({
            'success': False,
            'error': 'データベース名とテーブル名が必要です'
        }), 400

    # プロジェクト所属チェック
    try:
        conn   = mysql.connector.connect(**DatabaseConfig.default())
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT id, database_name FROM table_master_project_tables
            WHERE project_id = %s AND database_name = %s AND table_name = %s
            LIMIT 1
        """, (SOURCE_PROJECT_ID, database_name, table_name))
        row = cursor.fetchone()
        cursor.close(); conn.close()
        if not row:
            return jsonify({'success': False, 'error': 'アクセス不可'}), 403
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

    # データ取得（全件）
    try:
        conn   = mysql.connector.connect(
                     **DatabaseConfig.get_config(row['database_name']))
        cursor = conn.cursor(dictionary=True)

        cursor.execute("SELECT * FROM `%s`" % table_name)
        rows = cursor.fetchall()
        columns = [d[0] for d in cursor.description] if cursor.description else []
        cursor.close(); conn.close()

        # XLSX生成
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = table_name[:31]  # シート名は31文字まで

        # ヘッダー
        ws.append(columns)

        # データ行
        for row in rows:
            ws.append([
                (v.isoformat() if isinstance(v, (datetime.datetime, datetime.date))
                 else (str(v)
                       if isinstance(v, datetime.timedelta)
                       else (v.decode('utf-8', errors='replace')
                             if isinstance(v, bytes) else v)))
                for v in [row[c] for c in columns]
            ])

        # ヘッダー行のスタイル
        from openpyxl.styles import PatternFill, Font
        header_fill = PatternFill('solid', fgColor='1A3A5C')
        header_font = Font(color='FFFFFF', bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        # バッファに書き出し
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f"{table_name}.xlsx"
        return send_file(
            buf,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        logging.error("api_source_download error: %s", e)
        return jsonify({'success': False, 'error': str(e)}), 500

@data_center_bp.route('/return_to_fujin')
@login_required
def return_to_fujin():
    """FUJIN-Pダッシュボードに戻る"""
    return redirect_to_dashboard()

