# SPDX-FileCopyrightText: 2024-2026 Toyoaki Nishida
# SPDX-License-Identifier: AGPL-3.0-or-later
#
# This file is part of FUJIN-P.
# Copyright (C) 2024-2026 Toyoaki Nishida
#
# FUJIN-P is free software: you can redistribute it and/or modify
# it under the terms of the GNU Affero General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# FUJIN-P is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU Affero General Public License for more details.
#
# You should have received a copy of the GNU Affero General Public License
# along with FUJIN-P.  If not, see <https://www.gnu.org/licenses/>.
#
# Source: https://github.com/nishida-toyoaki/fujin-p

"""
index_review - ルート定義（プロジェクト型）
時系列指標分析プロジェクト単位で指標を管理する。
指標の閲覧・分析は access_policy に基づき一般ユーザに公開される。
"""
import datetime
import json
import logging
from pytz import timezone

from flask import render_template, request, jsonify, session, send_file
import mysql.connector
import io
import pandas as pd
from auth import redirect_to_dashboard

from config import Config
from db import DatabaseConfig, Tables
from decorators import login_required
from . import index_review_bp

JST = timezone('Asia/Tokyo')
TABLE_SET_ADMIN_GROUP = 'パブリックテーブル総管理者'
PROJECT_ACCESS_POLICIES = {'public', 'domestic', 'group', 'private'}


def get_jst_now():
    return datetime.datetime.now(JST).replace(tzinfo=None)


def _get_project_access_policy(data):
    """公開範囲を取得。未指定・NULL・空文字はprivateとして扱う。"""
    raw_policy = data.get('access_policy')

    if raw_policy is None or not str(raw_policy).strip():
        return 'private'

    policy = str(raw_policy).strip().lower()
    return policy if policy in PROJECT_ACCESS_POLICIES else None


# ─────────────────────────────────────────────────────────────
# アクセス制御ヘルパー
# ─────────────────────────────────────────────────────────────

def _is_admin(user_id):
    if not user_id:
        return False
    try:
        conn = mysql.connector.connect(**DatabaseConfig.default())
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT category FROM users WHERE id = %s", (user_id,))
        row = cursor.fetchone()
        return bool(row and row.get('category') == 'admin')
    except Exception:
        return False
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close(); conn.close()


def _is_group_member(user_id, group_id):
    if not user_id or not group_id:
        return False
    try:
        now = get_jst_now()
        conn = mysql.connector.connect(**DatabaseConfig.default())
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT id FROM user_group_memberships
            WHERE group_id = %s AND user_id = %s
              AND (valid_from  IS NULL OR valid_from  <= %s)
              AND (valid_until IS NULL OR valid_until >= %s)
        """, (group_id, user_id, now, now))
        return cursor.fetchone() is not None
    except Exception:
        return False
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close(); conn.close()

def _get_user_group_ids(user_id):
    if not user_id:
        return []
    try:
        now = get_jst_now()
        conn = mysql.connector.connect(**DatabaseConfig.default())
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT group_id FROM user_group_memberships
            WHERE user_id = %s
              AND (valid_from IS NULL OR valid_from <= %s)
              AND (valid_until IS NULL OR valid_until >= %s)
        """, (user_id, now, now))
        return [r['group_id'] for r in cursor.fetchall()]
    except Exception:
        return []
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close(); conn.close()

def _can_view_project(user_id, project):
    if _is_admin(user_id):
        return True
    if project['owner_user_id'] == user_id:
        return True

    policy = str(
        project.get('access_policy') or 'private'
    ).strip().lower()

    if policy == 'public':
        return True
    if policy == 'domestic':
        return session.get('user_category') in ('regular', 'admin')
    if policy == 'group':
        user_groups = _get_user_group_ids(user_id)
        access_groups = project.get('access_group_ids', [])
        return bool(set(user_groups) & set(access_groups))

    # privateおよび未知の値は、admin・所有者以外には公開しない
    return False

def _can_edit_project(user_id, project):
    if _is_admin(user_id):
        return True
    if project['owner_user_id'] == user_id:
        return True
    if project.get('group_id'):
        return _is_group_member(user_id, project['group_id'])
    return False


def _fetch_project(project_id):
    try:
        conn = mysql.connector.connect(**DatabaseConfig.default())
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT p.*, u.full_name AS owner_name, g.name AS group_name
            FROM ir_projects p
            LEFT JOIN users u ON u.id = p.owner_user_id
            LEFT JOIN user_groups g ON g.id = p.group_id
            WHERE p.id = %s
        """, (project_id,))
        p = cursor.fetchone()
        if p:
            cursor.execute(
                "SELECT group_id FROM ir_project_access_groups WHERE project_id = %s",
                (project_id,)
            )
            p['access_group_ids'] = [r['group_id'] for r in cursor.fetchall()]
        return p
    except Exception as e:
        logging.error("_fetch_project error: %s", e)
        return None
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close(); conn.close()


# ─────────────────────────────────────────────────────────────
# ページルート
# ─────────────────────────────────────────────────────────────

@index_review_bp.route('/')
@login_required
def index():
    return render_template('index_review/index.html')


@index_review_bp.route('/table_manager')
@login_required
def table_manager():
    return render_template('index_review/table_manager.html')


# ─────────────────────────────────────────────────────────────
# グループ一覧（プロジェクトフォーム用）
# ─────────────────────────────────────────────────────────────

@index_review_bp.route('/api/groups')
@login_required
def api_groups():
    try:
        conn = mysql.connector.connect(**DatabaseConfig.default())
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT id, name, description FROM user_groups ORDER BY name")
        groups = cursor.fetchall()
        return jsonify({'success': True, 'groups': groups})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close(); conn.close()


# ─────────────────────────────────────────────────────────────
# プロジェクト CRUD
# ─────────────────────────────────────────────────────────────

@index_review_bp.route('/api/projects', methods=['GET'])
@login_required
def api_projects():
    """ログインユーザがアクセス可能なプロジェクト一覧"""
    user_id = session.get('user_id')
    try:
        conn = mysql.connector.connect(**DatabaseConfig.default())
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT p.*,
                   u.full_name AS owner_name,
                   g.name      AS group_name,
                   COUNT(i.id) AS indicator_count
            FROM ir_projects p
            LEFT JOIN users u ON u.id = p.owner_user_id
            LEFT JOIN user_groups g ON g.id = p.group_id
            LEFT JOIN ir_project_indicators i ON i.project_id = p.id
            GROUP BY p.id
            ORDER BY p.name
        """)
        rows = cursor.fetchall()

        visible = []
        for r in rows:
            for k in ('created_at', 'updated_at'):
                if r.get(k):
                    r[k] = r[k].strftime('%Y-%m-%d %H:%M')
            # ★ _can_view_project の前に access_group_ids を取得
            if r.get('access_policy') == 'group':
                conn2 = mysql.connector.connect(**DatabaseConfig.default())
                cur2  = conn2.cursor(dictionary=True)
                cur2.execute(
                    "SELECT group_id FROM ir_project_access_groups WHERE project_id = %s",
                    (r['id'],)
                )
                r['access_group_ids'] = [row['group_id'] for row in cur2.fetchall()]
                cur2.close(); conn2.close()
            else:
                r['access_group_ids'] = []
            # ★ access_group_ids が揃ってから判定
            if _can_view_project(user_id, r):
                r['can_edit'] = _can_edit_project(user_id, r)
                visible.append(r)

        return jsonify({
            'success': True,
            'projects': visible,
            'current_user_id': user_id,
            'is_admin': _is_admin(user_id)
        })
    except Exception as e:
        logging.error("api_projects GET error: %s", e)
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close(); conn.close()


@index_review_bp.route('/api/projects', methods=['POST'])
@login_required
def api_projects_create():
    """プロジェクト新規作成（ログインユーザ全員可）"""
    user_id = session.get('user_id')
    data = request.json or {}
    name = data.get('name', '').strip()
    if not name:
        return jsonify({'success': False, 'error': 'プロジェクト名は必須です'}), 400
    access_policy = _get_project_access_policy(data)
    if access_policy is None:
        return jsonify({'success': False, 'error': '公開範囲の指定が不正です'}), 400

    now = get_jst_now()
    try:
        conn = mysql.connector.connect(**DatabaseConfig.default())
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO ir_projects
              (name, description, owner_user_id, group_id, access_policy, created_at, updated_at)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """, (name,
              data.get('description', '').strip(),
              user_id,
              data.get('group_id') or None,
              access_policy,
              now, now))
        new_id = cursor.lastrowid
        if access_policy == 'group':
            for gid in (data.get('access_group_ids') or []):
                cursor.execute(
                    "INSERT INTO ir_project_access_groups (project_id, group_id) VALUES (%s,%s)",
                    (new_id, gid)
                )

        conn.commit()
        return jsonify({'success': True, 'id': new_id})
    except Exception as e:
        logging.error("api_projects_create error: %s", e)
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close(); conn.close()


@index_review_bp.route('/api/projects/<int:pid>', methods=['GET'])
@login_required
def api_project_detail(pid):
    user_id = session.get('user_id')
    p = _fetch_project(pid)
    if not p:
        return jsonify({'success': False, 'error': '見つかりません'}), 404
    if not _can_view_project(user_id, p):
        return jsonify({'success': False, 'error': 'アクセス権がありません'}), 403
    for k in ('created_at', 'updated_at'):
        if p.get(k):
            p[k] = p[k].strftime('%Y-%m-%d %H:%M')
    p['can_edit'] = _can_edit_project(user_id, p)
    return jsonify({'success': True, 'project': p})


@index_review_bp.route('/api/projects/<int:pid>', methods=['PUT'])
@login_required
def api_project_update(pid):
    user_id = session.get('user_id')
    p = _fetch_project(pid)
    if not p:
        return jsonify({'success': False, 'error': '見つかりません'}), 404
    if not _can_edit_project(user_id, p):
        return jsonify({'success': False, 'error': '権限がありません'}), 403

    data = request.json or {}
    name = data.get('name', '').strip()
    if not name:
        return jsonify({'success': False, 'error': 'プロジェクト名は必須です'}), 400
    access_policy = _get_project_access_policy(data)
    if access_policy is None:
        return jsonify({'success': False, 'error': '公開範囲の指定が不正です'}), 400

    now = get_jst_now()
    try:
        conn = mysql.connector.connect(**DatabaseConfig.default())
        cursor = conn.cursor()
        cursor.execute("""
            UPDATE ir_projects
            SET name=%s, description=%s, group_id=%s, access_policy=%s, updated_at=%s
            WHERE id=%s
        """, (name,
              data.get('description', '').strip(),
              data.get('group_id') or None,
              access_policy,
              now, pid))
        cursor.execute("DELETE FROM ir_project_access_groups WHERE project_id = %s", (pid,))
        if access_policy == 'group':
            for gid in (data.get('access_group_ids') or []):
                cursor.execute(
                    "INSERT INTO ir_project_access_groups (project_id, group_id) VALUES (%s,%s)",
                    (pid, gid)
                )

        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        logging.error("api_project_update error: %s", e)
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close(); conn.close()


@index_review_bp.route('/api/projects/<int:pid>', methods=['DELETE'])
@login_required
def api_project_delete(pid):
    user_id = session.get('user_id')
    p = _fetch_project(pid)
    if not p:
        return jsonify({'success': False, 'error': '見つかりません'}), 404
    # 削除はオーナーまたは管理者のみ
    if not (_is_admin(user_id) or p['owner_user_id'] == user_id):
        return jsonify({'success': False, 'error': '削除はオーナーのみ可能です'}), 403

    try:
        conn = mysql.connector.connect(**DatabaseConfig.default())
        cursor = conn.cursor()
        cursor.execute("DELETE FROM ir_projects WHERE id = %s", (pid,))
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        logging.error("api_project_delete error: %s", e)
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close(); conn.close()

# ─────────────────────────────────────────────────────────────
# 指標別アノテーション CRUD
# ─────────────────────────────────────────────────────────────

@index_review_bp.route('/api/projects/<int:pid>/indicators/<int:iid>/annotations', methods=['GET'])
@login_required
def api_indicator_annotations_list(pid, iid):
    user_id = session.get('user_id')
    p = _fetch_project(pid)
    if not p or not _can_view_project(user_id, p):
        return jsonify({'success': False, 'error': 'アクセス権がありません'}), 403
    try:
        conn = mysql.connector.connect(**DatabaseConfig.default())
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT id, label, value, color, sort_order
            FROM ir_project_annotations
            WHERE project_id = %s AND indicator_id = %s
            ORDER BY sort_order, id
        """, (pid, iid))
        rows = cursor.fetchall()
        for r in rows:
            r['value'] = float(r['value'])
        return jsonify({'success': True, 'annotations': rows})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close(); conn.close()


@index_review_bp.route('/api/projects/<int:pid>/indicators/<int:iid>/annotations', methods=['POST'])
@login_required
def api_indicator_annotations_create(pid, iid):
    user_id = session.get('user_id')
    p = _fetch_project(pid)
    if not p or not _can_edit_project(user_id, p):
        return jsonify({'success': False, 'error': '権限がありません'}), 403
    data  = request.json or {}
    label = data.get('label', '').strip()
    color = data.get('color', '#ef4444').strip()
    try:
        value = float(data.get('value', ''))
    except (TypeError, ValueError):
        return jsonify({'success': False, 'error': '値は数値で入力してください'}), 400
    if not label:
        return jsonify({'success': False, 'error': 'ラベルは必須です'}), 400
    now = get_jst_now()
    try:
        conn = mysql.connector.connect(**DatabaseConfig.default())
        cursor = conn.cursor()
        # 指標がこのプロジェクトに属するか確認
        cursor.execute(
            "SELECT id FROM ir_project_indicators WHERE id = %s AND project_id = %s",
            (iid, pid)
        )
        if not cursor.fetchone():
            return jsonify({'success': False, 'error': '指標が見つかりません'}), 404
        cursor.execute("""
            INSERT INTO ir_project_annotations
              (project_id, indicator_id, label, value, color, sort_order, created_at)
            VALUES (%s, %s, %s, %s, %s, 0, %s)
        """, (pid, iid, label, value, color, now))
        conn.commit()
        return jsonify({'success': True, 'id': cursor.lastrowid})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close(); conn.close()


@index_review_bp.route('/api/projects/<int:pid>/indicators/<int:iid>/annotations/<int:aid>',
                       methods=['DELETE'])
@login_required
def api_indicator_annotations_delete(pid, iid, aid):
    user_id = session.get('user_id')
    p = _fetch_project(pid)
    if not p or not _can_edit_project(user_id, p):
        return jsonify({'success': False, 'error': '権限がありません'}), 403
    try:
        conn = mysql.connector.connect(**DatabaseConfig.default())
        cursor = conn.cursor()
        cursor.execute(
            "DELETE FROM ir_project_annotations WHERE id=%s AND indicator_id=%s AND project_id=%s",
            (aid, iid, pid)
        )
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close(); conn.close()

@index_review_bp.route('/api/projects/<int:pid>/indicators/<int:iid>/annotations/<int:aid>',
                       methods=['PUT'])
@login_required
def api_indicator_annotations_update(pid, iid, aid):
    user_id = session.get('user_id')
    p = _fetch_project(pid)
    if not p or not _can_edit_project(user_id, p):
        return jsonify({'success': False, 'error': '権限がありません'}), 403
    data  = request.json or {}
    label = data.get('label', '').strip()
    color = data.get('color', '#ef4444').strip()
    try:
        value = float(data.get('value', ''))
    except (TypeError, ValueError):
        return jsonify({'success': False, 'error': '値は数値で入力してください'}), 400
    if not label:
        return jsonify({'success': False, 'error': 'ラベルは必須です'}), 400
    try:
        conn = mysql.connector.connect(**DatabaseConfig.default())
        cursor = conn.cursor()
        cursor.execute(
            """UPDATE ir_project_annotations
               SET label=%s, value=%s, color=%s
               WHERE id=%s AND indicator_id=%s AND project_id=%s""",
            (label, value, color, aid, iid, pid)
        )
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close(); conn.close()

# ─────────────────────────────────────────────────────────────
# プロジェクト別指標 CRUD
# ─────────────────────────────────────────────────────────────

@index_review_bp.route('/api/projects/<int:pid>/indicators', methods=['GET'])
@login_required
def api_project_indicators(pid):
    user_id = session.get('user_id')
    p = _fetch_project(pid)
    if not p or not _can_view_project(user_id, p):
        return jsonify({'success': False, 'error': 'アクセス権がありません'}), 403
    try:
        conn = mysql.connector.connect(**DatabaseConfig.default())
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT id, name, description, value_column, year_column,
                   sort_order, created_at
            FROM ir_project_indicators
            WHERE project_id = %s
            ORDER BY name
        """, (pid,))
        rows = cursor.fetchall()
        for r in rows:
            if r.get('created_at'):
                r['created_at'] = r['created_at'].strftime('%Y-%m-%d')
        return jsonify({
            'success': True,
            'indicators': rows,
            'can_edit': _can_edit_project(user_id, p)
        })
    except Exception as e:
        logging.error("api_project_indicators GET error: %s", e)
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close(); conn.close()


@index_review_bp.route('/api/projects/<int:pid>/indicators', methods=['POST'])
@login_required
def api_project_indicators_create(pid):
    user_id = session.get('user_id')
    p = _fetch_project(pid)
    if not p:
        return jsonify({'success': False, 'error': '見つかりません'}), 404
    if not _can_edit_project(user_id, p):
        return jsonify({'success': False, 'error': '権限がありません'}), 403

    data = request.json or {}
    name         = data.get('name', '').strip()
    sql_query    = data.get('sql_query', '').strip()
    value_column = data.get('value_column', '').strip()
    year_column  = data.get('year_column', '年度').strip()
    description  = data.get('description', '').strip()

    if not name or not sql_query or not value_column:
        return jsonify({'success': False, 'error': '指標名・SQLクエリ・値カラムは必須です'}), 400
    if not sql_query.upper().startswith('SELECT'):
        return jsonify({'success': False, 'error': 'SELECT文のみ登録できます'}), 400

    now = get_jst_now()
    try:
        conn = mysql.connector.connect(**DatabaseConfig.default())
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO ir_project_indicators
              (project_id, name, sql_query, value_column, year_column,
               description, sort_order, created_at, updated_at)
            VALUES (%s, %s, %s, %s, %s, %s, 0, %s, %s)
        """, (pid, name, sql_query, value_column, year_column, description, now, now))
        conn.commit()
        return jsonify({'success': True, 'id': cursor.lastrowid})
    except Exception as e:
        logging.error("api_project_indicators_create error: %s", e)
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close(); conn.close()


@index_review_bp.route('/api/projects/<int:pid>/indicators/<int:iid>', methods=['GET'])
@login_required
def api_project_indicator_get(pid, iid):
    user_id = session.get('user_id')
    p = _fetch_project(pid)
    if not p or not _can_view_project(user_id, p):
        return jsonify({'success': False, 'error': 'アクセス権がありません'}), 403
    try:
        conn = mysql.connector.connect(**DatabaseConfig.default())
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT id, name, sql_query, value_column, year_column, description
            FROM ir_project_indicators
            WHERE id = %s AND project_id = %s
        """, (iid, pid))
        row = cursor.fetchone()
        if not row:
            return jsonify({'success': False, 'error': '見つかりません'}), 404
        return jsonify({'success': True, 'indicator': row})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close(); conn.close()


@index_review_bp.route('/api/projects/<int:pid>/indicators/<int:iid>', methods=['PUT'])
@login_required
def api_project_indicator_update(pid, iid):
    user_id = session.get('user_id')
    p = _fetch_project(pid)
    if not p or not _can_edit_project(user_id, p):
        return jsonify({'success': False, 'error': '権限がありません'}), 403

    data = request.json or {}
    name         = data.get('name', '').strip()
    sql_query    = data.get('sql_query', '').strip()
    value_column = data.get('value_column', '').strip()
    year_column  = data.get('year_column', '年度').strip()
    description  = data.get('description', '').strip()

    if not name or not sql_query or not value_column:
        return jsonify({'success': False, 'error': '必須項目が不足しています'}), 400
    if not sql_query.upper().startswith('SELECT'):
        return jsonify({'success': False, 'error': 'SELECT文のみ登録できます'}), 400

    now = get_jst_now()
    try:
        conn = mysql.connector.connect(**DatabaseConfig.default())
        cursor = conn.cursor(dictionary=True)
        # 名前重複チェック（同プロジェクト内、自分以外）
        cursor.execute("""
            SELECT id FROM ir_project_indicators
            WHERE project_id = %s AND name = %s AND id != %s
        """, (pid, name, iid))
        if cursor.fetchone():
            return jsonify({'success': False, 'error': f'指標名「{name}」はこのプロジェクトで既に使用中です'}), 400

        cursor.execute("""
            UPDATE ir_project_indicators
            SET name=%s, sql_query=%s, value_column=%s, year_column=%s,
                description=%s, updated_at=%s
            WHERE id=%s AND project_id=%s
        """, (name, sql_query, value_column, year_column, description, now, iid, pid))
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        logging.error("api_project_indicator_update error: %s", e)
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close(); conn.close()


@index_review_bp.route('/api/projects/<int:pid>/indicators/<int:iid>', methods=['DELETE'])
@login_required
def api_project_indicator_delete(pid, iid):
    user_id = session.get('user_id')
    p = _fetch_project(pid)
    if not p or not _can_edit_project(user_id, p):
        return jsonify({'success': False, 'error': '権限がありません'}), 403
    try:
        conn = mysql.connector.connect(**DatabaseConfig.default())
        cursor = conn.cursor()
        cursor.execute(
            "DELETE FROM ir_project_indicators WHERE id = %s AND project_id = %s",
            (iid, pid)
        )
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close(); conn.close()


@index_review_bp.route('/api/projects/<int:pid>/indicators/<int:iid>/data', methods=['GET'])
@login_required
def api_project_indicator_data(pid, iid):
    """指標の時系列データを返す（SQLを fujinp DB で実行）"""
    user_id = session.get('user_id')
    p = _fetch_project(pid)
    if not p or not _can_view_project(user_id, p):
        return jsonify({'success': False, 'error': 'アクセス権がありません'}), 403

    try:
        # 指標定義を default DB から取得
        conn_d = mysql.connector.connect(**DatabaseConfig.default())
        cur_d  = conn_d.cursor(dictionary=True)
        cur_d.execute("""
            SELECT sql_query, value_column, year_column
            FROM ir_project_indicators WHERE id = %s AND project_id = %s
        """, (iid, pid))
        ind = cur_d.fetchone()
        cur_d.close(); conn_d.close()

        if not ind:
            return jsonify({'success': False, 'error': '指標が見つかりません'}), 404

        sql = ind['sql_query'].strip()
        if not sql.upper().startswith('SELECT'):
            return jsonify({'success': False, 'error': 'SELECT文のみ実行できます'}), 400

        # fujinp DB で実行
        conn_f = mysql.connector.connect(**DatabaseConfig.fujinp())
        cur_f  = conn_f.cursor(dictionary=True)
        cur_f.execute(sql)
        rows = cur_f.fetchall()
        cur_f.close(); conn_f.close()

        year_col = ind['year_column']
        val_col  = ind['value_column']
        data = [
            {'年度': int(r[year_col]), '値': float(r[val_col])}
            for r in rows
            if r.get(year_col) is not None and r.get(val_col) is not None
        ]
        return jsonify({'success': True, 'data': data, 'unit': ''})

    except Exception as e:
        logging.error("api_project_indicator_data error [pid=%s, iid=%s]: %s", pid, iid, e)
        return jsonify({'success': False, 'error': str(e)}), 500


# ─────────────────────────────────────────────────────────────
# エクスポート / インポート
# ─────────────────────────────────────────────────────────────

@index_review_bp.route('/api/projects/<int:pid>/export', methods=['GET'])
@login_required
def api_project_export(pid):
    """プロジェクト内の全指標を JSON でエクスポート"""
    user_id = session.get('user_id')
    p = _fetch_project(pid)
    if not p or not _can_view_project(user_id, p):
        return jsonify({'success': False, 'error': 'アクセス権がありません'}), 403
    try:
        conn = mysql.connector.connect(**DatabaseConfig.default())
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT name, sql_query, value_column, year_column, description
            FROM ir_project_indicators
            WHERE project_id = %s
            ORDER BY name
        """, (pid,))
        rows = cursor.fetchall()
        return jsonify({'success': True, 'indicators': rows,
                        'project_name': p['name']})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close(); conn.close()


@index_review_bp.route('/api/projects/<int:pid>/import', methods=['POST'])
@login_required
def api_project_import(pid):
    """JSON から指標を一括インポート（編集権限が必要）"""
    user_id = session.get('user_id')
    p = _fetch_project(pid)
    if not p:
        return jsonify({'success': False, 'error': '見つかりません'}), 404
    if not _can_edit_project(user_id, p):
        return jsonify({'success': False, 'error': '権限がありません'}), 403

    data = request.json or {}
    indicators = data.get('indicators', [])
    if not isinstance(indicators, list) or not indicators:
        return jsonify({'success': False, 'error': '指標データが空です'}), 400

    now = get_jst_now()
    imported = 0
    skipped  = 0
    errors   = []

    try:
        conn = mysql.connector.connect(**DatabaseConfig.default())
        cursor = conn.cursor(dictionary=True)

        # 既存名を取得（重複スキップ用）
        cursor.execute(
            "SELECT name FROM ir_project_indicators WHERE project_id = %s", (pid,))
        existing_names = {r['name'] for r in cursor.fetchall()}

        for ind in indicators:
            name         = str(ind.get('name', '')).strip()
            sql_query    = str(ind.get('sql_query', '')).strip()
            value_column = str(ind.get('value_column', '')).strip()
            year_column  = str(ind.get('year_column', '年度')).strip() or '年度'
            description  = str(ind.get('description', '')).strip()

            if not name or not sql_query or not value_column:
                errors.append(f'スキップ（必須項目不足）: {name or "名称なし"}')
                skipped += 1
                continue
            if not sql_query.upper().startswith('SELECT'):
                errors.append(f'スキップ（SELECT文以外）: {name}')
                skipped += 1
                continue
            if name in existing_names:
                errors.append(f'スキップ（名前重複）: {name}')
                skipped += 1
                continue

            cursor.execute("""
                INSERT INTO ir_project_indicators
                  (project_id, name, sql_query, value_column, year_column,
                   description, sort_order, created_at, updated_at)
                VALUES (%s, %s, %s, %s, %s, %s, 0, %s, %s)
            """, (pid, name, sql_query, value_column, year_column,
                  description, now, now))
            existing_names.add(name)
            imported += 1

        conn.commit()
        return jsonify({
            'success': True,
            'imported': imported,
            'skipped': skipped,
            'messages': errors
        })
    except Exception as e:
        logging.error("api_project_import error: %s", e)
        if 'conn' in locals():
            conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close(); conn.close()


# ─────────────────────────────────────────────────────────────
# SQL テスト・テーブル一覧（変更なし）
# ─────────────────────────────────────────────────────────────

@index_review_bp.route('/api/sql_test', methods=['POST'])
@login_required
def api_sql_test():
    data = request.json or {}
    sql = data.get('sql', '').strip()
    if not sql:
        return jsonify({'success': False, 'error': 'SQLを入力してください'}), 400
    if not sql.upper().startswith('SELECT'):
        return jsonify({'success': False, 'error': 'SELECT文のみ実行できます'}), 400
    try:
        conn = mysql.connector.connect(**DatabaseConfig.fujinp())
        cursor = conn.cursor(dictionary=True)
        cursor.execute(sql)
        columns = list(cursor.column_names) if cursor.column_names else []
        raw_rows = cursor.fetchmany(20)
        serialized = []
        for r in raw_rows:
            row_out = {}
            for k, v in r.items():
                if isinstance(v, (datetime.date, datetime.datetime)):
                    row_out[k] = str(v)
                elif isinstance(v, bytes):
                    row_out[k] = v.decode('utf-8', errors='replace')
                elif v is None:
                    row_out[k] = None
                else:
                    try:
                        row_out[k] = float(v)
                    except (TypeError, ValueError):
                        row_out[k] = v
            serialized.append(row_out)
        return jsonify({'success': True, 'columns': columns,
                        'rows': serialized, 'row_count': len(serialized)})
    except Exception as e:
        logging.error("api_sql_test error: %s", e)
        return jsonify({'success': False, 'error': str(e)}), 200
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close(); conn.close()


@index_review_bp.route('/api/table_list')
@login_required
def api_table_list():
    try:
        conn = mysql.connector.connect(**DatabaseConfig.fujinp())
        cursor = conn.cursor()
        cursor.execute("SHOW TABLES")
        all_tables = [row[0] for row in cursor.fetchall()]
        t_tables = sorted(t for t in all_tables if t.upper().startswith('T_'))
        return jsonify({'success': True, 'tables': t_tables})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close(); conn.close()


# ─────────────────────────────────────────────────────────────
# テーブルブラウザ（変更なし）
# ─────────────────────────────────────────────────────────────

@index_review_bp.route('/api/browser/tables')
@login_required
def api_browser_tables():
    result = []
    try:
        conn = mysql.connector.connect(**DatabaseConfig.base())
        cursor = conn.cursor()
        user_prefix = Config.DB_ACCOUNT + "$"
        cursor.execute(
            "SELECT SCHEMA_NAME FROM INFORMATION_SCHEMA.SCHEMATA WHERE SCHEMA_NAME LIKE %s",
            (user_prefix + '%',))
        databases = [r[0] for r in cursor.fetchall()]
        cursor.close(); conn.close()
        for db in databases:
            try:
                db_conn = mysql.connector.connect(**DatabaseConfig.get_config(db))
                db_cursor = db_conn.cursor()
                db_cursor.execute("SHOW TABLES")
                for (tbl,) in db_cursor.fetchall():
                    try:
                        db_cursor.execute(f"SELECT COUNT(*) FROM `{tbl}`")
                        cnt = db_cursor.fetchone()[0]
                    except Exception:
                        cnt = -1
                    result.append({'database': db, 'table_name': tbl, 'row_count': cnt})
                db_cursor.close(); db_conn.close()
            except Exception as e:
                logging.warning("api_browser_tables: DB %s error: %s", db, e)
        return jsonify({'success': True, 'tables': result})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@index_review_bp.route('/api/browser/projects')
@login_required
def api_browser_projects():
    try:
        conn = mysql.connector.connect(**DatabaseConfig.default())
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT p.id, p.name, p.description, p.created_at, p.updated_at,
                   p.access_policy,
                   COUNT(DISTINCT pt.id) AS table_count,
                   COUNT(DISTINCT pv.id) AS view_count
            FROM table_master_projects p
            LEFT JOIN table_master_project_tables pt ON pt.project_id = p.id
            LEFT JOIN table_master_project_views  pv ON pv.project_id = p.id
            GROUP BY p.id ORDER BY p.name
        """)
        rows = cursor.fetchall()
        for r in rows:
            for k in ('created_at', 'updated_at'):
                if r.get(k): r[k] = r[k].isoformat()
        return jsonify({'success': True, 'projects': rows})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close(); conn.close()


@index_review_bp.route('/api/browser/project_tables')
@login_required
def api_browser_project_tables():
    project_id = request.args.get('project_id', '').strip()
    if not project_id:
        return jsonify({'success': False, 'error': 'project_id が必要です'}), 400
    try:
        conn = mysql.connector.connect(**DatabaseConfig.default())
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT id, database_name, table_name, added_at
            FROM table_master_project_tables WHERE project_id = %s
            ORDER BY database_name, table_name
        """, (project_id,))
        rows = cursor.fetchall()
        for r in rows:
            if r.get('added_at'): r['added_at'] = r['added_at'].isoformat()
            try:
                db_conn = mysql.connector.connect(**DatabaseConfig.get_config(r['database_name']))
                db_cur  = db_conn.cursor()
                db_cur.execute(f"SELECT COUNT(*) FROM `{r['table_name']}`")
                r['row_count'] = db_cur.fetchone()[0]
                db_cur.close(); db_conn.close()
            except Exception:
                r['row_count'] = -1
        return jsonify({'success': True, 'tables': rows})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close(); conn.close()


@index_review_bp.route('/api/browser/schema', methods=['POST'])
@login_required
def api_browser_schema():
    data = request.json or {}
    database   = data.get('database', '').strip()
    table_name = data.get('table_name', '').strip()
    if not database or not table_name:
        return jsonify({'success': False, 'error': 'パラメータ不足'}), 400
    for v in (database, table_name):
        if any(c in v for c in ('`', ';', "'", '"', '\x00')):
            return jsonify({'success': False, 'error': '不正な識別子'}), 400
    try:
        conn = mysql.connector.connect(**DatabaseConfig.get_config(database))
        cursor = conn.cursor()
        cursor.execute("SHOW TABLES LIKE %s", (table_name,))
        if not cursor.fetchone():
            return jsonify({'success': False, 'error': 'テーブルが見つかりません'}), 404
        cursor.execute(f"SHOW CREATE TABLE `{table_name}`")
        row = cursor.fetchone()
        return jsonify({'success': True, 'schema': row[1] if row else ''})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close(); conn.close()


@index_review_bp.route('/api/browser/data', methods=['POST'])
@login_required
def api_browser_data():
    data = request.json or {}
    database   = data.get('database', '').strip()
    table_name = data.get('table_name', '').strip()
    try:
        limit  = min(max(int(data.get('limit', 500)), 1), 500)
        offset = max(int(data.get('offset', 0)), 0)
    except (TypeError, ValueError):
        return jsonify({'success': False, 'error': 'limit と offset は整数で指定してください'}), 400
    if not database or not table_name:
        return jsonify({'success': False, 'error': 'パラメータ不足'}), 400
    for v in (database, table_name):
        if any(c in v for c in ('`', ';', "'", '"', '\x00')):
            return jsonify({'success': False, 'error': '不正な識別子'}), 400
    try:
        conn = mysql.connector.connect(**DatabaseConfig.get_config(database))
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SHOW TABLES LIKE %s", (table_name,))
        if not cursor.fetchone():
            return jsonify({'success': False, 'error': 'テーブルが見つかりません'}), 404
        cursor.execute(f"SELECT COUNT(*) as cnt FROM `{table_name}`")
        total = cursor.fetchone()['cnt']
        cursor.execute(f"SELECT * FROM `{table_name}` LIMIT %s OFFSET %s", (limit, offset))
        rows = cursor.fetchall()
        columns = list(cursor.column_names) if cursor.column_names else []

        def safe(v):
            if v is None: return None
            if isinstance(v, (datetime.datetime, datetime.date)): return v.isoformat()
            if isinstance(v, datetime.timedelta):
                s = int(v.total_seconds())
                return f"{s//3600:02d}:{(s%3600)//60:02d}"
            if isinstance(v, bytes): return v.decode('utf-8', errors='replace')
            return v

        serialized = [{k: safe(val) for k, val in row.items()} for row in rows]
        return jsonify({'success': True, 'columns': columns, 'rows': serialized,
                        'total_count': total, 'fetched_count': len(rows)})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close(); conn.close()


# ─────────────────────────────────────────────────────────────
# スキーマダンプ（変更なし）
# ─────────────────────────────────────────────────────────────

@index_review_bp.route('/api/schema_dump')
@login_required
def api_schema_dump():
    SAMPLE_ROWS = 3
    MAX_SAMPLE_LEN = 30
    def fmt(v):
        if v is None: return 'NULL'
        s = str(v)
        if isinstance(v, (datetime.date, datetime.datetime)): s = v.isoformat()
        return s if len(s) <= MAX_SAMPLE_LEN else s[:MAX_SAMPLE_LEN] + '…'

    table_set_id = request.args.get('table_set_id', type=int)

    try:
        if table_set_id:
            # ★ テーブル集スコープ：ir_table_set_items から対象テーブルを取得
            conn_def = mysql.connector.connect(**DatabaseConfig.default())
            cur_def  = conn_def.cursor(dictionary=True)
            cur_def.execute("""
                SELECT database_name, table_name
                FROM ir_table_set_items
                WHERE table_set_id = %s
                ORDER BY database_name, table_name
            """, (table_set_id,))
            items = cur_def.fetchall()
            cur_def.close(); conn_def.close()

            lines = []
            for item in items:
                db_name = item['database_name']
                tbl     = item['table_name']
                try:
                    conn2   = mysql.connector.connect(**DatabaseConfig.get_config(db_name))
                    cursor2 = conn2.cursor()
                    cursor2.execute(f"SELECT COUNT(*) FROM `{tbl}`")
                    row_count = cursor2.fetchone()[0]
                    cursor2.execute(f"SHOW COLUMNS FROM `{tbl}`")
                    columns = cursor2.fetchall()
                    cursor2.execute(f"SELECT * FROM `{tbl}` LIMIT {SAMPLE_ROWS}")
                    sample_rows = cursor2.fetchall()
                    cursor2.close(); conn2.close()

                    lines.append(f"### {tbl}（database: {db_name}、{row_count}行）")
                    lines.append("| カラム | 型 | サンプル値 |")
                    lines.append("|--------|-----|------------|")
                    for i, col in enumerate(columns):
                        sample_vals = [fmt(row[i]) for row in sample_rows if i < len(row)]
                        lines.append(f"| {col[0]} | {col[1]} | {', '.join(sample_vals) or '―'} |")
                    lines.append("")
                except Exception as e:
                    lines.append(f"### {tbl} ※取得エラー: {e}\n")

            return jsonify({'success': True, 'text': '\n'.join(lines), 'table_count': len(items)})

        else:
            # ★ 既存ロジック（変更なし）
            conn = mysql.connector.connect(**DatabaseConfig.fujinp())
            cursor = conn.cursor()
            cursor.execute("SHOW TABLES")
            all_tables = [r[0] for r in cursor.fetchall()]
            t_tables = sorted(t for t in all_tables if t.upper().startswith('T_'))
            lines = []
            db_name = DatabaseConfig.fujinp().get('database', 'fujinp')
            for tbl in t_tables:
                try:
                    cursor.execute(f"SELECT COUNT(*) FROM `{tbl}`")
                    row_count = cursor.fetchone()[0]
                    cursor.execute(f"SHOW COLUMNS FROM `{tbl}`")
                    columns = cursor.fetchall()
                    cursor.execute(f"SELECT * FROM `{tbl}` LIMIT {SAMPLE_ROWS}")
                    sample_rows = cursor.fetchall()
                    lines.append(f"### {tbl}（database: {db_name}、{row_count}行）")
                    lines.append("| カラム | 型 | サンプル値 |")
                    lines.append("|--------|-----|------------|")
                    for i, col in enumerate(columns):
                        sample_vals = [fmt(row[i]) for row in sample_rows if i < len(row)]
                        lines.append(f"| {col[0]} | {col[1]} | {', '.join(sample_vals) or '―'} |")
                    lines.append("")
                except Exception as e:
                    lines.append(f"### {tbl} ※取得エラー: {e}\n")
            return jsonify({'success': True, 'text': '\n'.join(lines), 'table_count': len(t_tables)})

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close(); conn.close()


# ─────────────────────────────────────────────────────────────
# テーブル集管理
# ─────────────────────────────────────────────────────────────

def _check_table_set_admin(user_id):
    if not user_id:
        return False
    try:
        conn = mysql.connector.connect(**DatabaseConfig.default())
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT category FROM users WHERE id = %s", (user_id,))
        row = cursor.fetchone()
        if row and row['category'] == 'admin':
            return True
        cursor.execute("SELECT id FROM user_groups WHERE name = %s", (TABLE_SET_ADMIN_GROUP,))
        grp = cursor.fetchone()
        if not grp:
            return False
        now = get_jst_now()
        cursor.execute("""
            SELECT id FROM user_group_memberships
            WHERE group_id = %s AND user_id = %s
              AND (valid_from IS NULL OR valid_from <= %s)
              AND (valid_until IS NULL OR valid_until >= %s)
        """, (grp['id'], user_id, now, now))
        return cursor.fetchone() is not None
    except Exception as e:
        logging.error("_check_table_set_admin error: %s", e)
        return False
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close(); conn.close()


@index_review_bp.route('/api/table_sets')
@login_required
def api_table_sets_list():
    try:
        conn = mysql.connector.connect(**DatabaseConfig.default())
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT ts.id, ts.name, ts.description, ts.owner_user_id,
                   ts.created_at, ts.updated_at, COUNT(ti.id) AS item_count
            FROM ir_table_sets ts
            LEFT JOIN ir_table_set_items ti ON ti.table_set_id = ts.id
            GROUP BY ts.id ORDER BY ts.id
        """)
        rows = cursor.fetchall()
        for r in rows:
            for k in ('created_at', 'updated_at'):
                if r.get(k): r[k] = r[k].strftime('%Y-%m-%d %H:%M')
        user_id  = session.get('user_id')
        can_administer = _check_table_set_admin(user_id)
        if rows:
            owner_ids = list({r['owner_user_id'] for r in rows if r['owner_user_id']})
            conn2 = mysql.connector.connect(**DatabaseConfig.default())
            cur2  = conn2.cursor(dictionary=True)
            if owner_ids:
                fmt = ','.join(['%s'] * len(owner_ids))
                cur2.execute(f"SELECT id, full_name FROM users WHERE id IN ({fmt})", owner_ids)
                name_map = {u['id']: u['full_name'] for u in cur2.fetchall()}
            else:
                name_map = {}
            cur2.close(); conn2.close()
            for r in rows:
                r['owner_name'] = name_map.get(r['owner_user_id'], '不明')
                r['can_edit']   = can_administer or (r['owner_user_id'] == user_id)
        return jsonify({'success': True, 'table_sets': rows, 'is_admin': can_administer,
                        'current_user_id': user_id})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close(); conn.close()


@index_review_bp.route('/api/table_sets', methods=['POST'])
@login_required
def api_table_sets_create():
    user_id = session.get('user_id')
    if not _check_table_set_admin(user_id):
        return jsonify({'success': False, 'error': '権限がありません'}), 403
    data = request.json or {}
    name = data.get('name', '').strip()
    if not name:
        return jsonify({'success': False, 'error': 'テーブル集名は必須です'}), 400
    now = get_jst_now()
    try:
        conn = mysql.connector.connect(**DatabaseConfig.default())
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO ir_table_sets (name, description, owner_user_id, created_at, updated_at)
            VALUES (%s, %s, %s, %s, %s)
        """, (name, data.get('description', ''), user_id, now, now))
        conn.commit()
        return jsonify({'success': True, 'id': cursor.lastrowid})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close(); conn.close()


@index_review_bp.route('/api/table_sets/<int:set_id>', methods=['PUT'])
@login_required
def api_table_sets_update(set_id):
    user_id = session.get('user_id')
    data = request.json or {}
    name = data.get('name', '').strip()
    if not name:
        return jsonify({'success': False, 'error': 'テーブル集名は必須です'}), 400
    try:
        conn = mysql.connector.connect(**DatabaseConfig.default())
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT owner_user_id FROM ir_table_sets WHERE id = %s", (set_id,))
        row = cursor.fetchone()
        if not row:
            return jsonify({'success': False, 'error': '見つかりません'}), 404
        if not (_check_table_set_admin(user_id) or row['owner_user_id'] == user_id):
            return jsonify({'success': False, 'error': '権限がありません'}), 403
        now = get_jst_now()
        cursor.execute("""
            UPDATE ir_table_sets SET name=%s, description=%s, updated_at=%s WHERE id=%s
        """, (name, data.get('description', ''), now, set_id))
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close(); conn.close()


@index_review_bp.route('/api/table_sets/<int:set_id>', methods=['DELETE'])
@login_required
def api_table_sets_delete(set_id):
    user_id = session.get('user_id')
    try:
        conn = mysql.connector.connect(**DatabaseConfig.default())
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT owner_user_id FROM ir_table_sets WHERE id = %s", (set_id,))
        row = cursor.fetchone()
        if not row:
            return jsonify({'success': False, 'error': '見つかりません'}), 404
        if not (_check_table_set_admin(user_id) or row['owner_user_id'] == user_id):
            return jsonify({'success': False, 'error': '権限がありません'}), 403
        cursor.execute("DELETE FROM ir_table_sets WHERE id = %s", (set_id,))
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close(); conn.close()


@index_review_bp.route('/api/table_sets/<int:set_id>/items')
@login_required
def api_table_set_items_list(set_id):
    try:
        user_id = session.get('user_id')
        conn = mysql.connector.connect(**DatabaseConfig.default())
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT id, database_name, table_name, manager_group_id, added_at
            FROM ir_table_set_items WHERE table_set_id = %s
            ORDER BY database_name, table_name
        """, (set_id,))
        rows = cursor.fetchall()
        for r in rows:
            if r.get('added_at'): r['added_at'] = r['added_at'].strftime('%Y-%m-%d %H:%M')
        group_ids = list({r['manager_group_id'] for r in rows if r['manager_group_id']})
        if group_ids:
            fmt = ','.join(['%s'] * len(group_ids))
            cursor.execute(f"SELECT id, name FROM user_groups WHERE id IN ({fmt})", group_ids)
            grp_map = {g['id']: g['name'] for g in cursor.fetchall()}
        else:
            grp_map = {}
        can_administer = _check_table_set_admin(user_id)
        cursor.execute("SELECT owner_user_id FROM ir_table_sets WHERE id = %s", (set_id,))
        ts = cursor.fetchone()
        is_owner = bool(ts and ts['owner_user_id'] == user_id)
        for r in rows:
            r['manager_group_name'] = grp_map.get(r['manager_group_id'], '―')
            r['can_update'] = (
                can_administer
                or _is_group_member(user_id, r['manager_group_id'])
            )
        can_edit = bool(ts) and (can_administer or is_owner)
        return jsonify({'success': True, 'items': rows, 'can_edit': can_edit})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close(); conn.close()


@index_review_bp.route('/api/table_sets/<int:set_id>/items', methods=['POST'])
@login_required
def api_table_set_items_add(set_id):
    user_id = session.get('user_id')
    data = request.json or {}
    db_name  = data.get('database_name', '').strip()
    tbl_name = data.get('table_name', '').strip()
    if not db_name or not tbl_name:
        return jsonify({'success': False, 'error': 'database_name と table_name は必須'}), 400
    try:
        conn = mysql.connector.connect(**DatabaseConfig.default())
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT owner_user_id FROM ir_table_sets WHERE id = %s", (set_id,))
        ts = cursor.fetchone()
        if not ts:
            return jsonify({'success': False, 'error': '見つかりません'}), 404
        if not (_check_table_set_admin(user_id) or ts['owner_user_id'] == user_id):
            return jsonify({'success': False, 'error': '権限がありません'}), 403
        now = get_jst_now()
        cursor.execute("""
            INSERT INTO ir_table_set_items
              (table_set_id, database_name, table_name, manager_group_id, added_at)
            VALUES (%s, %s, %s, %s, %s)
        """, (set_id, db_name, tbl_name, data.get('manager_group_id'), now))
        conn.commit()
        return jsonify({'success': True, 'id': cursor.lastrowid})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close(); conn.close()


@index_review_bp.route('/api/table_sets/<int:set_id>/items/<int:item_id>', methods=['PUT'])
@login_required
def api_table_set_items_update(set_id, item_id):
    user_id = session.get('user_id')
    data = request.json or {}
    try:
        conn = mysql.connector.connect(**DatabaseConfig.default())
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT owner_user_id FROM ir_table_sets WHERE id = %s", (set_id,))
        ts = cursor.fetchone()
        if not ts:
            return jsonify({'success': False, 'error': '見つかりません'}), 404
        if not (_check_table_set_admin(user_id) or ts['owner_user_id'] == user_id):
            return jsonify({'success': False, 'error': '権限がありません'}), 403
        cursor.execute("""
            UPDATE ir_table_set_items SET manager_group_id=%s WHERE id=%s AND table_set_id=%s
        """, (data.get('manager_group_id') or None, item_id, set_id))
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close(); conn.close()


@index_review_bp.route('/api/table_sets/<int:set_id>/items/<int:item_id>', methods=['DELETE'])
@login_required
def api_table_set_items_delete(set_id, item_id):
    user_id = session.get('user_id')
    try:
        conn = mysql.connector.connect(**DatabaseConfig.default())
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT owner_user_id FROM ir_table_sets WHERE id = %s", (set_id,))
        ts = cursor.fetchone()
        if not ts:
            return jsonify({'success': False, 'error': '見つかりません'}), 404
        if not (_check_table_set_admin(user_id) or ts['owner_user_id'] == user_id):
            return jsonify({'success': False, 'error': '権限がありません'}), 403
        cursor.execute("DELETE FROM ir_table_set_items WHERE id=%s AND table_set_id=%s",
                       (item_id, set_id))
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close(); conn.close()


@index_review_bp.route('/api/table_sets/groups')
@login_required
def api_table_sets_groups():
    user_id = session.get('user_id')
    if not _check_table_set_admin(user_id):
        return jsonify({'success': False, 'error': '権限がありません'}), 403
    try:
        conn = mysql.connector.connect(**DatabaseConfig.default())
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT id, name, description FROM user_groups ORDER BY name")
        groups = cursor.fetchall()
        return jsonify({'success': True, 'groups': groups})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close(); conn.close()


# ─────────────────────────────────────────────────────────────
# テーブルデータ更新
# ─────────────────────────────────────────────────────────────

@index_review_bp.route('/table_data_update/<int:item_id>')
@login_required
def table_data_update(item_id):
    user_id = session.get('user_id')
    try:
        conn = mysql.connector.connect(**DatabaseConfig.default())
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT i.id, i.database_name, i.table_name, i.manager_group_id,
                   ts.name AS set_name
            FROM ir_table_set_items i
            JOIN ir_table_sets ts ON ts.id = i.table_set_id
            WHERE i.id = %s
        """, (item_id,))
        item = cursor.fetchone()
        if not item:
            return "テーブルが見つかりません", 404
        if not (_check_table_set_admin(user_id) or
                _is_group_member(user_id, item['manager_group_id'])):
            return "権限がありません", 403
        return render_template('index_review/table_data_update.html', item=item)
    except Exception as e:
        logging.error("table_data_update error: %s", e)
        return "エラーが発生しました", 500
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close(); conn.close()


@index_review_bp.route('/table_data_update/<int:item_id>/download')
@login_required
def table_data_update_download(item_id):
    user_id = session.get('user_id')
    try:
        conn = mysql.connector.connect(**DatabaseConfig.default())
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT i.database_name, i.table_name, i.manager_group_id
            FROM ir_table_set_items i WHERE i.id = %s
        """, (item_id,))
        item = cursor.fetchone()
        cursor.close(); conn.close()
        if not item:
            return "テーブルが見つかりません", 404
        if not (_check_table_set_admin(user_id) or
                _is_group_member(user_id, item['manager_group_id'])):
            return "権限がありません", 403

        database   = item['database_name']
        table_name = item['table_name']
        conn2  = mysql.connector.connect(**DatabaseConfig.get_config(database))
        cur2   = conn2.cursor(dictionary=True)
        cur2.execute(f"DESCRIBE `{table_name}`")
        cols_info    = cur2.fetchall()
        time_columns = {c['Field'] for c in cols_info if c['Type'].lower().startswith('time')}
        cur2.execute(f"SELECT * FROM `{table_name}`")
        rows = cur2.fetchall()
        cur2.close(); conn2.close()

        # 0行のテーブルでも、ダウンロードExcelに実テーブルのヘッダーを残す。
        df = pd.DataFrame(rows, columns=[c['Field'] for c in cols_info])
        for col in time_columns:
            if col in df.columns:
                df[col] = df[col].apply(lambda x:
                    x.total_seconds() / 86400
                    if isinstance(x, (datetime.timedelta, pd.Timedelta))
                    else ((x.hour*3600 + x.minute*60 + x.second) / 86400
                          if isinstance(x, datetime.time) else x))

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=table_name[:31], index=False)
            from openpyxl.utils import get_column_letter
            ws = writer.sheets[table_name[:31]]
            for idx, col_name in enumerate(df.columns):
                if col_name in time_columns:
                    col_letter = get_column_letter(idx + 1)
                    for cell in ws[col_letter]:
                        if cell.row > 1:
                            cell.number_format = 'h:mm:ss'
        output.seek(0)
        return send_file(output, as_attachment=True,
                         download_name=f"{table_name}_download.xlsx",
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        logging.error("table_data_update_download error: %s", e)
        return str(e), 500


@index_review_bp.route('/table_data_update/<int:item_id>/upload', methods=['POST'])
@login_required
def table_data_update_upload(item_id):
    user_id = session.get('user_id')
    try:
        conn = mysql.connector.connect(**DatabaseConfig.default())
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT i.database_name, i.table_name, i.manager_group_id
            FROM ir_table_set_items i WHERE i.id = %s
        """, (item_id,))
        item = cursor.fetchone()
        cursor.close(); conn.close()
        if not item:
            return jsonify({'success': False, 'error': 'テーブルが見つかりません'}), 404
        if not (_check_table_set_admin(user_id) or
                _is_group_member(user_id, item['manager_group_id'])):
            return jsonify({'success': False, 'error': '権限がありません'}), 403

        database   = item['database_name']
        table_name = item['table_name']
        file = request.files.get('excel_file')
        if not file:
            return jsonify({'success': False, 'error': 'ファイルが指定されていません'}), 400

        df = pd.read_excel(file, engine='openpyxl')
        for col in df.columns:
            if pd.api.types.is_datetime64_any_dtype(df[col]):
                df[col] = df[col].dt.strftime('%Y-%m-%d %H:%M:%S').replace('NaT', None)

        def convert_value(v):
            try:
                if pd.isna(v): return None
            except Exception:
                pass
            if hasattr(v, 'item'): return v.item()
            return v

        conn2  = mysql.connector.connect(**DatabaseConfig.get_config(database))
        cursor2 = conn2.cursor()
        try:
            cursor2.execute(f"SHOW COLUMNS FROM `{table_name}`")
            expected_columns = [row[0] for row in cursor2.fetchall()]
            uploaded_columns = [str(column) for column in df.columns]
            if uploaded_columns != expected_columns:
                return jsonify({
                    'success': False,
                    'error': (
                        'Excelのカラム構成が対象テーブルと一致しません。'
                        '最新のダウンロードファイルを編集してアップロードしてください。'
                    ),
                    'expected_columns': expected_columns,
                    'uploaded_columns': uploaded_columns,
                }), 400

            cursor2.execute("SET FOREIGN_KEY_CHECKS = 0")
            conn2.start_transaction()
            cursor2.execute(f"DELETE FROM `{table_name}`")
            if not df.empty:
                placeholders = ', '.join(['%s'] * len(df.columns))
                columns      = '`, `'.join(str(c) for c in df.columns)
                insert_query = f"INSERT INTO `{table_name}` (`{columns}`) VALUES ({placeholders})"
                values       = [tuple(convert_value(v) for v in row) for row in df.values]
                cursor2.executemany(insert_query, values)
            cursor2.execute("SET FOREIGN_KEY_CHECKS = 1")
            conn2.commit()
            return jsonify({'success': True, 'message': f"{len(df)} 行をインポートしました。"})
        except Exception as e:
            conn2.rollback()
            cursor2.execute("SET FOREIGN_KEY_CHECKS = 1")
            return jsonify({'success': False, 'error': str(e)}), 500
    except Exception as e:
        logging.error("table_data_update_upload error: %s", e)
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if 'conn2' in locals() and conn2.is_connected():
            cursor2.close(); conn2.close()

@index_review_bp.route('/return_to_fujin')
@login_required
def return_to_fujin():
    """FUJIN-Pダッシュボードに戻る"""
    return redirect_to_dashboard()
