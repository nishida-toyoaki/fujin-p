# SPDX-FileCopyrightText: 2024-2026 Toyoaki Nishida
# SPDX-License-Identifier: AGPL-3.0-or-later
#
# This file is part of FUJIN-P.
# Copyright (C) 2024-2026 Toyoaki Nishida
#
# FUJIN-P is free software: you can redistribute it and/or modify
# it under the terms of the GNU Affero General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# FUJIN-P is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU Affero General Public License for more details.
#
# You should have received a copy of the GNU Affero General Public License
# along with FUJIN-P.  If not, see <https://www.gnu.org/licenses/>.
#
# Source: https://github.com/nishida-toyoaki/fujin-p

# /fujinp/colrep/scripts/excel_helper.py


import io
import json
import logging
import pandas as pd
from flask import Blueprint, request, jsonify, render_template
from auth import login_required
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Alignment

excel_helper_bp = Blueprint('excel_helper', __name__, url_prefix='/excel_helper', template_folder='../templates')

logging.basicConfig(level=logging.DEBUG)


@excel_helper_bp.route('/')
@login_required
def index():
    """エクセル→HTMLヘルパーのメイン画面"""
    return render_template('excel_helper.html')


@excel_helper_bp.route('/convert', methods=['POST'])
@login_required
def convert_excel_to_html():
    """エクセルをHTMLに変換"""
    try:
        file = request.files.get('excel_file')
        if not file:
            return jsonify({'success': False, 'error': 'ファイルが選択されていません'}), 400

        # 変換オプション
        style_type = request.form.get('style_type', 'simple')
        sheet_name = request.form.get('sheet_name', '0')

        # シート名を数値に変換（インデックス指定の場合）
        try:
            sheet_name = int(sheet_name)
        except ValueError:
            pass  # 文字列のシート名として扱う

        # スタイル情報を保持する場合はopenpyxlを使用
        if style_type == 'styled':
            html = convert_with_styles(file, sheet_name)
        else:
            # シンプル変換はpandasで高速処理
            html = convert_simple(file, sheet_name, style_type)

        # プレビュー用に整形
        preview_html = f"""
        <div style="overflow: auto; max-width: 100%;">
            {html}
        </div>
        """

        return jsonify({
            'success': True,
            'html': html,
            'preview': preview_html,
            'stats': get_table_stats(file, sheet_name)
        })

    except Exception as e:
        logging.error(f"Excel conversion error: {str(e)}")
        import traceback
        logging.error(f"Traceback: {traceback.format_exc()}")
        return jsonify({'success': False, 'error': str(e)}), 500


def convert_simple(file, sheet_name, style_type):
    """シンプルなHTML変換（完全インラインスタイル版）"""
    df = pd.read_excel(file, sheet_name=sheet_name)

    if style_type == 'simple':
        # シンプル版：基本的な罫線のみ
        table_style = (
            "border-collapse: collapse; "
            "border: 1px solid #000; "
            "width: 100%; "
            "font-family: Arial, sans-serif; "
            "font-size: 14px;"
        )
        th_style = (
            "border: 1px solid #000; "
            "padding: 8px; "
            "background-color: #f0f0f0; "
            "font-weight: bold; "
            "text-align: left;"
        )
        td_style = (
            "border: 1px solid #000; "
            "padding: 8px;"
        )

    elif style_type == 'bootstrap':
        # Bootstrap風：ストライプ
        table_style = (
            "border-collapse: collapse; "
            "width: 100%; "
            "margin-bottom: 1rem; "
            "font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif; "
            "font-size: 14px; "
            "color: #212529;"
        )
        th_style = (
            "border: 1px solid #dee2e6; "
            "padding: 0.75rem; "
            "background-color: #e9ecef; "
            "font-weight: bold; "
            "text-align: left; "
            "border-bottom: 2px solid #dee2e6;"
        )
        td_style = (
            "border: 1px solid #dee2e6; "
            "padding: 0.75rem;"
        )
    else:
        # デフォルト
        table_style = "border-collapse: collapse; width: 100%;"
        th_style = "border: 1px solid #ddd; padding: 8px;"
        td_style = "border: 1px solid #ddd; padding: 8px;"

    # HTML生成
    html_parts = [f'<table style="{table_style}">']

    # ヘッダー行
    html_parts.append('  <thead>')
    html_parts.append('    <tr>')
    for col in df.columns:
        html_parts.append(f'      <th style="{th_style}">{col}</th>')
    html_parts.append('    </tr>')
    html_parts.append('  </thead>')

    # データ行
    html_parts.append('  <tbody>')
    for idx, row in df.iterrows():
        # ストライプ効果（Bootstrap風の場合）
        if style_type == 'bootstrap' and idx % 2 == 1:
            row_td_style = td_style + " background-color: #f8f9fa;"
        else:
            row_td_style = td_style

        html_parts.append('    <tr>')
        for value in row:
            # 値の型に応じた配置
            if isinstance(value, (int, float)) and not pd.isna(value):
                cell_style = row_td_style + " text-align: right;"
            else:
                cell_style = row_td_style

            # NaN対策とHTMLエスケープ
            if pd.isna(value):
                display_value = ''
            else:
                display_value = str(value).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')

            html_parts.append(f'      <td style="{cell_style}">{display_value}</td>')
        html_parts.append('    </tr>')
    html_parts.append('  </tbody>')

    html_parts.append('</table>')

    return '\n'.join(html_parts)


def convert_with_styles(file, sheet_name):
    """スタイル情報を保持したHTML変換（完全インラインスタイル版）"""
    wb = load_workbook(file, data_only=False)

    if isinstance(sheet_name, int):
        ws = wb.worksheets[sheet_name]
    else:
        ws = wb[sheet_name]

    # テーブル全体のスタイル
    table_style = (
        "border-collapse: collapse; "
        "font-family: 'Calibri', Arial, sans-serif; "
        "font-size: 11pt;"
    )

    html_parts = [f'<table style="{table_style}">']

    # セル結合情報を取得
    merged_cells = {}
    for merged_range in ws.merged_cells.ranges:
        min_row, min_col = merged_range.min_row, merged_range.min_col
        max_row, max_col = merged_range.max_row, merged_range.max_col
        merged_cells[(min_row, min_col)] = {
            'rowspan': max_row - min_row + 1,
            'colspan': max_col - min_col + 1
        }
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                if (r, c) != (min_row, min_col):
                    merged_cells[(r, c)] = 'skip'

    # 各行を処理
    for row_idx, row in enumerate(ws.iter_rows(), start=1):
        html_parts.append('  <tr>')

        for col_idx, cell in enumerate(row, start=1):
            if merged_cells.get((row_idx, col_idx)) == 'skip':
                continue

            # セルのスタイルを完全にインラインCSSに変換
            inline_style = get_cell_inline_style(cell)

            # 結合セル情報
            merge_info = merged_cells.get((row_idx, col_idx), {})
            rowspan = merge_info.get('rowspan', 1)
            colspan = merge_info.get('colspan', 1)

            # HTMLタグ生成
            tag = 'th' if row_idx == 1 else 'td'
            attrs = []

            if rowspan > 1:
                attrs.append(f'rowspan="{rowspan}"')
            if colspan > 1:
                attrs.append(f'colspan="{colspan}"')
            attrs.append(f'style="{inline_style}"')

            attrs_str = ' ' + ' '.join(attrs)

            # セル値の取得とHTMLエスケープ
            cell_value = cell.value if cell.value is not None else ''
            cell_value = str(cell_value).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')

            html_parts.append(f'    <{tag}{attrs_str}>{cell_value}</{tag}>')

        html_parts.append('  </tr>')

    html_parts.append('</table>')

    return '\n'.join(html_parts)


def get_cell_inline_style(cell):
    """セルのスタイル情報を完全なインラインCSSに変換"""
    styles = []

    # デフォルトスタイル
    styles.append("padding: 5px")
    styles.append("border: 1px solid #000")

    # 背景色
    if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb:
        rgb = cell.fill.start_color.rgb
        if rgb and rgb not in ['00000000', 'FFFFFFFF']:
            if len(rgb) == 8:
                color = f"#{rgb[2:]}"
            else:
                color = f"#{rgb}"
            styles.append(f"background-color: {color}")

    # フォント設定
    if cell.font:
        # フォント色
        if cell.font.color and cell.font.color.rgb:
            rgb = cell.font.color.rgb
            if rgb and rgb not in ['00000000', '00000001']:
                if len(rgb) == 8:
                    color = f"#{rgb[2:]}"
                else:
                    color = f"#{rgb}"
                styles.append(f"color: {color}")

        # フォント名
        if cell.font.name:
            styles.append(f"font-family: '{cell.font.name}', Arial, sans-serif")

        # フォントサイズ
        if cell.font.size:
            styles.append(f"font-size: {cell.font.size}pt")

        # 太字
        if cell.font.bold:
            styles.append("font-weight: bold")

        # イタリック
        if cell.font.italic:
            styles.append("font-style: italic")

        # 下線
        if cell.font.underline:
            styles.append("text-decoration: underline")

        # 取り消し線
        if cell.font.strike:
            styles.append("text-decoration: line-through")

    # テキスト配置
    if cell.alignment:
        # 水平配置
        if cell.alignment.horizontal:
            h_align = cell.alignment.horizontal
            if h_align == 'center':
                styles.append("text-align: center")
            elif h_align == 'right':
                styles.append("text-align: right")
            elif h_align == 'left':
                styles.append("text-align: left")

        # 垂直配置
        if cell.alignment.vertical:
            v_align = cell.alignment.vertical
            if v_align == 'top':
                styles.append("vertical-align: top")
            elif v_align == 'center':
                styles.append("vertical-align: middle")
            elif v_align == 'bottom':
                styles.append("vertical-align: bottom")

        # 文字の折り返し
        if cell.alignment.wrap_text:
            styles.append("white-space: normal")
            styles.append("word-wrap: break-word")

    # 罫線
    if cell.border:
        if cell.border.left and cell.border.left.style:
            styles.append(get_border_style('left', cell.border.left))
        if cell.border.right and cell.border.right.style:
            styles.append(get_border_style('right', cell.border.right))
        if cell.border.top and cell.border.top.style:
            styles.append(get_border_style('top', cell.border.top))
        if cell.border.bottom and cell.border.bottom.style:
            styles.append(get_border_style('bottom', cell.border.bottom))

    return '; '.join(styles)


def get_border_style(position, border):
    """罫線スタイルを変換"""
    # 罫線の太さ
    if border.style == 'thin':
        width = '1px'
    elif border.style == 'medium':
        width = '2px'
    elif border.style == 'thick':
        width = '3px'
    else:
        width = '1px'

    # 罫線の色
    if border.color and border.color.rgb:
        rgb = border.color.rgb
        if len(rgb) == 8:
            color = f"#{rgb[2:]}"
        else:
            color = f"#{rgb}"
    else:
        color = '#000'

    # 罫線のスタイル
    if border.style in ['dashed', 'dotted']:
        line_style = border.style
    else:
        line_style = 'solid'

    return f"border-{position}: {width} {line_style} {color}"


def get_table_stats(file, sheet_name):
    """テーブルの統計情報を取得"""
    try:
        df = pd.read_excel(file, sheet_name=sheet_name)
        return {
            'rows': len(df),
            'columns': len(df.columns),
            'cells': len(df) * len(df.columns)
        }
    except:
        return {'rows': 0, 'columns': 0, 'cells': 0}


@excel_helper_bp.route('/get_sheets', methods=['POST'])
@login_required
def get_sheets():
    """エクセルファイルのシート一覧を取得"""
    try:
        file = request.files.get('excel_file')
        if not file:
            return jsonify({'success': False, 'error': 'ファイルが選択されていません'}), 400

        xl_file = pd.ExcelFile(file)
        sheets = xl_file.sheet_names

        return jsonify({
            'success': True,
            'sheets': sheets
        })

    except Exception as e:
        logging.error(f"Error getting sheets: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500