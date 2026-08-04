# SPDX-FileCopyrightText: 2024-2026 Toyoaki Nishida
# SPDX-License-Identifier: AGPL-3.0-or-later
#
# This file is part of FUJIN-P.
# Copyright (C) 2024-2026 Toyoaki Nishida
#
# FUJIN-P is free software: you can redistribute it and/or modify
# it under the terms of the GNU Affero General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# FUJIN-P is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU Affero General Public License for more details.
#
# You should have received a copy of the GNU Affero General Public License
# along with FUJIN-P.  If not, see <https://www.gnu.org/licenses/>.
#
# Source: https://github.com/nishida-toyoaki/fujin-p

import io
import os
import json
import uuid
import logging
import mysql.connector
import pandas as pd
from flask import Blueprint, request, jsonify, send_file, session, render_template, redirect, url_for, flash
from flask import Response  # 追加
from auth import login_required
import pytz
import datetime
from pytz import timezone
# from db import base_db_config, default_db_config
from config import Config
from db import DatabaseConfig, Tables
import re
# from medit.external import medit_external
# from markupsafe import Markup
import markdown
from markdown_converter import process_markdown
from markdown_converter import process_markdown_for_preview
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from werkzeug.utils import secure_filename
from auth import redirect_to_dashboard


# タイムゾーン設定
JST = timezone('Asia/Tokyo')

def get_jst_now():
    """現在の日時をJSTで取得（naive datetime）"""
    return datetime.datetime.now(JST).replace(tzinfo=None)

def serialize_for_json(obj):
    """datetime およびその他のオブジェクトを JSON シリアライズ可能な形式に変換"""
    # Jinja2 Undefined 型に対応
    from jinja2 import Undefined

    if isinstance(obj, Undefined):
        return None
    elif isinstance(obj, dict):
        return {k: serialize_for_json(v) for k, v in obj.items()}
    elif isinstance(obj, list):
        return [serialize_for_json(item) for item in obj]
    elif isinstance(obj, (datetime.datetime, datetime.date)):
        return obj.strftime('%Y-%m-%d %H:%M:%S')
    elif isinstance(obj, datetime.time):
        return obj.strftime('%H:%M:%S')
    elif obj is None:
        return None
    return obj


def build_project_export_json(project, rows, table_name):
    """プロジェクト＋タスクを、エクスポート／アーカイブ共通のJSON構造にまとめる。
    project は 責任者名 を含む行、rows は ユーザ名/氏名/content 等を含むタスク行のリスト。
    戻り値は serialize_for_json 済みの dict（そのまま json.dumps 可能）。"""
    export_data = {
        'project': {
            'id': project.get('id'),
            'プロジェクト名': project.get('プロジェクト名'),
            '責任者': project.get('責任者名'),
            'テーブル名': table_name,
            'Composer': project.get('Composer', ''),
            'is_public': project.get('is_public'),
            'export_date': get_jst_now().strftime('%Y-%m-%d %H:%M:%S'),
        },
        'tasks': [
            {
                'id': row.get('id'),
                '更新日時': row.get('更新日時'),
                'カラム名': row.get('カラム名'),
                '担当者': row.get('ユーザ名'),
                '担当者氏名': row.get('氏名'),
                '説明': row.get('説明'),
                '作業内容': row.get('content'),
                '備考': row.get('備考'),
                'ステータス': row.get('status'),
            }
            for row in rows
        ],
    }
    return serialize_for_json(export_data)

colrep_bp = Blueprint('colrep', __name__,
                     url_prefix='/colrep',
                     template_folder='../templates')

logging.basicConfig(level=logging.DEBUG)

# 定数定義
# ファイル先頭付近に追加
# UPLOAD_FOLDER = '/home/nishida/static/mdimgs'
# ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'webp'}
# COLREP_ADMIN_CATEGORY = "colrep総管理者"  # colrep総管理者
# COLREP_MANAGER_FEATURE = "CoRePo管理者"   # ★新規追加: CoRePo管理者フィーチャー名
# COLREP_PROJECTS_TABLE = "nishida$fujinp.colrep_projects"
# TARGET_DATABASE = "nishida$fujinp"

### 定数定義

UPLOAD_FOLDER = Config.UPLOAD_FOLDER
ALLOWED_EXTENSIONS = Config.ALLOWED_EXTENSIONS

# COLREP_ADMIN_CATEGORY = "colrep総管理者"   # 廃止: システム管理者(user_category=='admin')に統合
# COLREP_MANAGER_FEATURE = "CoRePo管理者"    # 廃止: user_groupベースに移行
COLREP_MANAGER_GROUP = "コレポ管理者"  # プロジェクト作成を許可するユーザーグループ名（完全一致）
COLREP_PROJECTS_TABLE = Tables.COLREP_PROJECTS
TARGET_DATABASE = Tables.DB_FUJINP

# SVG も許可する（Config 側に無くても確実に通す）
ALLOWED_EXTENSIONS = set(ALLOWED_EXTENSIONS) | {'svg'}


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def _sanitize_svg(data):
    """アップロードされた SVG から script / on* / javascript: を除去する簡易サニタイズ"""
    try:
        text = data.decode('utf-8', errors='replace')
    except Exception:
        return data
    # <script>...</script> 除去
    text = re.sub(r'<script\b[^>]*>.*?</script\s*>', '', text,
                  flags=re.IGNORECASE | re.DOTALL)
    # 自己終端の <script .../> 除去
    text = re.sub(r'<script\b[^>]*/\s*>', '', text, flags=re.IGNORECASE)
    # on*="..." イベントハンドラ除去
    text = re.sub(r'\son\w+\s*=\s*"[^"]*"', '', text, flags=re.IGNORECASE)
    text = re.sub(r"\son\w+\s*=\s*'[^']*'", '', text, flags=re.IGNORECASE)
    # javascript: URI 除去
    text = re.sub(r'javascript\s*:', '', text, flags=re.IGNORECASE)
    return text.encode('utf-8')

def get_user_info():
    """ユーザー情報を取得する関数"""
    user_id = session.get('user_id')
    logging.info(f"🔍 get_user_info called: user_id={user_id}")

    if not user_id:
        logging.warning("No user_id in session")
        return {'username': None, 'user_id': None, 'categories': [], 'subcategories': []}

    user_info = {
        'username': None,
        'user_id': user_id,
        'categories': [],
        'subcategories': []   # 互換性のため残す（権限判定にはもう使わない）
    }

    try:
        conn = mysql.connector.connect(**DatabaseConfig.default())
        cursor = conn.cursor(dictionary=True)

        # ユーザー基本情報の取得
        cursor.execute(f"""
            SELECT full_name, category
            FROM {Tables.USERS}
            WHERE id = %s
        """, (user_id,))

        user_data = cursor.fetchone()
        if user_data:
            user_info['username'] = user_data['full_name']
            user_info['categories'] = [user_data['category']] if user_data['category'] else []

    except Exception as e:
        logging.error(f"Error in get_user_info: {str(e)}")
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close()
            conn.close()

    return user_info

def get_user_active_group_names(user_id):
    """ユーザーが現在所属している有効なユーザーグループ名のリストを取得
    （default DB の user_groups / user_group_memberships を参照、有効期間チェック付き）"""
    if not user_id:
        return []
    try:
        now = get_jst_now()
        conn = mysql.connector.connect(**DatabaseConfig.default())
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT g.name
            FROM user_group_memberships m
            INNER JOIN user_groups g ON m.group_id = g.id
            WHERE m.user_id = %s
              AND (m.valid_from IS NULL OR m.valid_from <= %s)
              AND (m.valid_until IS NULL OR m.valid_until >= %s)
        """, (user_id, now, now))
        return [row['name'] for row in cursor.fetchall()]
    except Exception as e:
        logging.error(f"Error in get_user_active_group_names: {str(e)}")
        return []
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close()
            conn.close()

# 共有プレビューの公開範囲（文書アーカイブ／あわならと同じ区分）
VALID_ACCESS_POLICIES = {'public', 'domestic', 'private', 'group', 'domestic_group'}


def get_user_active_group_ids(user_id):
    """ユーザーが現在有効に所属しているグループIDのリスト（default DB）"""
    if not user_id:
        return []
    try:
        now = get_jst_now()
        conn = mysql.connector.connect(**DatabaseConfig.default())
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT group_id FROM user_group_memberships
            WHERE user_id = %s
              AND (valid_from IS NULL OR valid_from <= %s)
              AND (valid_until IS NULL OR valid_until >= %s)
        """, (user_id, now, now))
        return [r['group_id'] for r in cursor.fetchall()]
    except Exception as e:
        logging.error(f"Error in get_user_active_group_ids: {str(e)}")
        return []
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close()
            conn.close()


def get_all_user_groups():
    """全ユーザーグループの一覧（共有設定モーダルの選択肢用、default DB）"""
    try:
        conn = mysql.connector.connect(**DatabaseConfig.default())
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT id, name FROM user_groups ORDER BY id DESC")
        return cursor.fetchall()
    except Exception as e:
        logging.error(f"Error in get_all_user_groups: {str(e)}")
        return []
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close()
            conn.close()


def check_shared_preview_access(project_id, user_id):
    """
    共有プレビューの閲覧可否を判定。

    まず既存の check_project_access_permission
    （admin／責任者／担当者／is_public）を通し、
    ダメなら access_policy（文書アーカイブと同じルール）で判定する：
      public         - ログイン済みユーザなら誰でも
      domestic       - regular ユーザのみ
      group          - 指定グループの有効所属者（カテゴリ不問）
      domestic_group - regular または指定グループ所属者（和集合）
      private        - 上記の既存アクセス者のみ
    """
    has_access, _is_author = check_project_access_permission(project_id, user_id)
    if has_access:
        return True

    try:
        conn = mysql.connector.connect(**DatabaseConfig.fujinp())
        cursor = conn.cursor(dictionary=True)
        cursor.execute(f"SELECT access_policy FROM {COLREP_PROJECTS_TABLE} WHERE id = %s",
                       (project_id,))
        row = cursor.fetchone()
        if not row:
            return False
        policy = row.get('access_policy') or 'private'

        if policy == 'public':
            return True

        user_category = session.get('user_category', '')
        if policy == 'domestic':
            return user_category == 'regular'

        if policy in ('group', 'domestic_group'):
            if policy == 'domestic_group' and user_category == 'regular':
                return True
            cursor.execute(
                "SELECT group_id FROM colrep_access_groups WHERE project_id = %s",
                (project_id,))
            allowed = {r['group_id'] for r in cursor.fetchall()}
            mine = set(get_user_active_group_ids(user_id))
            return bool(allowed & mine)

        return False
    except Exception as e:
        logging.error(f"Error in check_shared_preview_access: {str(e)}")
        return False
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close()
            conn.close()

def check_colrep_admin_permission():
    """全プロジェクト管理権限をチェック（旧colrep総管理者はシステム管理者に統合）"""
    return session.get('user_category') == 'admin'

def get_user_permissions():
    """ユーザーの権限情報を取得（user_groupベース）
    - is_colrep_admin  : 全プロジェクトの閲覧・管理（システム管理者 user_category=='admin'）
    - is_colrep_manager: プロジェクト作成可（グループ「コレポ管理者」に有効所属）
    """
    user_info = get_user_info()
    user_id = user_info.get('user_id')

    is_system_admin = (session.get('user_category') == 'admin')

    # adminはグループ照会不要（is_colrep_admin側で全権限が付与される）
    if is_system_admin:
        group_names = []
    else:
        group_names = get_user_active_group_names(user_id)

    permissions = {
        'is_colrep_admin': is_system_admin,
        'is_colrep_manager': COLREP_MANAGER_GROUP in group_names,
        'user_info': user_info
    }

    return permissions


# routes.py の修正（index() 関数のみ）
# 既存の index() 関数を以下で置き換えてください

@colrep_bp.route('/')
@login_required
def index():
    """colrepダッシュボード"""
    permissions = get_user_permissions()
    user_info = get_user_info()
    user_id = user_info.get('user_id')
    user_category = session.get('user_category')

    try:
        # conn = mysql.connector.connect(**base_db_config, database=TARGET_DATABASE)
        conn = mysql.connector.connect(**DatabaseConfig.fujinp())
        cursor = conn.cursor(dictionary=True)

        # ✅ 修正: 管理者と一般ユーザーで異なるクエリを使用
        if user_category == 'admin' or permissions['is_colrep_admin']:
            # 管理者: すべてのプロジェクトを表示
            query = f"""
                SELECT cp.id, cp.プロジェクト名, cp.更新日時, cp.責任者, cp.テーブル名, cp.Composer, cp.is_public,
                       u.full_name as 責任者名
                FROM {COLREP_PROJECTS_TABLE} cp
                LEFT JOIN {Tables.USERS} u ON cp.責任者 = u.id
                ORDER BY cp.更新日時 DESC
            """
            cursor.execute(query)
        else:
            # 一般ユーザー: 自分が責任者になっているプロジェクトのみ表示
            query = f"""
                SELECT cp.id, cp.プロジェクト名, cp.更新日時, cp.責任者, cp.テーブル名, cp.Composer, cp.is_public,
                       u.full_name as 責任者名
                FROM {COLREP_PROJECTS_TABLE} cp
                LEFT JOIN {Tables.USERS} u ON cp.責任者 = u.id
                WHERE cp.責任者 = %s OR cp.is_public = TRUE
                ORDER BY cp.更新日時 DESC
            """
            cursor.execute(query, (user_id,))

        projects = cursor.fetchall()

        # 統計情報の計算
        stats = {
            'total_projects': len(projects),
            'recent_projects': len([p for p in projects if p['更新日時'] and
                                  (get_jst_now() - p['更新日時']).days <= 7])
        }

        # システム情報（DB・テーブル名を実際の設定から取得）
        system_info = {
            'database': TARGET_DATABASE,
            'table': COLREP_PROJECTS_TABLE.split('.')[-1],
        }

        return render_template('colrep_dashboard.html',
                              projects=projects,
                              stats=stats,
                              system_info=system_info,
                              permissions=permissions)

    except Exception as e:
        logging.error(f"Error in colrep index: {str(e)}")
        return render_template('colrep_dashboard.html',
                              error=str(e),
                              projects=[],
                              stats={'total_projects': 0, 'recent_projects': 0},
                              system_info={
                                  'database': TARGET_DATABASE,
                                  'table': COLREP_PROJECTS_TABLE.split('.')[-1],
                              },
                              permissions=permissions)
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close()
            conn.close()

@colrep_bp.route('/create_project', methods=['GET', 'POST'])
@login_required
def create_project():
    """プロジェクト作成"""

    # ★修正: 権限チェックを変更 (総管理者 OR CoRePo管理者 ならOK)
    permissions = get_user_permissions()
    if not (permissions['is_colrep_admin'] or permissions['is_colrep_manager']):
        flash('プロジェクト作成権限（CoRePo管理者以上）が必要です。', 'error')
        return redirect(url_for('colrep.index'))

    if request.method == 'GET':
        try:
            # conn = mysql.connector.connect(**base_db_config, database="nishida$default")
            conn = mysql.connector.connect(**DatabaseConfig.default())
            cursor = conn.cursor(dictionary=True)

            cursor.execute(f"SELECT id, full_name FROM {Tables.USERS} ORDER BY full_name")
            users = cursor.fetchall()

            return render_template('create_colrep_project.html',
                                 users=users,
                                 current_user_id=session.get('user_id'))

        except Exception as e:
            logging.error(f"Error in create_project GET: {str(e)}")
            flash(f'エラーが発生しました: {str(e)}', 'error')
            return redirect(url_for('colrep.index'))
        finally:
            if 'conn' in locals() and conn.is_connected():
                cursor.close()
                conn.close()

    elif request.method == 'POST':
        try:
            project_name = request.form.get('project_name')
            responsible_user = request.form.get('responsible_user')
            composer = request.form.get('composer', '')
            # ✅ 追加: is_public フラグを取得
            is_public = request.form.get('is_public') == 'on'
            # ★追加: 「本文」タスク自動作成フラグ（フォーム既定はON）
            create_main_task = request.form.get('create_main_task') == 'on'

            if not project_name or not responsible_user:
                flash('必須項目を入力してください。', 'error')
                return redirect(url_for('colrep.create_project'))
            # ★追加: Composerが空なら本文マクロを初期設定
            if create_main_task and not composer.strip():
                composer = '[[本文]]'
            table_name = f"colrep_{uuid.uuid4().hex[:12]}"

            # conn = mysql.connector.connect(**base_db_config, database=TARGET_DATABASE)
            conn = mysql.connector.connect(**DatabaseConfig.fujinp())
            cursor = conn.cursor()

            now = get_jst_now()

            create_table_sql = f"""
                CREATE TABLE `{table_name}` (
                    `id` INT AUTO_INCREMENT PRIMARY KEY,
                    `更新日時` DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                    `カラム名` VARCHAR(255) NOT NULL,
                    `担当者アカウント` INT NOT NULL,
                    `説明` TEXT COMMENT '管理者から入力者への説明',
                    `content` LONGTEXT COMMENT '入力内容',
                    `備考` TEXT COMMENT '入力者から管理者への説明',
                    `status` VARCHAR(20) DEFAULT '作業中' COMMENT '進捗状況：作業中/改訂中/完了',
                    INDEX `idx_担当者` (`担当者アカウント`),
                    INDEX `idx_カラム名` (`カラム名`),
                    INDEX `idx_status` (`status`),
                    FOREIGN KEY (`担当者アカウント`) REFERENCES {Tables.USERS}(`id`) ON DELETE RESTRICT
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
                COMMENT='CoRePoプロジェクト用データテーブル'
            """

            try:
                cursor.execute(create_table_sql)
                conn.commit()
                logging.info(f"テーブル作成成功: {table_name}")
            except Exception as table_error:
                logging.error(f"テーブル作成失敗: {str(table_error)}")
                if 'conn' in locals() and conn.is_connected():
                    conn.rollback()
                flash(f'テーブル作成中にエラーが発生しました。', 'error')
                return redirect(url_for('colrep.create_project'))

            # ✅ 修正: is_public カラムを追加
            insert_query = f"""
                INSERT INTO {COLREP_PROJECTS_TABLE}
                (プロジェクト名, 更新日時, 責任者, テーブル名, Composer, is_public)
                VALUES (%s, %s, %s, %s, %s, %s)
            """

            cursor.execute(insert_query, (
                project_name,
                now,
                responsible_user,
                table_name,
                composer,
                is_public
            ))

            project_id = cursor.lastrowid
            # ★追加: 「本文」タスクを自動作成（担当者＝責任者）
            if create_main_task:
                cursor.execute(f"""
                    INSERT INTO `{table_name}` (更新日時, カラム名, 担当者アカウント, 説明)
                    VALUES (%s, %s, %s, %s)
                """, (now, '本文', responsible_user, '本文の執筆'))

            conn.commit()

            access_type = "公開" if is_public else "限定"
            if create_main_task:
                flash(f'プロジェクト「{project_name}」を作成し、「本文」タスクを責任者に割り当てました。'
                      f'作業ダッシュボードから執筆を開始できます。（{access_type}プロジェクト）', 'success')
            else:
                flash(f'プロジェクト「{project_name}」を作成しました。（{access_type}プロジェクト）', 'success')
            return redirect(url_for('colrep.view_project', project_id=project_id))

        except mysql.connector.IntegrityError as e:
            logging.error(f"Integrity error in create_project: {str(e)}")
            flash('プロジェクト名またはテーブル名が重複している可能性があります。', 'error')
            return redirect(url_for('colrep.create_project'))
        except Exception as e:
            logging.error(f"Error in create_project POST: {str(e)}")
            if 'conn' in locals() and conn.is_connected():
                conn.rollback()
            flash(f'エラーが発生しました: {str(e)}', 'error')
            return redirect(url_for('colrep.create_project'))
        finally:
            if 'conn' in locals() and conn.is_connected():
                cursor.close()
                conn.close()

@colrep_bp.route('/quick_create_project', methods=['POST'])
@login_required
def quick_create_project():
    """クイック作成：「本文」タスクだけのプロジェクトを作成し、
    作成者を責任者兼執筆者にして即執筆を開始できるようにする（JSON API）"""
    permissions = get_user_permissions()
    if not (permissions['is_colrep_admin'] or permissions['is_colrep_manager']):
        return jsonify({'success': False,
                        'error': 'プロジェクト作成権限（CoRePo管理者以上）が必要です。'}), 403

    user_id = session.get('user_id')
    data = request.get_json() or {}
    project_name = (data.get('project_name') or '').strip()
    is_public = bool(data.get('is_public', False))

    if not project_name:
        return jsonify({'success': False, 'error': 'プロジェクト名を入力してください。'}), 400

    table_name = f"colrep_{uuid.uuid4().hex[:12]}"

    try:
        conn = mysql.connector.connect(**DatabaseConfig.fujinp())
        cursor = conn.cursor()
        now = get_jst_now()

        # プロジェクト固有テーブル（create_project と同一DDL）
        create_table_sql = f"""
            CREATE TABLE `{table_name}` (
                `id` INT AUTO_INCREMENT PRIMARY KEY,
                `更新日時` DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                `カラム名` VARCHAR(255) NOT NULL,
                `担当者アカウント` INT NOT NULL,
                `説明` TEXT COMMENT '管理者から入力者への説明',
                `content` LONGTEXT COMMENT '入力内容',
                `備考` TEXT COMMENT '入力者から管理者への説明',
                `status` VARCHAR(20) DEFAULT '作業中' COMMENT '進捗状況：作業中/改訂中/完了',
                INDEX `idx_担当者` (`担当者アカウント`),
                INDEX `idx_カラム名` (`カラム名`),
                INDEX `idx_status` (`status`),
                FOREIGN KEY (`担当者アカウント`) REFERENCES {Tables.USERS}(`id`) ON DELETE RESTRICT
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
            COMMENT='CoRePoプロジェクト用データテーブル'
        """
        cursor.execute(create_table_sql)

        # 台帳へ登録：責任者＝作成者、Composer初期値＝[[本文]]
        cursor.execute(f"""
            INSERT INTO {COLREP_PROJECTS_TABLE}
            (プロジェクト名, 更新日時, 責任者, テーブル名, Composer, is_public)
            VALUES (%s, %s, %s, %s, %s, %s)
        """, (project_name, now, user_id, table_name, '[[本文]]', is_public))
        project_id = cursor.lastrowid

        # 「本文」タスク：担当者＝作成者
        cursor.execute(f"""
            INSERT INTO `{table_name}` (更新日時, カラム名, 担当者アカウント, 説明)
            VALUES (%s, %s, %s, %s)
        """, (now, '本文', user_id, '本文の執筆（クイック作成）'))
        task_id = cursor.lastrowid

        conn.commit()
        logging.info(f"クイック作成成功: project={project_id}, task={task_id}, user={user_id}")

        return jsonify({
            'success': True,
            'project_id': project_id,
            'task_id': task_id,
            'project_name': project_name
        })

    except Exception as e:
        logging.error(f"Error in quick_create_project: {str(e)}")
        if 'conn' in locals() and conn.is_connected():
            conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close()
            conn.close()



@colrep_bp.route('/awami_create_project', methods=['POST'])
@login_required
def awami_create_project():
    """あわみ連携専用：プロジェクト＋「本文」タスク＋共有設定を一括作成し、
    実体URL（共有プレビュー）とエディタURLを返す（JSON API）。

    あわみのキャンバス上で「新規文書を生成してリンク」を押したときに呼ばれる。
    既存の quick_create_project / save_share_settings とは独立した専用ルートで、
    既存機能の挙動には影響しない。"""
    permissions = get_user_permissions()
    if not (permissions['is_colrep_admin'] or permissions['is_colrep_manager']):
        return jsonify({'success': False, 'error_code': 'forbidden',
                        'error': 'プロジェクト作成権限（コレポ管理者以上）が必要です。'}), 403

    user_id = session.get('user_id')
    data = request.get_json() or {}
    project_name = (data.get('project_name') or '').strip()
    access_policy = data.get('access_policy') or 'private'
    group_ids = data.get('access_group_ids') or []
    initial_content = data.get('initial_content') or None

    if not project_name:
        return jsonify({'success': False, 'error_code': 'name_required',
                        'error': 'プロジェクト名を入力してください。'}), 400
    if access_policy not in VALID_ACCESS_POLICIES:
        access_policy = 'private'          # 無効値は安全側に倒す
    if access_policy in ('group', 'domestic_group') and not group_ids:
        return jsonify({'success': False, 'error_code': 'group_required',
                        'error': 'グループ公開では、許可するグループを1つ以上指定してください。'}), 400

    table_name = f"colrep_{uuid.uuid4().hex[:12]}"

    try:
        conn = mysql.connector.connect(**DatabaseConfig.fujinp())
        cursor = conn.cursor(dictionary=True)
        now = get_jst_now()

        # --- 同名チェック（プロジェクト名は UNIQUE。あわみ側で再入力を促すため409で返す）---
        cursor.execute(
            f"SELECT id FROM {COLREP_PROJECTS_TABLE} WHERE プロジェクト名 = %s",
            (project_name,))
        if cursor.fetchone():
            return jsonify({'success': False, 'error_code': 'duplicate_name',
                            'error': f'同名のプロジェクト「{project_name}」が既にあります。'}), 409

        # --- プロジェクト固有テーブル（quick_create_project と同一DDL）---
        create_table_sql = f"""
            CREATE TABLE `{table_name}` (
                `id` INT AUTO_INCREMENT PRIMARY KEY,
                `更新日時` DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                `カラム名` VARCHAR(255) NOT NULL,
                `担当者アカウント` INT NOT NULL,
                `説明` TEXT COMMENT '管理者から入力者への説明',
                `content` LONGTEXT COMMENT '入力内容',
                `備考` TEXT COMMENT '入力者から管理者への説明',
                `status` VARCHAR(20) DEFAULT '作業中' COMMENT '進捗状況：作業中/改訂中/完了',
                INDEX `idx_担当者` (`担当者アカウント`),
                INDEX `idx_カラム名` (`カラム名`),
                INDEX `idx_status` (`status`),
                FOREIGN KEY (`担当者アカウント`) REFERENCES {Tables.USERS}(`id`) ON DELETE RESTRICT
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
            COMMENT='CoRePoプロジェクト用データテーブル'
        """
        cursor.execute(create_table_sql)

        # --- 台帳へ登録：責任者＝作成者、Composer初期値＝[[本文]]、公開範囲も同時設定 ---
        cursor.execute(f"""
            INSERT INTO {COLREP_PROJECTS_TABLE}
            (プロジェクト名, 更新日時, 責任者, テーブル名, Composer, is_public, access_policy)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """, (project_name, now, user_id, table_name, '[[本文]]', False, access_policy))
        project_id = cursor.lastrowid

        # --- 「本文」タスク：担当者＝作成者（初期本文があれば書き込む）---
        cursor.execute(f"""
            INSERT INTO `{table_name}` (更新日時, カラム名, 担当者アカウント, 説明, content)
            VALUES (%s, %s, %s, %s, %s)
        """, (now, '本文', user_id, '本文の執筆（あわみから作成）', initial_content))
        task_id = cursor.lastrowid

        # --- 共有許可グループ（save_share_settings と同じ保存規則：全削除→再挿入）---
        cursor.execute("DELETE FROM colrep_access_groups WHERE project_id = %s",
                       (project_id,))
        if access_policy in ('group', 'domestic_group'):
            for gid in group_ids:
                cursor.execute(
                    "INSERT INTO colrep_access_groups (project_id, group_id) VALUES (%s, %s)",
                    (project_id, int(gid)))

        conn.commit()
        logging.info(f"あわみ連携作成成功: project={project_id}, task={task_id}, "
                     f"user={user_id}, policy={access_policy}")

        return jsonify({
            'success': True,
            'project_id': project_id,
            'task_id': task_id,
            'project_name': project_name,
            'access_policy': access_policy,
            'entity_url': url_for('colrep.shared_preview', project_id=project_id),
            'editor_url': url_for('colrep.edit_content_external',
                                  project_id=project_id, task_id=task_id),
        })

    except mysql.connector.IntegrityError as e:
        logging.error(f"Integrity error in awami_create_project: {str(e)}")
        if 'conn' in locals() and conn.is_connected():
            conn.rollback()
        return jsonify({'success': False, 'error_code': 'duplicate_name',
                        'error': 'プロジェクト名またはテーブル名が重複しています。'}), 409
    except Exception as e:
        logging.error(f"Error in awami_create_project: {str(e)}")
        if 'conn' in locals() and conn.is_connected():
            conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close()
            conn.close()


@colrep_bp.route('/project/<int:project_id>')
@login_required
def view_project(project_id):
    """プロジェクト詳細表示"""
    user_id = session.get('user_id')          # [2026-07-25 改修] 権限チェック用
    try:
        # conn = mysql.connector.connect(**base_db_config, database=TARGET_DATABASE)
        conn = mysql.connector.connect(**DatabaseConfig.fujinp())
        cursor = conn.cursor(dictionary=True)

        # プロジェクト情報を取得
        query = f"""
            SELECT cp.*, u.full_name as 責任者名, u.full_name as 責任者氏名
            FROM {COLREP_PROJECTS_TABLE} cp
            LEFT JOIN {Tables.USERS} u ON cp.責任者 = u.id
            WHERE cp.id = %s
        """
        cursor.execute(query, (project_id,))
        project = cursor.fetchone()

        if not project:
            flash('指定されたプロジェクトが見つかりません。', 'error')
            return redirect(url_for('colrep.index'))

        # [2026-07-25 改修] 権限チェックを追加（従来はログインさえしていれば誰でも閲覧できた）
        # admin / 責任者 / 担当者 / 公開プロジェクト のいずれかであること
        has_access, _is_author = check_project_access_permission(project_id, user_id)
        if not has_access:
            flash('このプロジェクトを閲覧する権限がありません。', 'error')
            return redirect(url_for('colrep.index'))

        # テーブルが存在するかチェック
        table_exists = False
        table_info = None
        if project['テーブル名']:
            try:
                cursor.execute(f"SHOW TABLES LIKE '{project['テーブル名']}'")
                table_exists = cursor.fetchone() is not None

                if table_exists:
                    # テーブル情報を取得
                    cursor.execute(f"DESCRIBE `{project['テーブル名']}`")
                    table_structure = cursor.fetchall()

                    cursor.execute(f"SELECT COUNT(*) as record_count FROM `{project['テーブル名']}`")
                    record_count = cursor.fetchone()['record_count']

                    table_info = {
                        'structure': table_structure,
                        'record_count': record_count
                    }
            except Exception as e:
                logging.warning(f"Error checking table {project['テーブル名']}: {str(e)}")

        permissions = get_user_permissions()

        # 日本時間を文字列として取得
        now_dt = get_jst_now()
        now_date = now_dt.strftime('%Y%m%d')
        now_datetime = now_dt.strftime('%Y年%m月%d日 %H:%M')

        return render_template('colrep_project_detail.html',
                              project=project,
                              table_exists=table_exists,
                              table_info=table_info,
                              permissions=permissions,
                              now_date=now_date,
                              now_datetime=now_datetime)

    except Exception as e:
        logging.error(f"Error in view_project: {str(e)}")
        flash(f'エラーが発生しました: {str(e)}', 'error')
        return redirect(url_for('colrep.index'))
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close()
            conn.close()

# ============================================================
# 3. 既存関数の修正: edit_project
# ============================================================

@colrep_bp.route('/edit_project/<int:project_id>', methods=['GET', 'POST'])
@login_required
def edit_project(project_id):
    """プロジェクト編集"""
    user_info = get_user_info()
    user_id = user_info.get('user_id')
    permissions = get_user_permissions()

    conn = None  # ★追加：初期化
    cursor = None  # ★追加：初期化

    try:
        conn = mysql.connector.connect(**DatabaseConfig.fujinp())
        cursor = conn.cursor(dictionary=True)

        # プロジェクト情報取得
        cursor.execute(f"SELECT * FROM {COLREP_PROJECTS_TABLE} WHERE id = %s", (project_id,))
        project = cursor.fetchone()

        if not project:
            flash('指定されたプロジェクトが見つかりません。', 'error')
            return redirect(url_for('colrep.index'))

        # 権限チェック
        is_system_admin = (session.get('user_category') == 'admin')
        is_colrep_admin = permissions['is_colrep_admin']
        is_owner = (project['責任者'] == user_id)

        if not (is_system_admin or is_colrep_admin or is_owner):
            flash('権限がありません。', 'error')
            return redirect(url_for('colrep.index'))

        if request.method == 'GET':
            # ★修正：同じ接続を使い回す
            cursor.execute(f"SELECT id, full_name FROM {Tables.USERS} ORDER BY full_name")
            users = cursor.fetchall()

            return render_template('edit_colrep_project.html',
                                 project=project,
                                 users=users,
                                 permissions=permissions)

        elif request.method == 'POST':
            project_name = request.form.get('project_name')
            responsible_user = request.form.get('responsible_user')
            table_name = request.form.get('table_name')
            composer = request.form.get('composer', '')
            is_public = request.form.get('is_public') == 'on'

            if not project_name or not responsible_user or not table_name:
                flash('必須項目を入力してください。', 'error')
                return redirect(url_for('colrep.edit_project', project_id=project_id))

            now = get_jst_now()
            update_query = f"""
                UPDATE {COLREP_PROJECTS_TABLE}
                SET プロジェクト名 = %s, 更新日時 = %s, 責任者 = %s, テーブル名 = %s, Composer = %s, is_public = %s
                WHERE id = %s
            """

            cursor.execute(update_query, (
                project_name,
                now,
                responsible_user,
                table_name,
                composer,
                is_public,
                project_id
            ))

            conn.commit()
            access_type = "公開" if is_public else "限定"
            flash(f'プロジェクトを更新しました。（{access_type}プロジェクト）', 'success')
            return redirect(url_for('colrep.index'))

    except Exception as e:
        logging.error(f"Error in edit_project: {str(e)}")
        if conn and conn.is_connected():  # ★修正
            conn.rollback()
        flash(f'エラーが発生しました: {str(e)}', 'error')
        return redirect(url_for('colrep.view_project', project_id=project_id))
    finally:
        if cursor:  # ★修正
            try:
                cursor.close()
            except:
                pass
        if conn and conn.is_connected():  # ★修正
            conn.close()

# [2026-07-25 改修] 削除: /toggle_project_access
# どのテンプレートからも呼ばれていないデッドルートだったため削除。
# 公開/限定の切り替えは /edit_project のフォーム（is_public）で行う。

@colrep_bp.route('/delete_project/<int:project_id>', methods=['POST'])
@login_required
def delete_project(project_id):
    """プロジェクト削除"""
    # ユーザー情報の取得
    user_info = get_user_info()
    user_id = user_info.get('user_id')
    permissions = get_user_permissions()

    try:
        # conn = mysql.connector.connect(**base_db_config, database=TARGET_DATABASE)
        conn = mysql.connector.connect(**DatabaseConfig.fujinp())
        cursor = conn.cursor(dictionary=True)

        # プロジェクト情報を取得
        cursor.execute(f"SELECT プロジェクト名, テーブル名, 責任者 FROM {COLREP_PROJECTS_TABLE} WHERE id = %s", (project_id,))
        project = cursor.fetchone()

        if not project:
            return jsonify({'success': False, 'error': 'プロジェクトが見つかりません。'}), 404

        # ★★★ 権限チェック修正: 総管理者 OR プロジェクト責任者のみ許可 ★★★
        is_system_admin = (session.get('user_category') == 'admin')
        is_colrep_admin = permissions['is_colrep_admin']
        is_owner = (project['責任者'] == user_id)

        # 誰か一人がTrueならアクセス許可
        if not (is_system_admin or is_colrep_admin or is_owner):
            flash('権限がありません。', 'error')
            return redirect(url_for('colrep.index'))
        # ★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★
        # テーブル名を確認
        table_name = project.get('テーブル名')

        # ========== 新規追加: テーブルが存在すれば削除 ==========
        if table_name:
            try:
                cursor.execute(f"SHOW TABLES LIKE %s", (table_name,))
                if cursor.fetchone():
                    # テーブルを削除
                    cursor.execute(f"DROP TABLE IF EXISTS `{table_name}`")
                    conn.commit()
                    logging.info(f"プロジェクト {project_id} のテーブル削除: {table_name}")
            except Exception as table_error:
                logging.error(f"テーブル削除エラー: {str(table_error)}")
                # テーブル削除エラーでもプロジェクト削除は続行
        # =============================================

        # プロジェクトを削除
        cursor.execute(f"DELETE FROM {COLREP_PROJECTS_TABLE} WHERE id = %s", (project_id,))
        conn.commit()

        return jsonify({
            'success': True,
            'message': 'プロジェクトとテーブルを削除しました。',
            'redirect_url': url_for('colrep.index')
        })
    except Exception as e:
        logging.error(f"Error in delete_project: {str(e)}")
        if 'conn' in locals() and conn.is_connected():
            conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close()
            conn.close()

# [2026-07-25 改修] 削除: /project_stats
# 全プロジェクトの統計を無条件で返す未使用APIだったため削除。
# ダッシュボードの統計は index() 内で権限に応じて算出している。

@colrep_bp.route('/manage_table/<int:project_id>')
@login_required
def manage_project_table(project_id):
    """プロジェクトテーブル管理ページ（プロジェクト責任者のみ）"""
    user_info = get_user_info()
    user_id = user_info.get('user_id')

    try:
        # conn = mysql.connector.connect(**base_db_config, database=TARGET_DATABASE)
        conn = mysql.connector.connect(**DatabaseConfig.fujinp())
        cursor = conn.cursor(dictionary=True)

        # プロジェクト情報を取得
        cursor.execute(f"""
            SELECT cp.*, u.full_name as 責任者名
            FROM {COLREP_PROJECTS_TABLE} cp
            LEFT JOIN {Tables.USERS} u ON cp.責任者 = u.id
            WHERE cp.id = %s
        """, (project_id,))
        project = cursor.fetchone()

        if not project:
            flash('プロジェクトが見つかりません。', 'error')
            return redirect(url_for('colrep.index'))

        # 権限チェック：システム管理者、CoRePo総管理者、または責任者
        permissions = get_user_permissions()
        is_system_admin = (session.get('user_category') == 'admin')
        is_colrep_admin = permissions['is_colrep_admin']
        is_owner = (project['責任者'] == user_id)

        if not (is_system_admin or is_colrep_admin or is_owner):
            flash('テーブル管理はプロジェクト責任者または管理者のみ実行可能です。', 'error')
            return redirect(url_for('colrep.view_project', project_id=project_id))

        table_name = project['テーブル名']

        # テーブルが存在するかチェック
        cursor.execute(f"SHOW TABLES LIKE '{table_name}'")
        table_exists = cursor.fetchone() is not None

        if not table_exists:
            flash('テーブルがまだ作成されていません。総管理者にテーブル作成を依頼してください。', 'warning')
            return redirect(url_for('colrep.view_project', project_id=project_id))

        table_data = []
        # ✅ 修正: エイリアスを追加してテンプレートで使えるようにする
        cursor.execute(f"""
            SELECT t.*,
                   u.full_name as ユーザ名,
                   u.full_name as 氏名
            FROM `{table_name}` t
            LEFT JOIN {Tables.USERS} u ON t.担当者アカウント = u.id
            ORDER BY t.id ASC
        """)
        table_data = cursor.fetchall()

        # ユーザー一覧を取得（担当者選択用）
        # conn_users = mysql.connector.connect(**base_db_config, database="nishida$default")
        conn_users = mysql.connector.connect(**DatabaseConfig.default())
        cursor_users = conn_users.cursor(dictionary=True)
        cursor_users.execute(f"SELECT id, full_name FROM {Tables.USERS} ORDER BY full_name")
        users = cursor_users.fetchall()
        cursor_users.close()
        conn_users.close()

        return render_template('manage_project_table.html',
                              project=project,
                              table_exists=table_exists,
                              table_data=table_data,
                              users=users,
                              user_info=user_info)  # ★ 追加

    except Exception as e:
        logging.error(f"Error in manage_project_table: {str(e)}")
        flash(f'エラーが発生しました: {str(e)}', 'error')
        return redirect(url_for('colrep.view_project', project_id=project_id))
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close()
            conn.close()

@colrep_bp.route('/add_table_row/<int:project_id>', methods=['POST'])
@login_required
def add_table_row(project_id):
    """テーブルに新しい行を追加（プロジェクト責任者のみ）"""
    user_info = get_user_info()
    user_id = user_info.get('user_id')
    permissions = get_user_permissions()

    try:
        # conn = mysql.connector.connect(**base_db_config, database=TARGET_DATABASE)
        conn = mysql.connector.connect(**DatabaseConfig.fujinp())
        cursor = conn.cursor(dictionary=True)

        # プロジェクト情報を取得
        cursor.execute(f"SELECT テーブル名, 責任者 FROM {COLREP_PROJECTS_TABLE} WHERE id = %s", (project_id,))
        project = cursor.fetchone()

        if not project:
            return jsonify({'success': False, 'error': 'プロジェクトが見つかりません。'}), 404

        # 権限チェック：システム管理者、CoRePo総管理者、または責任者
        is_system_admin = (session.get('user_category') == 'admin')
        is_colrep_admin = permissions['is_colrep_admin']
        is_owner = (project['責任者'] == user_id)

        if not (is_system_admin or is_colrep_admin or is_owner):
            return jsonify({'success': False, 'error': 'タスクの追加はプロジェクト責任者または管理者のみ実行可能です。'}), 403
        data = request.get_json()
        カラム名 = data.get('カラム名')
        担当者アカウント = data.get('担当者アカウント')
        説明 = data.get('説明', '')

        if not カラム名 or not 担当者アカウント:
            return jsonify({'success': False, 'error': '必須項目を入力してください。'}), 400

        table_name = project['テーブル名']

        # 新しい行を追加
        now = get_jst_now()
        insert_query = f"""
            INSERT INTO `{table_name}` (更新日時, カラム名, 担当者アカウント, 説明)
            VALUES (%s, %s, %s, %s)
        """

        cursor.execute(insert_query, (now, カラム名, 担当者アカウント, 説明))
        conn.commit()

        return jsonify({
            'success': True,
            'message': '新しいタスクを追加しました。'
        })

    except Exception as e:
        logging.error(f"Error in add_table_row: {str(e)}")
        if 'conn' in locals() and conn.is_connected():
            conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close()
            conn.close()

# worker_dashboard ルート内の修正箇所

def safe_process_content(content):
    """安全なコンテンツ処理関数"""
    if content is None:
        return ""

    # 型安全性の確保
    if not isinstance(content, str):
        try:
            content = str(content)
        except (TypeError, ValueError):
            return ""

    # 文字列の正規化
    content = content.strip()
    if not content:
        return ""

    try:
        # 危険な制御文字の除去（改行、タブは保持）
        cleaned = ''.join(c for c in content if ord(c) >= 32 or c in '\t\n\r')

        if not cleaned.strip():
            return ""

        return cleaned

    except Exception as e:
        logging.error(f"Content processing error: {str(e)}")
        return ""

def safe_calculate_length(content):
    """安全な文字数計算"""
    try:
        if content is None:
            return 0
        if not isinstance(content, str):
            content = str(content)
        return len(content.strip())
    except Exception:
        return 0

def calculate_safe_statistics(all_tasks):
    """統計情報の安全な計算（関数として分離）"""
    try:
        total_tasks = len(all_tasks)
        completed_tasks = 0
        pending_tasks = 0

        for task in all_tasks:
            try:
                content = task.get('content', '')
                if isinstance(content, str) and len(content.strip()) > 0:
                    completed_tasks += 1
                else:
                    pending_tasks += 1
            except Exception:
                pending_tasks += 1

        # 二重チェック
        if completed_tasks + pending_tasks != total_tasks:
            pending_tasks = total_tasks - completed_tasks

        # プロジェクト数を安全に計算
        project_ids = set()
        for task in all_tasks:
            try:
                pid = task.get('project_id')
                if pid is not None:
                    project_ids.add(pid)
            except Exception:
                pass

        return {
            'total_tasks': total_tasks,
            'completed_tasks': completed_tasks,
            'pending_tasks': pending_tasks,
            'projects_count': len(project_ids)
        }

    except Exception as stats_error:
        logging.error(f"Error calculating stats: {str(stats_error)}")
        return {
            'total_tasks': len(all_tasks) if all_tasks else 0,
            'completed_tasks': 0,
            'pending_tasks': len(all_tasks) if all_tasks else 0,
            'projects_count': 0
        }

@colrep_bp.route('/worker_dashboard')
@login_required
def worker_dashboard():
    """入力者用ダッシュボード（ステータス表示対応版）"""
    user_info = get_user_info()
    user_id = user_info.get('user_id')

    if not user_id:
        flash('ログインが必要です。', 'error')
        return redirect(url_for('auth.login'))
    # ★追加: クイック作成ボタン・担当管理ボタンの表示判定用
    permissions = get_user_permissions()

    try:
        # conn = mysql.connector.connect(**base_db_config, database=TARGET_DATABASE)
        conn = mysql.connector.connect(**DatabaseConfig.fujinp())
        cursor = conn.cursor(dictionary=True)

        # ユーザーにアサインされたタスクを全プロジェクトから取得
        cursor.execute(f"""
            SELECT p.id as project_id, p.プロジェクト名, p.テーブル名, p.is_public, p.責任者
            FROM {COLREP_PROJECTS_TABLE} p
            WHERE p.テーブル名 IS NOT NULL AND p.テーブル名 != ''
        """)

        projects = cursor.fetchall()
        all_tasks = []

        logging.info(f"Worker dashboard - User ID: {user_id}")
        logging.info(f"Found {len(projects)} projects to check")

        for project in projects:
            table_name = project['テーブル名']
            project_id = project['project_id']

            try:
                # テーブルが存在するかチェック
                cursor.execute(f"SHOW TABLES LIKE '{table_name}'")
                if not cursor.fetchone():
                    logging.warning(f"Table {table_name} does not exist")
                    continue

                # ステータスカラムを追加で取得
                cursor.execute(f"""
                    SELECT id, 更新日時, カラム名, 説明, content, 備考, 担当者アカウント,
                           status
                    FROM `{table_name}`
                    WHERE 担当者アカウント = %s
                    ORDER BY 更新日時 DESC
                """, (user_id,))

                tasks = cursor.fetchall()

                logging.info(f"Project {project['プロジェクト名']} ({table_name}): Found {len(tasks)} tasks for user {user_id}")

                for task in tasks:
                    try:
                        # safe_process_contentで統一的に処理
                        content = safe_process_content(task.get('content'))
                        remarks = safe_process_content(task.get('備考'))
                        column_name = safe_process_content(task.get('カラム名', ''))
                        description = safe_process_content(task.get('説明', ''))

                        # ステータスの取得（デフォルト値は '作業中'）
                        status = task.get('status', '作業中')
                        if not status:
                            status = '作業中'

                        # タスク情報を構築
                        task_info = {
                            'id': task.get('id'),
                            'project_id': project_id,
                            'プロジェクト名': project['プロジェクト名'],
                            'テーブル名': table_name,
                            '更新日時': task.get('更新日時'),
                            'カラム名': column_name,
                            '説明': description,
                            'content': content,
                            '備考': remarks,
                            '担当者アカウント': task.get('担当者アカウント'),
                            'status': status,
                            'is_public': project.get('is_public', False),
                            'is_owner': project.get('責任者') == user_id
                        }

                        all_tasks.append(task_info)

                    except Exception as task_error:
                        logging.error(f"Error processing task {task.get('id', 'unknown')}: {str(task_error)}")
                        error_task_info = {
                            'id': task.get('id', 0),
                            'project_id': project_id,
                            'プロジェクト名': project['プロジェクト名'],
                            'テーブル名': table_name,
                            '更新日時': task.get('更新日時'),
                            'カラム名': 'データ処理エラー',
                            '説明': '',
                            'content': '',
                            '備考': '',
                            '担当者アカウント': task.get('担当者アカウント'),
                            'status': '作業中'
                        }
                        all_tasks.append(error_task_info)

            except Exception as table_error:
                logging.error(f"Error accessing table {table_name}: {str(table_error)}")
                continue

        # 統計情報を安全に計算
        stats = calculate_safe_statistics(all_tasks)

        logging.info(f"Final stats: {stats}")
        logging.info(f"Total tasks retrieved: {len(all_tasks)}")

        return render_template('worker_dashboard.html',
                              tasks=all_tasks,
                              stats=stats,
                              user_info=user_info,
                              permissions=permissions)

    except Exception as e:
        logging.error(f"Error in worker_dashboard: {str(e)}")
        return render_template('worker_dashboard.html',
                              tasks=[],
                              stats={'total_tasks': 0, 'completed_tasks': 0, 'pending_tasks': 0, 'projects_count': 0},
                              user_info=user_info,
                              permissions=permissions,
                              error_message=f'データ取得中にエラーが発生しました: {str(e)}')
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close()
            conn.close()


# ============================================================
# 2. ステータス更新エンドポイント（新規追加）
# ============================================================

@colrep_bp.route('/update_task_status/<int:project_id>/<int:task_id>', methods=['POST'])
@login_required
def update_task_status(project_id, task_id):
    """タスクのステータスを更新（作業者用）"""
    user_info = get_user_info()
    user_id = user_info.get('user_id')

    data = request.get_json()
    new_status = data.get('status', '').strip()

    # ステータス値の検証
    valid_statuses = ['作業中', '改訂中', '完了']
    if new_status not in valid_statuses:
        return jsonify({'success': False, 'error': '無効なステータスです'}), 400

    try:
        # conn = mysql.connector.connect(**base_db_config, database=TARGET_DATABASE)
        conn = mysql.connector.connect(**DatabaseConfig.fujinp())
        cursor = conn.cursor(dictionary=True)

        # プロジェクト情報を取得
        cursor.execute(f"SELECT テーブル名 FROM {COLREP_PROJECTS_TABLE} WHERE id = %s", (project_id,))
        project = cursor.fetchone()

        if not project:
            return jsonify({'success': False, 'error': 'プロジェクトが見つかりません。'}), 404

        table_name = project['テーブル名']

        # タスク所有者チェック（自分のタスクのみ更新可能）
        cursor.execute(f"""
            SELECT id FROM `{table_name}`
            WHERE id = %s AND 担当者アカウント = %s
        """, (task_id, user_id))

        if not cursor.fetchone():
            return jsonify({'success': False, 'error': '権限がありません。'}), 403

        # ステータスを更新
        now = get_jst_now()
        update_query = f"""
            UPDATE `{table_name}`
            SET status = %s, 更新日時 = %s
            WHERE id = %s AND 担当者アカウント = %s
        """

        cursor.execute(update_query, (new_status, now, task_id, user_id))
        conn.commit()

        return jsonify({
            'success': True,
            'message': 'ステータスを更新しました。'
        })

    except Exception as e:
        logging.error(f"Error in update_task_status: {str(e)}")
        if 'conn' in locals() and conn.is_connected():
            conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close()
            conn.close()


# ============================================================
# 3. 作業者用統合プレビューエンドポイント（新規追加）
# ============================================================

@colrep_bp.route('/get_integrated_preview/<int:project_id>')
@login_required
def get_integrated_preview(project_id):
    """統合プレビューをプレーン HTML で返す"""
    user_info = get_user_info()
    user_id = user_info.get('user_id')

    try:
        # ✅ 修正: アクセス権限をチェック
        has_access, is_author = check_project_access_permission(project_id, user_id)
        if not has_access:
            return Response("<p>このプロジェクトへのアクセス権限がありません。</p>", mimetype='text/html'), 403

        # conn = mysql.connector.connect(**base_db_config, database=TARGET_DATABASE)
        conn = mysql.connector.connect(**DatabaseConfig.fujinp())
        cursor = conn.cursor(dictionary=True)

        cursor.execute(f"""
            SELECT p.id, p.プロジェクト名, p.テーブル名, p.Composer, p.is_public
            FROM {COLREP_PROJECTS_TABLE} p
            WHERE p.id = %s
        """, (project_id,))

        project = cursor.fetchone()
        if not project:
            return Response("<p>プロジェクトが見つかりません。</p>", mimetype='text/html'), 404

        table_name = project['テーブル名']
        cursor.execute(f"SHOW TABLES LIKE '{table_name}'")
        if not cursor.fetchone():
            return Response("<p>プロジェクトテーブルが見つかりません。</p>", mimetype='text/html'), 404

        if not project.get('Composer'):
            return Response("<p>統合指示書（Composer）がまだ作成されていません。</p>", mimetype='text/html'), 400

        cursor.execute(f"""
            SELECT t.*,
               (SELECT full_name FROM {Tables.USERS} WHERE id = t.担当者アカウント) as full_name
            FROM `{table_name}` t
            ORDER BY t.id ASC
        """)
        raw_data = cursor.fetchall()

        integrated_markdown = _process_composer_integration(
            project['Composer'],
            raw_data,
            cursor,
            table_name
        )

        # ✅ 修正：拡張機能を削除してシンプルに
        try:
            # 拡張機能を最小限に
            integrated_html = process_markdown(integrated_markdown)
            #integrated_html = markdown.markdown(integrated_markdown, extensions=['extra', 'nl2br', 'sane_lists', 'fenced_code'])
        except Exception as e:
            logging.error(f"Markdown変換エラー: {str(e)}")
            import html as html_module
            escaped_content = html_module.escape(integrated_markdown)
            integrated_html = f"<pre style='white-space: pre-wrap; font-family: inherit; background: #f5f5f5; padding: 15px; border-radius: 4px;'>{escaped_content}</pre>"

        return Response(integrated_html, mimetype='text/html')

    except Exception as e:
        logging.error(f"Error in get_integrated_preview: {str(e)}")
        error_html = f"<p style='color: red;'>エラーが発生しました: {str(e)}</p>"
        return Response(error_html, mimetype='text/html'), 500
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close()
            conn.close()

@colrep_bp.route('/save_content_external/<int:project_id>/<int:task_id>', methods=['POST'])
@login_required
def save_content_external(project_id, task_id):
    """外部エディタからの作業内容保存"""
    user_info = get_user_info()
    user_id = user_info.get('user_id')

    data = request.get_json()
    content = data.get('content', '').strip()
    remarks = data.get('remarks', '').strip()

    try:
        # conn = mysql.connector.connect(**base_db_config, database=TARGET_DATABASE)
        conn = mysql.connector.connect(**DatabaseConfig.fujinp())
        cursor = conn.cursor(dictionary=True)

        # プロジェクト情報を取得
        cursor.execute(f"SELECT テーブル名 FROM {COLREP_PROJECTS_TABLE} WHERE id = %s", (project_id,))
        project = cursor.fetchone()

        if not project:
            return jsonify({'success': False, 'error': 'プロジェクトが見つかりません。'}), 404

        table_name = project['テーブル名']

        # 担当者チェック
        cursor.execute(f"SELECT id FROM `{table_name}` WHERE id = %s AND 担当者アカウント = %s",
                      (task_id, user_id))
        if not cursor.fetchone():
            return jsonify({'success': False, 'error': '権限がありません。'}), 403

        # 内容を更新
        now = get_jst_now()
        update_query = f"""
            UPDATE `{table_name}`
            SET 更新日時 = %s, content = %s, 備考 = %s
            WHERE id = %s AND 担当者アカウント = %s
        """

        cursor.execute(update_query, (now, content, remarks, task_id, user_id))
        conn.commit()

        return jsonify({
            'success': True,
            'message': '作業内容を保存しました。'
        })

    except Exception as e:
        logging.error(f"Error in save_content_external: {str(e)}")
        if 'conn' in locals() and conn.is_connected():
            conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close()
            conn.close()


@colrep_bp.app_template_filter('truncate')
def truncate_filter(text, length=100, suffix='...'):
    """テキスト切り詰めフィルター"""
    if text and len(text) > length:
        return text[:length] + suffix
    return text or ""

# colrep routes.py に追加するComposer関連ルート

@colrep_bp.route('/edit_composer_external/<int:project_id>')
@login_required
def edit_composer_external(project_id):
    """Composer外部エディタ"""
    user_info = get_user_info()
    user_id = user_info.get('user_id')
    user_category = session.get('user_category') # ★セッションからカテゴリ取得

    try:
        # conn = mysql.connector.connect(**base_db_config, database=TARGET_DATABASE)
        conn = mysql.connector.connect(**DatabaseConfig.fujinp())
        cursor = conn.cursor(dictionary=True)

        # プロジェクト情報を取得
        cursor.execute(f"""
            SELECT p.*, u.full_name as 責任者名
            FROM {COLREP_PROJECTS_TABLE} p
            LEFT JOIN {Tables.USERS} u ON p.責任者 = u.id
            WHERE p.id = %s
        """, (project_id,))

        project = cursor.fetchone()
        if not project:
            return render_template('error.html', error='プロジェクトが見つかりません。')

        # ★★★ 権限チェックの修正箇所 ★★★
        # 「システムAdminではない」かつ「プロジェクト責任者でもない」場合に拒否する
        is_system_admin = (user_category == 'admin')
        is_owner = (project['責任者'] == user_id)

        if not (is_system_admin or is_owner):
            return render_template('error.html',
                                  error='Composerの編集はプロジェクト責任者またはシステム管理者のみ実行可能です。')
        # ★★★★★★★★★★★★★★★★★★★★★

        current_date = get_jst_now().strftime('%Y-%m-%d')

        # Composer外部エディタページを表示
        return render_template('colrep_composer_external.html',
                               project=project,
                               project_id=project_id,
                               current_date=current_date)

    except Exception as e:
        logging.error(f"Error in edit_composer_external: {str(e)}")
        return render_template('error.html', error=str(e))
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close()
            conn.close()

@colrep_bp.route('/save_composer_external/<int:project_id>', methods=['POST'])
@login_required
def save_composer_external(project_id):
    """Composer外部エディタからの保存"""
    user_info = get_user_info()
    user_id = user_info.get('user_id')
    user_category = session.get('user_category') # ★システム管理者判定用

    data = request.get_json()
    composer_content = data.get('composer', '').strip()

    try:
        # conn = mysql.connector.connect(**base_db_config, database=TARGET_DATABASE)
        conn = mysql.connector.connect(**DatabaseConfig.fujinp())
        cursor = conn.cursor(dictionary=True)

        # プロジェクト情報を取得
        cursor.execute(f"SELECT 責任者 FROM {COLREP_PROJECTS_TABLE} WHERE id = %s", (project_id,))
        project = cursor.fetchone()

        if not project:
            return jsonify({'success': False, 'error': 'プロジェクトが見つかりません。'}), 404

        # ★★★ 修正箇所1: 権限チェックの緩和 ★★★
        is_system_admin = (user_category == 'admin')
        is_owner = (project['責任者'] == user_id)

        if not (is_system_admin or is_owner):
            return jsonify({'success': False, 'error': '権限がありません。'}), 403
        # ★★★★★★★★★★★★★★★★★★★★★★★

        # Composerを更新
        now = get_jst_now()

        # ★★★ 修正箇所2: SQLのWHERE句を変更 ★★★
        # adminの場合は責任者IDが異なっても更新できるように、条件を id のみにする
        if is_system_admin:
            update_query = f"""
                UPDATE {COLREP_PROJECTS_TABLE}
                SET Composer = %s, 更新日時 = %s
                WHERE id = %s
            """
            cursor.execute(update_query, (composer_content, now, project_id))
        else:
            # 一般ユーザーは従来通り責任者IDの一致を確認
            update_query = f"""
                UPDATE {COLREP_PROJECTS_TABLE}
                SET Composer = %s, 更新日時 = %s
                WHERE id = %s AND 責任者 = %s
            """
            cursor.execute(update_query, (composer_content, now, project_id, user_id))
        # ★★★★★★★★★★★★★★★★★★★★★

        conn.commit()

        return jsonify({
            'success': True,
            'message': 'Composerを保存しました。'
        })

    except Exception as e:
        logging.error(f"Error in save_composer_external: {str(e)}")
        if 'conn' in locals() and conn.is_connected():
            conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close()
            conn.close()

@colrep_bp.route('/get_composer/<int:project_id>')
@login_required
def get_composer(project_id):
    """Composer内容取得API（プレビュー用）"""
    user_info = get_user_info()
    user_id = user_info.get('user_id')

    try:
        # conn = mysql.connector.connect(**base_db_config, database=TARGET_DATABASE)
        conn = mysql.connector.connect(**DatabaseConfig.fujinp())
        cursor = conn.cursor(dictionary=True)

        # プロジェクト情報を取得
        cursor.execute(f"""
            SELECT プロジェクト名, Composer, 責任者
            FROM {COLREP_PROJECTS_TABLE}
            WHERE id = %s
        """, (project_id,))

        project = cursor.fetchone()
        if not project:
            return jsonify({'success': False, 'error': 'プロジェクトが見つかりません。'}), 404

        # プロジェクト責任者または総管理者のみ表示可能
        permissions = get_user_permissions()
        if not (project['責任者'] == user_id or permissions['is_colrep_admin']):
            return jsonify({'success': False, 'error': '権限がありません。'}), 403

        # ComposerコンテンツをHTMLに変換
        html_content = ""
        if project['Composer']:
            html_content = process_markdown_for_preview(project['Composer'])
        else:
            html_content = "<p class='text-muted'>Composerがまだ作成されていません</p>"

        return jsonify({
            'success': True,
            'project_name': project['プロジェクト名'],
            'md_content': project['Composer'] or '',
            'html_content': html_content
        })

    except Exception as e:
        logging.error(f"Error in get_composer: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close()
            conn.close()

@colrep_bp.route('/export_project/<int:project_id>')
@login_required
def export_project(project_id):
    """プロジェクト統合エクスポート（将来の機能）"""
    user_info = get_user_info()
    user_id = user_info.get('user_id')

    try:
        # conn = mysql.connector.connect(**base_db_config, database=TARGET_DATABASE)
        conn = mysql.connector.connect(**DatabaseConfig.fujinp())
        cursor = conn.cursor(dictionary=True)

        # プロジェクト情報を取得
        cursor.execute(f"""
            SELECT p.*, u.full_name as 責任者名
            FROM {COLREP_PROJECTS_TABLE} p
            LEFT JOIN {Tables.USERS} u ON p.責任者 = u.id
            WHERE p.id = %s
        """, (project_id,))

        project = cursor.fetchone()
        if not project:
            flash('プロジェクトが見つかりません。', 'error')
            return redirect(url_for('colrep.index'))

        # 権限チェック：プロジェクト責任者または総管理者
        permissions = get_user_permissions()
        if not (project['責任者'] == user_id or permissions['is_colrep_admin']):
            flash('エクスポート権限がありません。', 'error')
            return redirect(url_for('colrep.view_project', project_id=project_id))

        # テーブルが存在するかチェック
        table_name = project['テーブル名']
        cursor.execute(f"SHOW TABLES LIKE '{table_name}'")
        if not cursor.fetchone():
            flash('プロジェクトテーブルが存在しません。', 'error')
            return redirect(url_for('colrep.view_project', project_id=project_id))

        # すべてのタスクデータを取得
        cursor.execute(f"""
            SELECT t.*, u.full_name, u.full_name
            FROM `{table_name}` t
            LEFT JOIN {Tables.USERS} u ON t.担当者アカウント = u.id
            ORDER BY t.id ASC
        """)
        tasks = cursor.fetchall()

        # Composer情報も取得
        composer_content = project.get('Composer', '')

        # 統合データの作成（ここでComposerの指示に従ってデータを統合）
        # 現時点では簡単な一覧形式で出力
        export_data = {
            'project': project,
            'composer': composer_content,
            'tasks': tasks,
            'export_date': get_jst_now().strftime('%Y-%m-%d %H:%M:%S')
        }

        # 将来的にはComposerの指示に従って様々な形式でエクスポート可能
        # 現在は簡易的にJSONで返す
        return jsonify(serialize_for_json(export_data))

    except Exception as e:
        logging.error(f"Error in export_project: {str(e)}")
        flash(f'エクスポートに失敗しました: {str(e)}', 'error')
        return redirect(url_for('colrep.view_project', project_id=project_id))
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close()
            conn.close()

# colrep routes.py に追加する部分

# ファイル上部のインポートに追加


@colrep_bp.route('/get_task_content/<int:project_id>/<int:task_id>')
@login_required
def get_task_content(project_id, task_id):
    """タスク内容をJSON で返す（外部エディタ用）"""
    user_info = get_user_info()
    user_id = user_info.get('user_id')

    try:
        # conn = mysql.connector.connect(**base_db_config, database=TARGET_DATABASE)
        conn = mysql.connector.connect(**DatabaseConfig.fujinp())
        cursor = conn.cursor(dictionary=True)

        # プロジェクト情報を取得
        cursor.execute(f"""
            SELECT p.テーブル名, p.プロジェクト名
            FROM {COLREP_PROJECTS_TABLE} p
            WHERE p.id = %s
        """, (project_id,))

        project = cursor.fetchone()
        if not project:
            return jsonify({
                'success': False,
                'error': 'プロジェクトが見つかりません。'
            }), 404

        table_name = project['テーブル名']
        project_name = project['プロジェクト名']

        # タスク情報を取得（担当者チェック付き）
        cursor.execute(f"""
            SELECT カラム名, content, 備考 FROM `{table_name}`
            WHERE id = %s AND 担当者アカウント = %s
        """, (task_id, user_id))

        task = cursor.fetchone()
        if not task:
            return jsonify({
                'success': False,
                'error': 'タスクが見つかりません。'
            }), 404

        return jsonify({
            'success': True,
            'task_title': task['カラム名'],
            'project_name': project_name,
            'md_content': task['content'] or '',
            'remarks': task['備考'] or ''
        })

    except Exception as e:
        logging.error(f"Error in get_task_content: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close()
            conn.close()




# [2026-07-25 改修] 追加: /get_task_content_admin
# manage_project_table.html の「本文」「ソース」ボタンが呼ぶAPIが未実装だったため新規追加。
# 担当者本人向けの /get_task_content と異なり、admin／プロジェクト責任者が
# 担当者を問わず任意のタスクの本文を参照できる（閲覧のみ・更新機能は持たない）。
@colrep_bp.route('/get_task_content_admin/<int:project_id>/<int:task_id>')
@login_required
def get_task_content_admin(project_id, task_id):
    """タスク本文の取得（admin／プロジェクト責任者用・閲覧のみ）"""
    user_id = session.get('user_id')
    conn = None
    try:
        conn = mysql.connector.connect(**DatabaseConfig.fujinp())
        cursor = conn.cursor(dictionary=True)

        cursor.execute(
            f"SELECT テーブル名, 責任者 FROM {COLREP_PROJECTS_TABLE} WHERE id = %s",
            (project_id,))
        project = cursor.fetchone()
        if not project:
            return jsonify({'success': False, 'error': 'プロジェクトが見つかりません'}), 404

        # 権限チェック：システム管理者 または プロジェクト責任者
        is_system_admin = (session.get('user_category') == 'admin')
        if not (is_system_admin or project['責任者'] == user_id):
            return jsonify({'success': False,
                            'error': 'この操作はプロジェクト責任者または管理者のみ実行可能です'}), 403

        table_name = project['テーブル名']
        cursor.execute(f"SHOW TABLES LIKE '{table_name}'")
        if not cursor.fetchone():
            return jsonify({'success': False,
                            'error': 'プロジェクトテーブルがまだ作成されていません'}), 404

        cursor.execute(
            f"SELECT id, カラム名, content, 備考, status, 更新日時 FROM `{table_name}` WHERE id = %s",
            (task_id,))
        task = cursor.fetchone()
        if not task:
            return jsonify({'success': False, 'error': 'タスクが見つかりません'}), 404

        return jsonify(serialize_for_json({
            'success': True,
            'task_id': task['id'],
            'task_title': task['カラム名'],
            'md_content': task.get('content') or '',
            'remarks': task.get('備考') or '',
            'status': task.get('status') or '作業中',
            '更新日時': task.get('更新日時'),
        }))

    except Exception as e:
        logging.error(f"Error in get_task_content_admin: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if conn is not None and conn.is_connected():
            cursor.close()
            conn.close()


@colrep_bp.route('/edit_content_external/<int:project_id>/<int:task_id>')
@login_required
def edit_content_external(project_id, task_id):
    """外部エディタでの作業内容編集"""
    user_info = get_user_info()
    user_id = user_info.get('user_id')

    try:
        # conn = mysql.connector.connect(**base_db_config, database=TARGET_DATABASE)
        conn = mysql.connector.connect(**DatabaseConfig.fujinp())
        cursor = conn.cursor(dictionary=True)

        # プロジェクト情報を取得
        cursor.execute(f"""
            SELECT p.プロジェクト名, p.テーブル名
            FROM {COLREP_PROJECTS_TABLE} p
            WHERE p.id = %s
        """, (project_id,))

        project = cursor.fetchone()
        if not project:
            flash('プロジェクトが見つかりません。', 'error')
            return render_template('error.html', error='プロジェクトが見つかりません。')

        table_name = project['テーブル名']

        # タスク情報を取得（担当者チェック付き）
        cursor.execute(f"""
            SELECT * FROM `{table_name}`
            WHERE id = %s AND 担当者アカウント = %s
        """, (task_id, user_id))

        task = cursor.fetchone()
        if not task:
            return render_template('error.html', error='タスクが見つからないか、あなたにアサインされていません。')

        # medit external エディタページを表示
        return render_template('colrep_external_editor.html',
                              project=project,
                              task=task,
                              project_id=project_id,
                              task_id=task_id)

    except Exception as e:
        logging.error(f"Error in edit_content_external: {str(e)}")
        return render_template('error.html', error=str(e))
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close()
            conn.close()



# 既存のedit_taskルートを削除するか、コメントアウト
# 新しいワークフローでは使用しないため
# colrep routes.py に追加するタスク管理関連ルート

# [2026-07-25 改修] 削除: /update_task_description
# どのテンプレートからも呼ばれていないデッドルートだったため削除。
# タスク説明の更新は /save_description_external（責任者）または /update_task（admin/責任者）で行う。

@colrep_bp.route('/delete_task/<int:project_id>/<int:task_id>', methods=['POST'])
@login_required
def delete_task(project_id, task_id):
    """タスク削除（プロジェクト責任者のみ）"""
    user_info = get_user_info()
    user_id = user_info.get('user_id')
    permissions = get_user_permissions()

    try:
        # conn = mysql.connector.connect(**base_db_config, database=TARGET_DATABASE)
        conn = mysql.connector.connect(**DatabaseConfig.fujinp())
        cursor = conn.cursor(dictionary=True)

        # プロジェクト情報を取得
        cursor.execute(f"SELECT テーブル名, 責任者 FROM {COLREP_PROJECTS_TABLE} WHERE id = %s", (project_id,))
        project = cursor.fetchone()

        if not project:
            return jsonify({'success': False, 'error': 'プロジェクトが見つかりません。'}), 404

        # 権限チェック：システム管理者、CoRePo総管理者、または責任者
        is_system_admin = (session.get('user_category') == 'admin')
        is_colrep_admin = permissions['is_colrep_admin']
        is_owner = (project['責任者'] == user_id)

        if not (is_system_admin or is_colrep_admin or is_owner):
            return jsonify({'success': False, 'error': '権限がありません。'}), 403

        table_name = project['テーブル名']

        # タスクを削除
        cursor.execute(f"DELETE FROM `{table_name}` WHERE id = %s", (task_id,))
        conn.commit()

        return jsonify({
            'success': True,
            'message': 'タスクを削除しました。'
        })

    except Exception as e:
        logging.error(f"Error in delete_task: {str(e)}")
        if 'conn' in locals() and conn.is_connected():
            conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close()
            conn.close()

@colrep_bp.route('/preview_composer', methods=['POST'])
@login_required
def preview_composer():
    """Composerプレビュー（統合処理版）"""
    data = request.get_json()
    markdown_text = data.get('markdown', '')

    # リクエストヘッダーまたはセッションからproject_idを取得
    # フロントエンドから送信するか、URLパラメータで取得
    project_id = data.get('project_id')  # フロントエンドから送信

    if not project_id:
        # フォールバック：従来の処理（マクロ展開なし。自分が書いた文字列を変換するだけ）
        html = process_markdown_for_preview(markdown_text)
        return jsonify({'html': html})

    # [2026-07-25 改修] 権限チェックを追加
    # project_id を指定すると [[カラム名]] マクロが実データで展開されるため、
    # 従来は project_id を知っていれば誰でも部品の中身を取得できた。
    has_access, _is_author = check_project_access_permission(project_id, session.get('user_id'))
    if not has_access:
        return jsonify({
            'html': '<p style="color: red;">このプロジェクトのプレビュー権限がありません。</p>',
            'error': 'forbidden'
        }), 403

    try:
        # conn = mysql.connector.connect(**base_db_config, database=TARGET_DATABASE)
        conn = mysql.connector.connect(**DatabaseConfig.fujinp())
        cursor = conn.cursor(dictionary=True)

        # プロジェクト情報を取得
        cursor.execute(f"SELECT テーブル名 FROM {COLREP_PROJECTS_TABLE} WHERE id = %s", (project_id,))
        project = cursor.fetchone()

        if not project:
            html = process_markdown_for_preview(markdown_text)
            return jsonify({'html': html})

        table_name = project['テーブル名']

        # タスクデータを取得
        cursor.execute(f"""
            SELECT t.*, u.full_name, u.full_name
            FROM `{table_name}` t
            LEFT JOIN {Tables.USERS} u ON t.担当者アカウント = u.id
            ORDER BY t.id ASC
        """)
        raw_data = cursor.fetchall()

        # 統合プレビューと同じ処理を使用
        integrated_markdown = _process_composer_integration(
            markdown_text, raw_data, cursor, table_name
        )

        html = process_markdown_for_preview(integrated_markdown)
        return jsonify({'html': html})

    except Exception as e:
        logging.error(f"Composerプレビューエラー: {str(e)}")
        # エラー時は従来の処理にフォールバック
        html = process_markdown_for_preview(markdown_text)
        return jsonify({'html': html})
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close()
            conn.close()

# 既存のcolrep routes.pyファイルの末尾に以下のコードを追加


@colrep_bp.route('/preview_integrated/<int:project_id>')
@login_required
def preview_integrated(project_id):
    """統合コンテンツプレビュー（作業者側と同じシンプルなスタイル）"""
    user_info = get_user_info()
    user_id = user_info.get('user_id')
    user_category = session.get('user_category') # ★追加: セッションからカテゴリを取得

    try:
        # conn = mysql.connector.connect(**base_db_config, database=TARGET_DATABASE)
        conn = mysql.connector.connect(**DatabaseConfig.fujinp())
        cursor = conn.cursor(dictionary=True)

        # プロジェクト情報を取得
        cursor.execute(f"""
            SELECT p.*, u.full_name as 責任者名
            FROM {COLREP_PROJECTS_TABLE} p
            LEFT JOIN {Tables.USERS} u ON p.責任者 = u.id
            WHERE p.id = %s
        """, (project_id,))

        project = cursor.fetchone()
        if not project:
            return Response("<p>プロジェクトが見つかりません。</p>", mimetype='text/html'), 404

        # 権限チェック：プロジェクト責任者または総管理者
        # ★★★ ここを修正：権限チェックにシステムAdminを追加 ★★★
        permissions = get_user_permissions()
        is_system_admin = (user_category == 'admin') # ★追加
        is_colrep_admin = permissions['is_colrep_admin']
        is_owner = (project['責任者'] == user_id)

        # いずれかの管理者権限、または本人のプロジェクトなら通過
        if not (is_system_admin or is_colrep_admin or is_owner):
            return Response("<p>統合プレビューの表示権限がありません。</p>", mimetype='text/html'), 403
        # ★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★

        table_name = project['テーブル名']

        # テーブルが存在するかチェック
        cursor.execute(f"SHOW TABLES LIKE '{table_name}'")
        table_exists = cursor.fetchone() is not None

        if not table_exists:
            return Response("<p>プロジェクトテーブルがまだ作成されていません。</p>", mimetype='text/html'), 400

        # Composerが存在するかチェック
        if not project.get('Composer'):
            return Response("<p>統合指示書（Composer）がまだ作成されていません。</p>", mimetype='text/html'), 400

        # すべてのタスクデータを取得
        cursor.execute(f"""
            SELECT t.*, u.full_name, u.full_name
            FROM `{table_name}` t
            LEFT JOIN {Tables.USERS} u ON t.担当者アカウント = u.id
            ORDER BY t.id ASC
        """)
        raw_data = cursor.fetchall()

        # Composerの指示を解析して統合コンテンツを生成
        integrated_markdown = _process_composer_integration(
            project['Composer'],
            raw_data,
            cursor,
            table_name
        )

        # ✅ markdown.markdown() で変換
        try:
            #integrated_html = markdown.markdown(
            #    integrated_markdown,
            #    extensions=['extra', 'nl2br', 'sane_lists', 'fenced_code']
            #)
            integrated_html = process_markdown(integrated_markdown)
        except Exception as e:
            logging.error(f"Markdown変換エラー: {str(e)}")
            import html as html_module
            escaped_content = html_module.escape(integrated_markdown)
            integrated_html = f"<pre style='white-space: pre-wrap;'>{escaped_content}</pre>"

        # ✅ 修正：作業者側と同じシンプルなHTMLを返す
        complete_html = f"""<!DOCTYPE html>
<html lang="ja">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{project['プロジェクト名']} - 統合プレビュー</title>
    {_math_diagram_assets()}
</head>
<body>
    <div id="integrated-content">
        {integrated_html}
    </div>
</body>
</html>"""

        return Response(complete_html, mimetype='text/html')

    except Exception as e:
        logging.error(f"Error in preview_integrated: {str(e)}")
        error_html = f"<p style='color: red;'>エラーが発生しました: {str(e)}</p>"
        return Response(error_html, mimetype='text/html'), 500
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close()
            conn.close()

def _math_diagram_assets():
    """[2026-07-25 改修]
    サーバ側で組み立てるプレビューHTML用の KaTeX / Mermaid 読み込み断片。
    従来これらを読み込んでいなかったため、数式（$...$）や mermaid 図が
    エディタ内プレビューでしか描画されなかった。外部エディタと同じCDN・同じ設定を使う。"""
    return """
    <link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/katex@0.16.9/dist/katex.min.css">
    <style>
        body { font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", "Hiragino Sans",
               "Noto Sans JP", sans-serif; line-height: 1.7; margin: 0; }
        /* 上部の固定バー（共有プレビュー）を全幅のまま保つため、余白は本文側に付ける */
        #integrated-content { max-width: 960px; margin: 0 auto; padding: 24px; }
        #integrated-content img, #integrated-content svg { max-width: 100%; height: auto; }
        .katex { font-size: 1.1em !important; }
        .mermaid { text-align: center; }
        table { border-collapse: collapse; }
        table, th, td { border: 1px solid #ccc; }
        th, td { padding: 6px 10px; }
        pre { background: #f5f5f5; padding: 12px; border-radius: 4px; overflow-x: auto; }
    </style>
    <script defer src="https://cdn.jsdelivr.net/npm/katex@0.16.9/dist/katex.min.js"></script>
    <script defer src="https://cdn.jsdelivr.net/npm/katex@0.16.9/dist/contrib/auto-render.min.js"></script>
    <script src="https://cdn.jsdelivr.net/npm/mermaid@10.9.0/dist/mermaid.min.js"></script>
    <script>
    window.addEventListener('load', function () {
        var root = document.getElementById('integrated-content') || document.body;
        // ```mermaid のコードブロックを .mermaid 要素に変換（Markdown変換系の違いを吸収）
        root.querySelectorAll('pre > code.language-mermaid, pre > code.mermaid').forEach(function (code) {
            var div = document.createElement('div');
            div.className = 'mermaid';
            div.textContent = code.textContent;
            code.parentNode.replaceWith(div);
        });
        function renderMath() {
            if (!window.renderMathInElement) return;
            renderMathInElement(root, {
                delimiters: [
                    { left: '$$', right: '$$', display: true },
                    { left: '$', right: '$', display: false }
                ],
                throwOnError: false
            });
        }
        // Mermaid の描画完了後に KaTeX を走らせる（図の定義中の $ を誤変換しないため）
        if (window.mermaid) {
            try {
                mermaid.initialize({ startOnLoad: false });
                var p = mermaid.run({ nodes: root.querySelectorAll('.mermaid') });
                if (p && typeof p.then === 'function') { p.then(renderMath, renderMath); }
                else { renderMath(); }
            } catch (e) { console.error('mermaid error', e); renderMath(); }
        } else {
            renderMath();
        }
    });
    </script>"""


def _process_composer_integration(composer_content, raw_data, cursor, table_name):
    """Composer内のマクロを実際のコンテンツで置換"""
    result = composer_content

    # パターン 1: [[カラム名]] 形式のマクロ（新規追加）
    macro_pattern = r'\[\[([^\]]+)\]\]'

    def replace_macro(match):
        column_name = match.group(1).strip()
        # raw_data から該当するカラム名のデータを検索
        matching_tasks = [task for task in raw_data if task['カラム名'] == column_name]
        if matching_tasks:
            content = matching_tasks[0].get('content')
            # NULL または空の場合は空文字列を返す
            return content if content else ''
        # マッチしない場合も空文字列を返す
        return ''

    result = re.sub(macro_pattern, replace_macro, result)

    # パターン 2: 既存の SQL クエリパターン（互換性のため）
    sql_pattern = r'```sql_md_texts\s+select\s+content\s+from\s+' + re.escape(table_name) + r'\s+where\s+カラム名\s*=\s*[\'"]([^\'\"]+)[\'\"]\s*;\s*```'

    def replace_sql(match):
        column_name = match.group(1)
        matching_tasks = [task for task in raw_data if task['カラム名'] == column_name]
        if matching_tasks:
            return matching_tasks[0].get('content', '')
        return ''

    result = re.sub(sql_pattern, replace_sql, result, flags=re.IGNORECASE | re.MULTILINE | re.DOTALL)

    return result

def _canvas_ref_bar():
    """キャンバス参照バー（全閲覧者向け・自己完結JS付き）のHTML断片を返す。

    あわみ（/awami/api/find_by_url）とあわなら（/awanara/api/find_by_url）の
    両方を照会し，このプレビューを実体URLとして参照しているキャンバスを列挙する。
    片方のアプリが未導入でも他方の結果だけで動く（失敗は空扱い）。
    2026-07-19: あわなら専用バーをあわみ対応に拡張（旧 _awanara_ref_bar）。
    """
    return """
    <div style="position: sticky; top: 0; z-index: 9; background: #eef4fb;
                border-bottom: 1px solid #cfe0f2; padding: 8px 16px;
                font-family: sans-serif; display: flex; align-items: center;
                gap: 12px; flex-wrap: wrap;">
        <button id="cvRefBtn" onclick="findCanvasRefs()"
           style="background:#35618f;color:#fff;border:none;padding:6px 14px;
                  border-radius:6px;font-size:14px;cursor:pointer;">
            \U0001F578 キャンバスからの参照
        </button>
        <span id="cvRefMsg" style="color:#5a6672;font-size:13px;"></span>
        <div id="cvRefList" style="display:none;flex-wrap:wrap;gap:6px;"></div>
    </div>
    <script>
    function findCanvasRefs() {
        var btn = document.getElementById('cvRefBtn');
        var msg = document.getElementById('cvRefMsg');
        var list = document.getElementById('cvRefList');
        list.style.display = 'none'; list.innerHTML = '';
        msg.textContent = '検索中…';
        btn.disabled = true;
        var here = window.location.origin + window.location.pathname;
        var sources = [
            { app: 'あわみ',   api: '/awami/api/find_by_url' },
            { app: 'あわなら', api: '/awanara/api/find_by_url' }
        ];
        Promise.all(sources.map(function(s) {
            return fetch(s.api + '?url=' + encodeURIComponent(here),
                         { credentials: 'same-origin' })
                .then(function(r) { return r.ok ? r.json() : null; })
                .then(function(d) {
                    if (!d || !d.success) return [];
                    return (d.candidates || []).map(function(c) {
                        c.app = s.app; return c;
                    });
                })
                .catch(function() { return []; });
        })).then(function(results) {
            btn.disabled = false;
            var c = [].concat.apply([], results);
            if (c.length === 0) {
                msg.textContent = 'このプレビューを参照しているキャンバスは見つかりませんでした。';
            } else if (c.length === 1) {
                msg.textContent = '「' + c[0].canvas_name + '」（' + c[0].app + '）で開きます…';
                window.open(c[0].url, '_blank');
            } else {
                msg.textContent = '参照するキャンバスを選んでください（' + c.length + '件）:';
                list.style.display = 'flex';
                c.forEach(function(item) {
                    var a = document.createElement('a');
                    a.href = item.url; a.target = '_blank';
                    a.textContent = '[' + item.app + '] ' + item.canvas_name +
                        (item.node_label ? '（' + item.node_label + '）' : '');
                    a.style.cssText = 'background:#fff;border:1px solid #ccd4dc;' +
                        'border-radius:6px;padding:5px 10px;font-size:13px;' +
                        'color:#35618f;text-decoration:none;';
                    list.appendChild(a);
                });
            }
        });
    }
    </script>"""


@colrep_bp.route('/shared_preview/<int:project_id>')
@login_required
def shared_preview(project_id):
    """共有プレビュー：共有設定（access_policy）で許可された人が統合プレビューを閲覧できる"""
    user_info = get_user_info()
    user_id = user_info.get('user_id')

    if not check_shared_preview_access(project_id, user_id):
        return Response(
            "<p>このプロジェクトのプレビューを閲覧する権限がありません。"
            "プロジェクト責任者に共有設定を確認してください。</p>",
            mimetype='text/html'), 403

    try:
        conn = mysql.connector.connect(**DatabaseConfig.fujinp())
        cursor = conn.cursor(dictionary=True)

        cursor.execute(f"""
            SELECT p.*, u.full_name as 責任者名
            FROM {COLREP_PROJECTS_TABLE} p
            LEFT JOIN {Tables.USERS} u ON p.責任者 = u.id
            WHERE p.id = %s
        """, (project_id,))
        project = cursor.fetchone()
        if not project:
            return Response("<p>プロジェクトが見つかりません。</p>", mimetype='text/html'), 404

        table_name = project['テーブル名']
        cursor.execute(f"SHOW TABLES LIKE '{table_name}'")
        if not cursor.fetchone():
            return Response("<p>プロジェクトテーブルがまだ作成されていません。</p>",
                            mimetype='text/html'), 400
        if not project.get('Composer'):
            return Response("<p>統合指示書（Composer）がまだ作成されていません。</p>",
                            mimetype='text/html'), 400

# ★追加: 責任者・adminには編集ボタン付きのバーを表示
        is_manager = (session.get('user_category') == 'admin'
                      or project.get('責任者') == user_id)
        manager_bar = ''
        if is_manager:
            manager_bar = f"""
    <div style="position: sticky; top: 0; z-index: 10; background: #f8f9fa;
                border-bottom: 1px solid #dee2e6; padding: 8px 16px;
                display: flex; justify-content: space-between; align-items: center;
                font-family: sans-serif;">
        <span style="color: #6c757d; font-size: 14px;">プレビュー表示中（編集権限があります）</span>
        <a href="/colrep/edit_project/{project_id}"
           style="background: #0d6efd; color: #fff; padding: 6px 14px;
                  border-radius: 6px; text-decoration: none; font-size: 14px;">
            ✏ プロジェクト編集
        </a>
    </div>"""
        cursor.execute(f"""
            SELECT t.*, u.full_name
            FROM `{table_name}` t
            LEFT JOIN {Tables.USERS} u ON t.担当者アカウント = u.id
            ORDER BY t.id ASC
        """)
        raw_data = cursor.fetchall()

        integrated_markdown = _process_composer_integration(
            project['Composer'], raw_data, cursor, table_name)

        try:
            integrated_html = process_markdown(integrated_markdown)
        except Exception as e:
            logging.error(f"Markdown変換エラー: {str(e)}")
            import html as html_module
            escaped_content = html_module.escape(integrated_markdown)
            integrated_html = f"<pre style='white-space: pre-wrap;'>{escaped_content}</pre>"

        canvas_bar = _canvas_ref_bar()

        complete_html = f"""<!DOCTYPE html>
<html lang="ja">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{project['プロジェクト名']} - プレビュー</title>
    {_math_diagram_assets()}
</head>
<body>
    {manager_bar}
    {canvas_bar}
    <div id="integrated-content">
        {integrated_html}
    </div>
</body>
</html>"""
        return Response(complete_html, mimetype='text/html')

    except Exception as e:
        logging.error(f"Error in shared_preview: {str(e)}")
        return Response(f"<p style='color: red;'>エラーが発生しました: {str(e)}</p>",
                        mimetype='text/html'), 500
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close()
            conn.close()

@colrep_bp.route('/get_share_settings/<int:project_id>')
@login_required
def get_share_settings(project_id):
    """共有設定を取得（プロジェクト責任者・システム管理者のみ）"""
    user_id = session.get('user_id')
    try:
        conn = mysql.connector.connect(**DatabaseConfig.fujinp())
        cursor = conn.cursor(dictionary=True)
        cursor.execute(f"SELECT 責任者, access_policy FROM {COLREP_PROJECTS_TABLE} WHERE id = %s",
                       (project_id,))
        project = cursor.fetchone()
        if not project:
            return jsonify({'success': False, 'error': 'プロジェクトが見つかりません。'}), 404

        if not (session.get('user_category') == 'admin' or project['責任者'] == user_id):
            return jsonify({'success': False,
                            'error': '共有設定はプロジェクト責任者のみ操作できます。'}), 403

        cursor.execute("SELECT group_id FROM colrep_access_groups WHERE project_id = %s",
                       (project_id,))
        group_ids = [r['group_id'] for r in cursor.fetchall()]

        return jsonify({
            'success': True,
            'access_policy': project.get('access_policy') or 'private',
            'group_ids': group_ids,
            'groups': get_all_user_groups()
        })
    except Exception as e:
        logging.error(f"Error in get_share_settings: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close()
            conn.close()


@colrep_bp.route('/save_share_settings/<int:project_id>', methods=['POST'])
@login_required
def save_share_settings(project_id):
    """共有設定を保存（プロジェクト責任者・システム管理者のみ）"""
    user_id = session.get('user_id')
    data = request.get_json() or {}
    policy = data.get('access_policy', 'private')
    group_ids = data.get('access_group_ids') or []

    if policy not in VALID_ACCESS_POLICIES:
        return jsonify({'success': False, 'error': '不正な公開範囲です。'}), 400
    if policy in ('group', 'domestic_group') and not group_ids:
        return jsonify({'success': False,
                        'error': 'グループ公開では、許可するグループを1つ以上選んでください。'}), 400

    try:
        conn = mysql.connector.connect(**DatabaseConfig.fujinp())
        cursor = conn.cursor(dictionary=True)
        cursor.execute(f"SELECT 責任者 FROM {COLREP_PROJECTS_TABLE} WHERE id = %s", (project_id,))
        project = cursor.fetchone()
        if not project:
            return jsonify({'success': False, 'error': 'プロジェクトが見つかりません。'}), 404

        if not (session.get('user_category') == 'admin' or project['責任者'] == user_id):
            return jsonify({'success': False,
                            'error': '共有設定はプロジェクト責任者のみ操作できます。'}), 403

        cursor.execute(f"""
            UPDATE {COLREP_PROJECTS_TABLE}
            SET access_policy = %s, 更新日時 = %s
            WHERE id = %s
        """, (policy, get_jst_now(), project_id))

        cursor.execute("DELETE FROM colrep_access_groups WHERE project_id = %s", (project_id,))
        if policy in ('group', 'domestic_group'):
            for gid in group_ids:
                cursor.execute(
                    "INSERT INTO colrep_access_groups (project_id, group_id) VALUES (%s, %s)",
                    (project_id, int(gid)))
        conn.commit()

        return jsonify({'success': True, 'message': '共有設定を保存しました。'})
    except Exception as e:
        logging.error(f"Error in save_share_settings: {str(e)}")
        if 'conn' in locals() and conn.is_connected():
            conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close()
            conn.close()

@colrep_bp.route('/archive_project/<int:project_id>', methods=['POST'])
@login_required
def archive_project(project_id):
    """プロジェクトをアーカイブシステムに保存（統合コンテンツ版）"""
    user_info = get_user_info()
    user_id = user_info.get('user_id')

    try:
        # conn = mysql.connector.connect(**base_db_config, database=TARGET_DATABASE)
        conn = mysql.connector.connect(**DatabaseConfig.fujinp())
        cursor = conn.cursor(dictionary=True)

        # プロジェクト情報を取得
        cursor.execute(f"""
            SELECT cp.*, u.full_name as 責任者名, u.full_name as 責任者氏名
            FROM {COLREP_PROJECTS_TABLE} cp
            LEFT JOIN {Tables.USERS} u ON cp.責任者 = u.id
            WHERE cp.id = %s
        """, (project_id,))
        project = cursor.fetchone()

        if not project:
            flash('プロジェクトが見つかりません。', 'error')
            return redirect(url_for('colrep.index'))

        # 権限チェック：プロジェクト責任者または総管理者
        permissions = get_user_permissions()
        if not (project['責任者'] == user_id or permissions['is_colrep_admin']):
            flash('このプロジェクトをアーカイブする権限がありません。', 'error')
            return redirect(url_for('colrep.view_project', project_id=project_id))

        table_name = project['テーブル名']

        # テーブルが存在するかチェック
        cursor.execute(f"SHOW TABLES LIKE '{table_name}'")
        table_exists = cursor.fetchone() is not None

        if not table_exists:
            flash('プロジェクトテーブルが存在しません。', 'error')
            return redirect(url_for('colrep.view_project', project_id=project_id))

        # Composerが存在するかチェック
        if not project.get('Composer'):
            flash('統合指示書（Composer）が作成されていません。', 'error')
            return redirect(url_for('colrep.view_project', project_id=project_id))

        # すべてのタスクデータを取得
        cursor.execute(f"""
            SELECT t.*, u.full_name, u.full_name
            FROM `{table_name}` t
            LEFT JOIN {Tables.USERS} u ON t.担当者アカウント = u.id
            ORDER BY t.id ASC
        """)
        raw_data = cursor.fetchall()

        # Composerの指示を解析して統合コンテンツを生成（既存関数を流用）
        integrated_markdown = _process_composer_integration(
            project['Composer'],
            raw_data,
            cursor,
            table_name
        )

        # MarkdownをHTMLに変換
        try:
            integrated_html = process_markdown_for_preview(integrated_markdown)
            logging.info("medit.markdown_converter import successful")
        except ImportError:
            # フォールバック処理
            import html
            escaped_content = html.escape(integrated_markdown)
            integrated_html = f"<pre style='white-space: pre-wrap; font-family: inherit;'>{escaped_content}</pre>"
            logging.info("Using python-markdown as fallback")
        # フォームデータを取得
        title = request.form.get('archive_title', '')
        public_description = request.form.get('public_description', '')
        owner_memo = request.form.get('owner_memo', '')

        if not title.strip():
            flash('アーカイブタイトルは必須です。', 'error')
            return redirect(url_for('colrep.view_project', project_id=project_id))

        # 統合HTMLコンテンツを完全なHTMLドキュメントとして整形
        now = get_jst_now()
        now_datetime_str = now.strftime('%Y年%m月%d日 %H:%M')

        # ★ プロジェクトのソース（JSON）を生成してアーカイブに同梱する
        #   エクスポート機能と同一構造。復元・再インポート用。
        cursor.execute(f"""
            SELECT
                t.id, t.更新日時, t.カラム名,
                u.full_name as ユーザ名, u.full_name as 氏名,
                t.説明, t.content, t.備考, t.status
            FROM `{table_name}` t
            LEFT JOIN {Tables.USERS} u ON t.担当者アカウント = u.id
            ORDER BY t.id ASC
        """)
        json_rows = cursor.fetchall()
        source_json = json.dumps(
            build_project_export_json(project, json_rows, table_name),
            ensure_ascii=False, indent=2
        )

        final_html = generate_complete_archive_html(
            title=title,
            project_name=project['プロジェクト名'],
            integrated_content=integrated_html,
            archive_datetime=now_datetime_str,
            project_id=project_id
        )

        # archive_project 関数内で以下のログを追加

        # 1. 統合Markdown生成後
        logging.info(f"Integrated Markdown (first 200 chars): {integrated_markdown[:200]}...")

        # 2. HTML変換後
        logging.info(f"Integrated HTML (first 200 chars): {integrated_html[:200]}...")

        # 3. 最終HTML生成後
        logging.info(f"Final HTML content length: {len(final_html)}")
        logging.info(f"Final HTML preview (chars 1000-1200): {final_html[1000:1200]}...")

        # アーカイブシステムに保存
        success, message = save_to_colrep_archive(
            title=title,
            public_description=public_description,
            owner_memo=owner_memo,
            html_content=final_html,
            source_info=f"CoRePoプロジェクト統合レポート (プロジェクトID: {project_id})",
            source_json=source_json
        )

        if success:
            flash(f'プロジェクト「{project["プロジェクト名"]}」の統合レポートをアーカイブに保存しました', 'success')
        else:
            flash(f'アーカイブ保存中にエラーが発生しました: {message}', 'error')

    except Exception as e:
        logging.error(f"Error in archive_project: {str(e)}")
        flash(f'エラーが発生しました: {str(e)}', 'error')
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close()
            conn.close()

    return redirect(url_for('colrep.view_project', project_id=project_id))

def generate_complete_archive_html(title, project_name, integrated_content, archive_datetime, project_id):
    """統合レポート用のプレーンなHTMLドキュメントを生成"""

    # 最小限のHTMLドキュメント構造のみ
    html_content = """<!DOCTYPE html>
<html lang="ja">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>""" + title + """</title>
</head>
<body>
    """ + integrated_content + """
</body>
</html>"""

    return html_content

def save_to_colrep_archive(title, public_description, owner_memo, html_content, source_info="", source_json=None):
    """CoRePoアーカイブシステムに保存。
    source_json を渡すと、プロジェクトのソース(JSON)を corepo_source_json 列に格納する。"""
    try:
        # component_editor2と同じアーカイブテーブルを使用
        # connection = mysql.connector.connect(**base_db_config, database="nishida$fujinp")
        connection = mysql.connector.connect(**DatabaseConfig.fujinp())
        cursor = connection.cursor()

        now = get_jst_now()
        formatted_now = now.strftime('%Y-%m-%d %H:%M:%S')

        current_user_id = session.get('user_id')

        cursor.execute("""
            INSERT INTO public_documents
            (title, public_description, owner_memo, content, corepo_source_json,
             created_by, created_at, updated_at, access_policy)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (title, public_description, owner_memo, html_content, source_json,
              current_user_id, formatted_now, formatted_now, 'private'))
        connection.commit()
        doc_id = cursor.lastrowid

        cursor.close()
        connection.close()

        logging.info(f"CoRePoプロジェクトをアーカイブに保存しました (ID: {doc_id}): {title}")
        return True, f"ID:{doc_id} として保存されました"

    except Exception as e:
        logging.error(f"CoRePoアーカイブ保存エラー: {str(e)}")
        return False, str(e)

@colrep_bp.route('/update_task/<int:project_id>/<int:task_id>', methods=['POST'])
@login_required
def update_task(project_id, task_id):
    """タスクの担当者、説明、備考を更新"""
    user_info = get_user_info()
    user_id = user_info.get('user_id')
    permissions = get_user_permissions()

    try:
        print(f"タスク更新開始: project_id={project_id}, task_id={task_id}")

        # conn = mysql.connector.connect(**base_db_config, database=TARGET_DATABASE)
        conn = mysql.connector.connect(**DatabaseConfig.fujinp())
        cursor = conn.cursor(dictionary=True)

        # プロジェクト情報を取得
        cursor.execute(f"SELECT テーブル名, 責任者 FROM {COLREP_PROJECTS_TABLE} WHERE id = %s", (project_id,))
        project = cursor.fetchone()
        print(f"プロジェクト確認OK: {project}")

        if not project:
            return jsonify({'success': False, 'error': 'プロジェクトが見つかりません'}), 404

        # 権限チェック：システム管理者、CoRePo総管理者、または責任者
        is_system_admin = (session.get('user_category') == 'admin')
        is_colrep_admin = permissions['is_colrep_admin']
        is_owner = (project['責任者'] == user_id)

        if not (is_system_admin or is_colrep_admin or is_owner):
            return jsonify({'success': False, 'error': 'タスクの編集はプロジェクト責任者または管理者のみ実行可能です'}), 403

        # リクエストデータの取得
        data = request.get_json()
        print(f"受信データ: {data}")
        if not data:
            return jsonify({'success': False, 'error': 'データが送信されていません'}), 400

        # 必須フィールドの確認
        担当者アカウント = data.get('担当者アカウント')
        if not 担当者アカウント:
            return jsonify({'success': False, 'error': '担当者は必須項目です'}), 400

        # 担当者の存在確認
        conn_users = mysql.connector.connect(**DatabaseConfig.default())
        cursor_users = conn_users.cursor(dictionary=True)
        cursor_users.execute(f"SELECT id FROM {Tables.USERS} WHERE id = %s", (担当者アカウント,))
        user_exists = cursor_users.fetchone()
        cursor_users.close()
        conn_users.close()

        if not user_exists:
            return jsonify({'success': False, 'error': '指定された担当者が見つかりません'}), 400

        table_name = project['テーブル名']

        # タスクの存在確認
        cursor.execute(f"SELECT id FROM `{table_name}` WHERE id = %s", (task_id,))
        task_exists = cursor.fetchone()

        if not task_exists:
            return jsonify({'success': False, 'error': 'タスクが見つかりません'}), 404

        # タスクの更新
        now = get_jst_now()
        説明 = data.get('説明', '').strip() if data.get('説明') else None
        備考 = data.get('備考', '').strip() if data.get('備考') else None

        update_query = f"""
            UPDATE `{table_name}`
            SET 担当者アカウント = %s, 説明 = %s, 備考 = %s, 更新日時 = %s
            WHERE id = %s
        """

        cursor.execute(update_query, (担当者アカウント, 説明, 備考, now, task_id))
        conn.commit()

        return jsonify({
            'success': True,
            'message': 'タスクが正常に更新されました'
        })

    except Exception as e:
        print(f"タスク更新エラー: {str(e)}")
        if 'conn' in locals() and conn.is_connected():
            conn.rollback()
        return jsonify({
            'success': False,
            'error': f'タスクの更新中にエラーが発生しました: {str(e)}'
        }), 500
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close()
            conn.close()

# routes.py に追加/修正する関数

# ============================================================
# 1. 権限チェック関数（新規追加）
# ============================================================

def check_project_access_permission(project_id, user_id):
    """
    プロジェクトへのアクセス権限をチェック

    条件：
    - 管理者: すべてのプロジェクトにアクセス可能
    - public プロジェクト: 全ユーザーにアクセス可能
    - private プロジェクト: 責任者と執筆者のみアクセス可能

    戻り値: (has_access, is_author) の tuple
    """
    # ★システム基盤adminは常にオールマイティ
    if session.get('user_category') == 'admin':
        return True, True

    try:
        # conn = mysql.connector.connect(**base_db_config, database=TARGET_DATABASE)
        conn = mysql.connector.connect(**DatabaseConfig.fujinp())
        cursor = conn.cursor(dictionary=True)

        # プロジェクト情報を取得
        cursor.execute(f"""
            SELECT id, is_public, 責任者, テーブル名
            FROM {COLREP_PROJECTS_TABLE}
            WHERE id = %s
        """, (project_id,))

        project = cursor.fetchone()
        if not project:
            return False, False

        # ★システム基盤adminは常にオールマイティ
        if session.get('user_category') == 'admin':
            return True, True
        # 管理者は常にアクセス可能
        permissions = get_user_permissions()
        if permissions['is_colrep_admin']:
            return True, True

        # publicプロジェクトは全ユーザーアクセス可能
        if project['is_public']:
            return True, False

        # privateプロジェクト：責任者のみアクセス可能
        if project['責任者'] == user_id:
            return True, True

        # ★★★ 追加：プロジェクトテーブルの担当者であればアクセス可能 ★★★
        if project['テーブル名']:
            try:
                cursor.execute(f"""
                    SELECT id FROM `{project['テーブル名']}`
                    WHERE 担当者アカウント = %s
                    LIMIT 1
                """, (user_id,))
                if cursor.fetchone():
                    return True, False  # アクセス可能だが編集権限（is_author）はFalse
            except Exception as e:
                logging.warning(f"担当者チェック時エラー: {str(e)}")
        # ★★★ ここまで ★★★
        # それ以外はアクセス不可
        return False, False

    except Exception as e:
        logging.error(f"Error in check_project_access_permission: {str(e)}")
        return False, False
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close()
            conn.close()


def get_project_contributors(project_id):
    """
    プロジェクトの執筆者（コンテンツが入っているユーザー）を取得
    """
    try:
        # conn = mysql.connector.connect(**base_db_config, database=TARGET_DATABASE)
        conn = mysql.connector.connect(**DatabaseConfig.fujinp())
        cursor = conn.cursor(dictionary=True)

        # プロジェクト情報を取得
        cursor.execute(f"""
            SELECT テーブル名 FROM {COLREP_PROJECTS_TABLE}
            WHERE id = %s
        """, (project_id,))

        project = cursor.fetchone()
        if not project:
            return []

        table_name = project['テーブル名']

        # コンテンツが入っている執筆者を取得
        cursor.execute(f"""
            SELECT DISTINCT 担当者アカウント
            FROM `{table_name}`
            WHERE content IS NOT NULL AND content != ''
        """)

        contributors = cursor.fetchall()
        return [c['担当者アカウント'] for c in contributors]

    except Exception as e:
        logging.error(f"Error in get_project_contributors: {str(e)}")
        return []
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close()
            conn.close()

# ============================================================
# 5. 新規エンドポイント: 公開ドキュメントビューア用API
# ============================================================

@colrep_bp.route('/get_project_info/<int:project_id>')
@login_required
def get_project_info(project_id):
    """プロジェクト情報取得API（アクセス権限チェック付き）"""
    user_info = get_user_info()
    user_id = user_info.get('user_id')

    try:
        # ✅ アクセス権限をチェック
        has_access, is_author = check_project_access_permission(project_id, user_id)
        if not has_access:
            return jsonify({'success': False, 'error': 'アクセス権限がありません。'}), 403

        # conn = mysql.connector.connect(**base_db_config, database=TARGET_DATABASE)
        conn = mysql.connector.connect(**DatabaseConfig.fujinp())
        cursor = conn.cursor(dictionary=True)

        cursor.execute(f"""
            SELECT id, プロジェクト名, Composer, is_public, 責任者
            FROM {COLREP_PROJECTS_TABLE}
            WHERE id = %s
        """, (project_id,))

        project = cursor.fetchone()
        if not project:
            return jsonify({'success': False, 'error': 'プロジェクトが見つかりません。'}), 404

        return jsonify({
            'success': True,
            'project': {
                'id': project['id'],
                'プロジェクト名': project['プロジェクト名'],
                'is_public': project['is_public'],
                'has_composer': bool(project['Composer']),
                'can_edit': is_author  # 編集可能かどうか
            }
        })

    except Exception as e:
        logging.error(f"Error in get_project_info: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close()
            conn.close()

# routes.py に追加する新規エンドポイント

# ============================================================
# 公開ドキュメントビューア用エンドポイント
# ============================================================

@colrep_bp.route('/public_documents')
@login_required
def public_documents():
    """公開ドキュメントビューア（全ユーザーアクセス可能）"""
    return render_template('public_documents_viewer.html')


@colrep_bp.route('/list_public_projects')
@login_required
def list_public_projects():
    """
    アクセス可能なプロジェクト一覧を取得（JSON）

    条件：
    - 管理者: すべてのプロジェクト
    - 一般ユーザー: 公開プロジェクト + 自分が関わっているプロジェクト
    """
    user_info = get_user_info()
    user_id = user_info.get('user_id')
    permissions = get_user_permissions()

    try:
        # conn = mysql.connector.connect(**base_db_config, database=TARGET_DATABASE)
        conn = mysql.connector.connect(**DatabaseConfig.fujinp())
        cursor = conn.cursor(dictionary=True)

        if permissions['is_colrep_admin']:
            # 管理者: すべてのプロジェクト
            cursor.execute(f"""
                SELECT cp.id, cp.プロジェクト名, cp.更新日時, cp.is_public,
                       cp.責任者, u.full_name as 責任者名
                FROM {COLREP_PROJECTS_TABLE} cp
                LEFT JOIN {Tables.USERS} u ON cp.責任者 = u.id
                ORDER BY cp.更新日時 DESC
            """)
        else:
            # 一般ユーザー: 公開プロジェクト + 責任者プロジェクト + 執筆者プロジェクト
            cursor.execute(f"""
                SELECT DISTINCT cp.id, cp.プロジェクト名, cp.更新日時, cp.is_public,
                       cp.責任者, u.full_name as 責任者名, cp.テーブル名
                FROM {COLREP_PROJECTS_TABLE} cp
                LEFT JOIN {Tables.USERS} u ON cp.責任者 = u.id
                WHERE cp.is_public = TRUE
                   OR cp.責任者 = %s
                ORDER BY cp.更新日時 DESC
            """, (user_id,))

        projects = cursor.fetchall()

        # 一般ユーザーの場合、執筆者プロジェクトも追加
        if not permissions['is_colrep_admin']:
            # ユーザーが執筆者であるプロジェクトを取得
            author_projects = set()
            cursor.execute(f"""
                SELECT DISTINCT cp.id
                FROM {COLREP_PROJECTS_TABLE} cp
                WHERE cp.テーブル名 IS NOT NULL AND cp.テーブル名 != ''
            """)

            potential_projects = cursor.fetchall()
            for proj in potential_projects:
                table_name = None
                cursor.execute(f"""
                    SELECT テーブル名 FROM {COLREP_PROJECTS_TABLE}
                    WHERE id = %s
                """, (proj['id'],))
                proj_info = cursor.fetchone()
                if proj_info:
                    table_name = proj_info['テーブル名']

                    # このテーブルでユーザーが执筆者かチェック
                    cursor.execute(f"""
                        SELECT id FROM `{table_name}`
                        WHERE 担当者アカウント = %s AND content IS NOT NULL AND content != ''
                        LIMIT 1
                    """, (user_id,))

                    if cursor.fetchone():
                        author_projects.add(proj['id'])

            # 执筆者プロジェクトを追加
            if author_projects:
                cursor.execute(f"""
                    SELECT cp.id, cp.プロジェクト名, cp.更新日時, cp.is_public,
                           cp.責任者, u.full_name as 責任者名
                    FROM {COLREP_PROJECTS_TABLE} cp
                    LEFT JOIN {Tables.USERS} u ON cp.責任者 = u.id
                    WHERE cp.id IN ({','.join(str(p) for p in author_projects)})
                """)
                author_projects_data = cursor.fetchall()

                # 既存のプロジェクトに追加（重複排除）
                existing_ids = {p['id'] for p in projects}
                for proj in author_projects_data:
                    if proj['id'] not in existing_ids:
                        projects.append(proj)

        return jsonify({
            'success': True,
            'projects': serialize_for_json(projects)
        })

    except Exception as e:
        logging.error(f"Error in list_public_projects: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close()
            conn.close()


@colrep_bp.route('/get_composer_source/<int:project_id>')
@login_required
def get_composer_source(project_id):
    """
    Composerソースコードを取得（アクセス権限チェック付き）
    """
    user_info = get_user_info()
    user_id = user_info.get('user_id')

    try:
        # ✅ アクセス権限をチェック
        has_access, is_author = check_project_access_permission(project_id, user_id)
        if not has_access:
            return jsonify({'success': False, 'error': 'アクセス権限がありません。'}), 403

        # conn = mysql.connector.connect(**base_db_config, database=TARGET_DATABASE)
        conn = mysql.connector.connect(**DatabaseConfig.fujinp())
        cursor = conn.cursor(dictionary=True)

        cursor.execute(f"""
            SELECT プロジェクト名, Composer
            FROM {COLREP_PROJECTS_TABLE}
            WHERE id = %s
        """, (project_id,))

        project = cursor.fetchone()
        if not project:
            return jsonify({'success': False, 'error': 'プロジェクトが見つかりません。'}), 404

        if not project['Composer']:
            return jsonify({'success': False, 'error': 'Composerがまだ作成されていません。'}), 400

        return jsonify({
            'success': True,
            'project_name': project['プロジェクト名'],
            'composer': project['Composer']
        })

    except Exception as e:
        logging.error(f"Error in get_composer_source: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close()
            conn.close()

# このコードを routes.py の worker_dashboard() の直前に挿入してください
# デバッグ用のルートを追加

# [2026-07-25 改修] 削除: /debug_session, /debug_user_tasks/<user_id>, /debug_database
# いずれも @login_required が無く、セッション内容やDBの中身を無認証で返していたため削除。

@colrep_bp.route('/preview', methods=['POST'])
@login_required          # [2026-07-25 改修] 無認証だったためログイン必須に変更
def preview_markdown():
    """作業内容のMarkdownプレビュー"""
    data = request.get_json()
    markdown_text = data.get('markdown', '')

    try:
        # シンプルな拡張のみを使用
        # html = markdown.markdown(markdown_text, extensions=['extra', 'nl2br', 'sane_lists', 'fenced_code'])
        html = process_markdown(markdown_text)
        return jsonify({'html': html})
    except Exception as e:
        import traceback
        logging.error(f"Preview error: {str(e)}\n{traceback.format_exc()}")
        return jsonify({
            'html': f'<p style="color: red;">エラー: {str(e)}</p>',
            'error_details': traceback.format_exc()
        })


@colrep_bp.route('/export_project_data/<int:project_id>')
@login_required          # [2026-07-25 改修] コメントアウトされていた認証を復活
def export_project_data(project_id):
    """プロジェクトのテーブルデータをExcelファイルとしてエクスポート"""
    user_info = get_user_info()
    user_id = user_info.get('user_id')
    user_category = session.get('user_category')  # ★システム管理者判定用
    permissions = get_user_permissions()

    try:
        # ✅ 修正: use_pure=True を追加
        # conn = mysql.connector.connect(**base_db_config, database=TARGET_DATABASE)
        conn = mysql.connector.connect(**DatabaseConfig.fujinp())
        cursor = conn.cursor(dictionary=True)

        # プロジェクト情報の取得
        cursor.execute(f"""
            SELECT cp.*, u.full_name as 責任者名
            FROM {COLREP_PROJECTS_TABLE} cp
            LEFT JOIN {Tables.USERS} u ON cp.責任者 = u.id
            WHERE cp.id = %s
        """, (project_id,))

        project = cursor.fetchone()
        if not project:
            flash('プロジェクトが見つかりません。', 'error')
            return redirect(url_for('colrep.index'))

        # ★★★ 権限チェック修正: システムadmin, CoRePo総管理者, 責任者, 公開設定のいずれかなら許可 ★★★
        user_category = session.get('user_category')
        permissions = get_user_permissions()

        is_system_admin = (user_category == 'admin')
        is_colrep_admin = permissions['is_colrep_admin']
        is_owner = (project['責任者'] == user_id)

        if not (is_system_admin or is_colrep_admin or is_owner or project['is_public']):
            flash('エクスポート権限がありません。', 'error')
            return redirect(url_for('colrep.index'))

        # ★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★

        table_name = project['テーブル名']

        # テーブルが存在するか確認
        cursor.execute(f"SHOW TABLES LIKE '{table_name}'")
        if not cursor.fetchone():
            flash('プロジェクトテーブルが存在しません。', 'error')
            return redirect(url_for('colrep.view_project', project_id=project_id))

        # ✅ 修正: 新プラットフォーム用のスキーマに変更
        cursor.execute(f"""
            SELECT
                t.id,
                t.更新日時,
                t.カラム名,
                u.full_name as ユーザ名,
                u.full_name as 氏名,
                t.説明,
                t.content,
                t.備考,
                t.status
            FROM `{table_name}` t
            LEFT JOIN {Tables.USERS} u ON t.担当者アカウント = u.id
            ORDER BY t.id ASC
        """)

        rows = cursor.fetchall()

        # ✅ 修正: タスクが空でもComposerがあればエクスポート可能に
        # if not rows:
        #     flash('エクスポートするデータがありません。', 'warning')
        #     return redirect(url_for('colrep.view_project', project_id=project_id))

        # ✅ Excel容量制限の定数
        EXCEL_CELL_MAX_CHARS = 32767

        # ✅ 容量超過チェック用リスト
        truncation_warnings = []

        # Excelワークブックを作成（UTF-8対応）
        wb = openpyxl.Workbook()

        # スタイル定義
        title_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        title_font = Font(color="FFFFFF", bold=True, size=14)
        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        header_font = Font(bold=True, size=11)
        data_header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        data_header_font = Font(color="FFFFFF", bold=True, size=11)
        warning_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        warning_font = Font(color="9C0006", bold=True)
        info_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
        info_font = Font(color="1565C0", size=10)
        border_thin = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # ========== シート1: Composer ==========
        ws_composer = wb.active
        ws_composer.title = "Composer"

        # Composerタイトル
        ws_composer.merge_cells('A1:B1')
        composer_title = ws_composer['A1']
        composer_title.value = "統合指示書 (Composer)"
        composer_title.fill = title_fill
        composer_title.font = title_font
        composer_title.alignment = Alignment(horizontal="center", vertical="center")
        composer_title.border = border_thin
        ws_composer.row_dimensions[1].height = 25

        # プロジェクト情報
        ws_composer['A2'] = "プロジェクト名"
        ws_composer['B2'] = project['プロジェクト名']
        ws_composer['A3'] = "責任者"
        ws_composer['B3'] = project['責任者名']
        ws_composer['A4'] = "エクスポート日時"
        ws_composer['B4'] = get_jst_now().strftime('%Y年%m月%d日 %H:%M')

        # 情報セクションのスタイル
        for row in range(2, 5):
            ws_composer[f'A{row}'].fill = header_fill
            ws_composer[f'A{row}'].font = header_font
            ws_composer[f'A{row}'].border = border_thin
            ws_composer[f'B{row}'].border = border_thin

        # 空白行
        ws_composer.append([])

        # Composer内容ヘッダー
        ws_composer['A6'] = "Composer内容"
        ws_composer['A6'].fill = header_fill
        ws_composer['A6'].font = header_font
        ws_composer['A6'].border = border_thin
        ws_composer.merge_cells('A6:B6')

        # ✅ Composerを1セル（A7）にテキストとして格納（容量チェック付き）
        composer_content = project.get('Composer', '')

        # ✅ デバッグログ追加
        logging.info(f"Composer original length: {len(composer_content) if composer_content else 0} chars")

        if composer_content:
            composer_length = len(composer_content)

            if composer_length > EXCEL_CELL_MAX_CHARS:
                # 容量超過の場合
                truncated_composer = composer_content[:EXCEL_CELL_MAX_CHARS]
                ws_composer['A7'] = truncated_composer

                # ✅ デバッグログ
                logging.warning(f"Composer truncated: {composer_length} → {len(truncated_composer)} chars")

                truncation_warnings.append(
                    f"⚠️ Composer: {composer_length:,}文字 → {EXCEL_CELL_MAX_CHARS:,}文字に切り詰められました"
                )
                # 警告メッセージをB7に表示
                ws_composer['B7'] = f"⚠️ 警告: Composerが長すぎます（{composer_length:,}文字）\nExcelの制限により{EXCEL_CELL_MAX_CHARS:,}文字まで表示"
                ws_composer['B7'].fill = warning_fill
                ws_composer['B7'].font = warning_font
                ws_composer['B7'].alignment = Alignment(vertical="top", wrap_text=True)
                ws_composer['B7'].border = border_thin
            else:
                # ✅ 通常の場合（全文格納）
                ws_composer['A7'] = composer_content

                # ✅ 長文の場合は注意書きをB7に追加
                if composer_length > 5000:  # 5,000文字以上の場合
                    ws_composer['B7'] = f"ℹ️ 注意: Composerは{composer_length:,}文字あります\n\n📋 全文をコピーする方法:\n1. A7セルをダブルクリック\n2. Ctrl+A（全選択）\n3. Ctrl+C（コピー）\n\n※Excelの行高制限により画面上は一部のみ表示されますが、データは完全に格納されています"
                    ws_composer['B7'].fill = info_fill
                    ws_composer['B7'].font = info_font
                    ws_composer['B7'].alignment = Alignment(vertical="top", wrap_text=True)
                    ws_composer['B7'].border = border_thin

                # ✅ デバッグログ
                logging.info(f"Composer stored completely: {composer_length} chars")
        else:
            ws_composer['A7'] = "(Composerが設定されていません)"

        # A7セルのスタイル設定
        ws_composer['A7'].alignment = Alignment(vertical="top", wrap_text=True)
        ws_composer['A7'].border = border_thin

        # ✅ セルに実際に格納されたデータの長さを確認
        actual_stored_length = len(ws_composer['A7'].value) if ws_composer['A7'].value else 0
        logging.info(f"Composer actually stored in cell A7: {actual_stored_length} chars")

        # 列幅設定（A列を非常に広く）
        ws_composer.column_dimensions['A'].width = 120
        ws_composer.column_dimensions['B'].width = 40

        # ========== シート2: タスクデータ ==========
        ws_data = wb.create_sheet(title="タスクデータ")

        # プロジェクト情報（上部）
        ws_data.merge_cells('A1:I1')
        info_cell = ws_data['A1']
        info_cell.value = f"プロジェクト: {project['プロジェクト名']}"
        info_cell.font = Font(bold=True, size=14)
        info_cell.alignment = Alignment(horizontal="left", vertical="center")

        ws_data.merge_cells('A2:I2')
        info_cell2 = ws_data['A2']
        info_cell2.value = f"責任者: {project['責任者名']}  |  エクスポート日時: {get_jst_now().strftime('%Y年%m月%d日 %H:%M')}"
        info_cell2.font = Font(size=10)
        info_cell2.alignment = Alignment(horizontal="left", vertical="center")

        # 空白行
        ws_data.append([])

        # ヘッダー行
        headers = ['ID', '更新日時', 'カラム名', '担当者', '担当者氏名', '説明', '作業内容', '備考', 'ステータス']
        ws_data.append(headers)

        # ヘッダー行のスタイル設定
        header_row = ws_data[4]
        for cell in header_row:
            cell.fill = data_header_fill
            cell.font = data_header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border_thin

        # ✅ データ行を追加（容量チェック付き）
        for row_data in rows:
            # datetimeオブジェクトを文字列に変換
            updated_at = row_data['更新日時'].strftime('%Y/%m/%d %H:%M') if row_data['更新日時'] else ''

            # LONGTEXTフィールドの容量チェック
            def check_and_truncate(text, field_name, task_id):
                if not text:
                    return ''
                text_length = len(text)
                if text_length > EXCEL_CELL_MAX_CHARS:
                    truncation_warnings.append(
                        f"⚠️ タスクID {task_id} の「{field_name}」: {text_length:,}文字 → {EXCEL_CELL_MAX_CHARS:,}文字に切り詰められました"
                    )
                    return text[:EXCEL_CELL_MAX_CHARS] + f"\n\n⚠️【切り詰め警告】元の文字数: {text_length:,}文字"
                return text

            description = check_and_truncate(row_data['説明'], '説明', row_data['id'])
            content = check_and_truncate(row_data['content'], '作業内容', row_data['id'])
            remarks = check_and_truncate(row_data['備考'], '備考', row_data['id'])

            ws_data.append([
                row_data['id'],
                updated_at,
                row_data['カラム名'],
                row_data['ユーザ名'],
                row_data['氏名'],
                description,
                content,
                remarks,
                row_data['status']
            ])

        # 列幅の調整
        column_widths = {
            'A': 8,   # ID
            'B': 18,  # 更新日時
            'C': 25,  # カラム名
            'D': 15,  # 担当者
            'E': 15,  # 担当者氏名
            'F': 30,  # 説明
            'G': 60,  # 作業内容（LONGTEXT）
            'H': 30,  # 備考
            'I': 12   # ステータス
        }

        for col, width in column_widths.items():
            ws_data.column_dimensions[col].width = width

        # データ行のスタイル設定
        for row in ws_data.iter_rows(min_row=5, max_row=ws_data.max_row):
            for cell in row:
                cell.border = border_thin
                # LONGTEXT列（F, G, H列）は折り返しを有効に
                if cell.column_letter in ['F', 'G', 'H']:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                    # 切り詰め警告が含まれている場合は背景色を変更
                    if cell.value and '⚠️【切り詰め警告】' in str(cell.value):
                        cell.fill = warning_fill
                else:
                    cell.alignment = Alignment(vertical="top")

        # 行の高さを自動調整（LONGTEXTがある行）
        for row in ws_data.iter_rows(min_row=5, max_row=ws_data.max_row):
            # content列のセルを確認
            content_cell = row[6]  # G列（インデックス6）
            if content_cell.value and len(str(content_cell.value)) > 100:
                # 長いコンテンツがある場合は行の高さを調整
                ws_data.row_dimensions[content_cell.row].height = 80

        # メモリ上でExcelファイルを生成
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # ファイル名を生成
        now = get_jst_now()
        filename = f"{project['プロジェクト名']}_{now.strftime('%Y%m%d_%H%M')}.xlsx"
        filename = re.sub(r'[\\/:*?"<>|]', '_', filename)

        # ✅ 切り詰め警告がある場合はフラッシュメッセージで通知
        if truncation_warnings:
            warning_message = "エクスポート完了（一部データが切り詰められました）:\n" + "\n".join(truncation_warnings)
            flash(warning_message, 'warning')
            logging.warning(f"Export truncation warnings for project {project_id}: {truncation_warnings}")
        else:
            flash(f'プロジェクト「{project["プロジェクト名"]}」をエクスポートしました。', 'success')

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        logging.error(f"Error in export_project_data: {str(e)}")
        import traceback
        logging.error(traceback.format_exc())
        flash(f'エクスポート中にエラーが発生しました: {str(e)}', 'error')
        return redirect(url_for('colrep.view_project', project_id=project_id))
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close()
            conn.close()


@colrep_bp.route('/export_project_data_json/<int:project_id>')
@login_required          # [2026-07-25 改修] コメントアウトされていた認証を復活
def export_project_data_json(project_id):
    """プロジェクトのテーブルデータをJSONファイルとしてエクスポート"""
    user_info = get_user_info()
    user_id = user_info.get('user_id')
    user_category = session.get('user_category')  # ★システム管理者判定用
    permissions = get_user_permissions()

    try:
        conn = mysql.connector.connect(**DatabaseConfig.fujinp())
        cursor = conn.cursor(dictionary=True)

        # プロジェクト情報の取得
        cursor.execute(f"""
            SELECT cp.*, u.full_name as 責任者名
            FROM {COLREP_PROJECTS_TABLE} cp
            LEFT JOIN {Tables.USERS} u ON cp.責任者 = u.id
            WHERE cp.id = %s
        """, (project_id,))

        project = cursor.fetchone()
        if not project:
            flash('プロジェクトが見つかりません。', 'error')
            return redirect(url_for('colrep.index'))

        # ★★★ 権限チェック: システムadmin, CoRePo総管理者, 責任者, 公開設定のいずれかなら許可 ★★★
        is_system_admin = (user_category == 'admin')
        is_colrep_admin = permissions['is_colrep_admin']
        is_owner = (project['責任者'] == user_id)

        if not (is_system_admin or is_colrep_admin or is_owner or project['is_public']):
            flash('エクスポート権限がありません。', 'error')
            return redirect(url_for('colrep.index'))

        table_name = project['テーブル名']

        # テーブルが存在するか確認
        cursor.execute(f"SHOW TABLES LIKE '{table_name}'")
        if not cursor.fetchone():
            flash('プロジェクトテーブルが存在しません。', 'error')
            return redirect(url_for('colrep.view_project', project_id=project_id))

        # タスクデータの取得（Excelエクスポートと同一スキーマ）
        cursor.execute(f"""
            SELECT
                t.id,
                t.更新日時,
                t.カラム名,
                u.full_name as ユーザ名,
                u.full_name as 氏名,
                t.説明,
                t.content,
                t.備考,
                t.status
            FROM `{table_name}` t
            LEFT JOIN {Tables.USERS} u ON t.担当者アカウント = u.id
            ORDER BY t.id ASC
        """)

        rows = cursor.fetchall()

        # JSON構造を組み立て（エクスポート／アーカイブ共通ヘルパー）
        serialized = build_project_export_json(project, rows, table_name)
        json_bytes = json.dumps(serialized, ensure_ascii=False, indent=2).encode('utf-8')

        output = io.BytesIO(json_bytes)
        output.seek(0)

        # ファイル名を生成
        now = get_jst_now()
        filename = f"{project['プロジェクト名']}_{now.strftime('%Y%m%d_%H%M')}.json"
        filename = re.sub(r'[\\/:*?"<>|]', '_', filename)

        flash(f'プロジェクト「{project["プロジェクト名"]}」をJSONエクスポートしました。', 'success')

        return send_file(
            output,
            mimetype='application/json; charset=utf-8',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        logging.error(f"Error in export_project_data_json: {str(e)}")
        import traceback
        logging.error(traceback.format_exc())
        flash(f'JSONエクスポート中にエラーが発生しました: {str(e)}', 'error')
        return redirect(url_for('colrep.view_project', project_id=project_id))
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close()
            conn.close()

# routes.py に追加するインポート機能のコード（FUJINP用 - nishida$fujinp）

@colrep_bp.route('/import_project', methods=['GET', 'POST'])
@login_required
def import_project():
    """Excelファイルからプロジェクトをインポート（新プラットフォーム用 - インポート者が責任者・担当者）"""
    # ★修正: 権限チェックを変更 (総管理者 OR CoRePo管理者 ならOK)
    permissions = get_user_permissions()
    if not (permissions['is_colrep_admin'] or permissions['is_colrep_manager']):
        flash('プロジェクト作成権限（CoRePo管理者以上）が必要です。', 'error')
        return redirect(url_for('colrep.index'))

    if request.method == 'POST':
        try:
            # ✅ インポートを実行したユーザーのIDを取得
            import_user_id = session.get('user_id')
            if not import_user_id:
                flash('ログインが必要です。', 'error')
                return redirect(url_for('colrep.index'))

            # アップロードされたファイルを取得
            if 'excel_file' not in request.files:
                flash('ファイルが選択されていません。', 'error')
                return redirect(url_for('colrep.index'))

            file = request.files['excel_file']
            if file.filename == '':
                flash('ファイルが選択されていません。', 'error')
                return redirect(url_for('colrep.index'))

            if not file.filename.endswith('.xlsx'):
                flash('xlsxファイルのみアップロード可能です。', 'error')
                return redirect(url_for('colrep.index'))

            # Excelファイルを読み込み
            wb = openpyxl.load_workbook(file)

            # ✅ シート2「タスクデータ」を読み込む
            if 'タスクデータ' in wb.sheetnames:
                ws = wb['タスクデータ']
            else:
                # フォールバック: 最初のシート
                ws = wb.active

            # プロジェクト情報を抽出（行1-2）
            project_info = ws['A1'].value  # "プロジェクト: XXX"
            if not project_info or not project_info.startswith('プロジェクト:'):
                flash('Excelファイルの形式が正しくありません（プロジェクト名が見つかりません）。', 'error')
                return redirect(url_for('colrep.index'))

            original_project_name = project_info.replace('プロジェクト:', '').strip()

            # データベース接続
            # conn = mysql.connector.connect(**base_db_config, database=TARGET_DATABASE)
            conn = mysql.connector.connect(**DatabaseConfig.fujinp())
            cursor = conn.cursor(dictionary=True)

            # プロジェクト名の重複チェックと番号付加
            project_name = original_project_name
            counter = 1
            while True:
                cursor.execute(f"SELECT id FROM {COLREP_PROJECTS_TABLE} WHERE プロジェクト名 = %s", (project_name,))
                if not cursor.fetchone():
                    break  # 重複なし
                project_name = f"{original_project_name} ({counter})"
                counter += 1

            # ✅ 新プラットフォーム用: インポート者を責任者に設定
            responsible_user_id = import_user_id

            # 新しいテーブルを作成
            table_name = f"colrep_{uuid.uuid4().hex[:12]}"

            create_table_sql = f"""
                CREATE TABLE `{table_name}` (
                    `id` INT AUTO_INCREMENT PRIMARY KEY,
                    `更新日時` DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                    `カラム名` VARCHAR(255) NOT NULL,
                    `担当者アカウント` INT NOT NULL,
                    `説明` TEXT COMMENT '管理者から入力者への説明',
                    `content` LONGTEXT COMMENT '入力内容',
                    `備考` TEXT COMMENT '入力者から管理者への説明',
                    `status` VARCHAR(20) DEFAULT '作業中' COMMENT '進捗状況：作業中/改訂中/完了',
                    INDEX `idx_担当者` (`担当者アカウント`),
                    INDEX `idx_カラム名` (`カラム名`),
                    INDEX `idx_status` (`status`),
                    FOREIGN KEY (`担当者アカウント`) REFERENCES {Tables.USERS}(`id`) ON DELETE RESTRICT
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
                COMMENT='CoRePoプロジェクト用データテーブル（インポート）'
            """

            cursor.execute(create_table_sql)
            conn.commit()

            # Composerを読み込む（シート1から）
            composer_content = ''
            if 'Composer' in wb.sheetnames:
                ws_composer = wb['Composer']
                # A7セルにComposerが格納されている
                if ws_composer['A7'].value:
                    composer_content = str(ws_composer['A7'].value)

            # プロジェクトを作成
            now = get_jst_now()
            insert_query = f"""
                INSERT INTO {COLREP_PROJECTS_TABLE}
                (プロジェクト名, 更新日時, 責任者, テーブル名, Composer, is_public)
                VALUES (%s, %s, %s, %s, %s, %s)
            """
            cursor.execute(insert_query, (project_name, now, responsible_user_id, table_name, composer_content, False))
            project_id = cursor.lastrowid
            conn.commit()

            # データ行をインポート（行5以降）
            imported_count = 0
            skipped_count = 0

            for row_idx, row in enumerate(ws.iter_rows(min_row=5, values_only=True), start=5):
                # 空行はスキップ
                if not any(row):
                    continue

                # カラムを解析
                # ID, 更新日時, カラム名, 担当者, 担当者氏名, 説明, 作業内容, 備考, ステータス
                row_id, updated_at, column_name, user_name, user_fullname, description, content, remarks, status = row[:9]

                if not column_name:  # カラム名は必須
                    skipped_count += 1
                    continue

                # ✅ 新プラットフォーム用: すべてのタスクの担当者をインポート者に設定
                assigned_user_id = import_user_id

                # データを挿入
                insert_data_query = f"""
                    INSERT INTO `{table_name}`
                    (更新日時, カラム名, 担当者アカウント, 説明, content, 備考, status)
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                """

                cursor.execute(insert_data_query, (
                    now,
                    column_name,
                    assigned_user_id,
                    description or '',
                    content or '',
                    remarks or '',
                    status or '作業中'
                ))
                imported_count += 1

            conn.commit()

            # ✅ メッセージを改善
            if original_project_name != project_name:
                flash(f'プロジェクト「{project_name}」をインポートしました（元の名前: {original_project_name}、{imported_count}件のタスク、{skipped_count}件スキップ）\n※あなたがプロジェクト責任者およびすべてのタスクの担当者として設定されました', 'success')
            else:
                flash(f'プロジェクト「{project_name}」をインポートしました（{imported_count}件のタスク、{skipped_count}件スキップ）\n※あなたがプロジェクト責任者およびすべてのタスクの担当者として設定されました', 'success')

            return redirect(url_for('colrep.view_project', project_id=project_id))

        except openpyxl.utils.exceptions.InvalidFileException:
            flash('無効なExcelファイルです。', 'error')
            return redirect(url_for('colrep.index'))
        except Exception as e:
            logging.error(f"Error in import_project: {str(e)}")
            import traceback
            logging.error(traceback.format_exc())
            if 'conn' in locals() and conn.is_connected():
                conn.rollback()
            flash(f'インポート中にエラーが発生しました: {str(e)}', 'error')
            return redirect(url_for('colrep.index'))
        finally:
            if 'conn' in locals() and conn.is_connected():
                cursor.close()
                conn.close()


def _create_project_from_json_data(data, owner_user_id):
    """共通処理: JSON構造(dict)から新規プロジェクト＋テーブルを作成する。
    責任者および全タスクの担当者を owner_user_id に設定する。
    戻り値: (project_id, project_name, original_project_name, imported_count, skipped_count)
    例外はそのまま呼び出し側に伝播する。"""
    project_data = data.get('project') or {}
    tasks = data.get('tasks') or []

    original_project_name = (project_data.get('プロジェクト名') or '').strip()
    if not original_project_name:
        raise ValueError('JSONの形式が正しくありません（プロジェクト名が見つかりません）。')

    composer_content = project_data.get('Composer') or ''

    conn = mysql.connector.connect(**DatabaseConfig.fujinp())
    try:
        cursor = conn.cursor(dictionary=True)

        # プロジェクト名の重複チェックと番号付加
        project_name = original_project_name
        counter = 1
        while True:
            cursor.execute(f"SELECT id FROM {COLREP_PROJECTS_TABLE} WHERE プロジェクト名 = %s", (project_name,))
            if not cursor.fetchone():
                break
            project_name = f"{original_project_name} ({counter})"
            counter += 1

        # 新しいテーブルを作成
        table_name = f"colrep_{uuid.uuid4().hex[:12]}"
        create_table_sql = f"""
            CREATE TABLE `{table_name}` (
                `id` INT AUTO_INCREMENT PRIMARY KEY,
                `更新日時` DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                `カラム名` VARCHAR(255) NOT NULL,
                `担当者アカウント` INT NOT NULL,
                `説明` TEXT COMMENT '管理者から入力者への説明',
                `content` LONGTEXT COMMENT '入力内容',
                `備考` TEXT COMMENT '入力者から管理者への説明',
                `status` VARCHAR(20) DEFAULT '作業中' COMMENT '進捗状況：作業中/改訂中/完了',
                INDEX `idx_担当者` (`担当者アカウント`),
                INDEX `idx_カラム名` (`カラム名`),
                INDEX `idx_status` (`status`),
                FOREIGN KEY (`担当者アカウント`) REFERENCES {Tables.USERS}(`id`) ON DELETE RESTRICT
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
            COMMENT='CoRePoプロジェクト用データテーブル（JSONインポート）'
        """
        cursor.execute(create_table_sql)
        conn.commit()

        # プロジェクトを作成
        now = get_jst_now()
        cursor.execute(f"""
            INSERT INTO {COLREP_PROJECTS_TABLE}
            (プロジェクト名, 更新日時, 責任者, テーブル名, Composer, is_public)
            VALUES (%s, %s, %s, %s, %s, %s)
        """, (project_name, now, owner_user_id, table_name, composer_content, False))
        project_id = cursor.lastrowid
        conn.commit()

        # タスク行をインポート
        imported_count = 0
        skipped_count = 0
        for task in tasks:
            column_name = task.get('カラム名')
            if not column_name:
                skipped_count += 1
                continue
            cursor.execute(f"""
                INSERT INTO `{table_name}`
                (更新日時, カラム名, 担当者アカウント, 説明, content, 備考, status)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            """, (
                now,
                column_name,
                owner_user_id,
                task.get('説明') or '',
                task.get('作業内容') or '',
                task.get('備考') or '',
                task.get('ステータス') or '作業中'
            ))
            imported_count += 1
        conn.commit()

        return project_id, project_name, original_project_name, imported_count, skipped_count
    finally:
        if conn.is_connected():
            cursor.close()
            conn.close()


@colrep_bp.route('/import_project_json', methods=['GET', 'POST'])
@login_required
def import_project_json():
    """JSONファイルからプロジェクトをインポート（インポート者が責任者・担当者）"""
    permissions = get_user_permissions()
    if not (permissions['is_colrep_admin'] or permissions['is_colrep_manager']):
        flash('プロジェクト作成権限（CoRePo管理者以上）が必要です。', 'error')
        return redirect(url_for('colrep.index'))

    if request.method == 'POST':
        try:
            # ✅ インポートを実行したユーザーのIDを取得
            import_user_id = session.get('user_id')
            if not import_user_id:
                flash('ログインが必要です。', 'error')
                return redirect(url_for('colrep.index'))

            # アップロードされたファイルを取得
            if 'json_file' not in request.files:
                flash('ファイルが選択されていません。', 'error')
                return redirect(url_for('colrep.index'))

            file = request.files['json_file']
            if file.filename == '':
                flash('ファイルが選択されていません。', 'error')
                return redirect(url_for('colrep.index'))

            if not file.filename.endswith('.json'):
                flash('jsonファイルのみアップロード可能です。', 'error')
                return redirect(url_for('colrep.index'))

            # JSONファイルを読み込み
            try:
                data = json.loads(file.read().decode('utf-8'))
            except (json.JSONDecodeError, UnicodeDecodeError):
                flash('無効なJSONファイルです。', 'error')
                return redirect(url_for('colrep.index'))

            try:
                (project_id, project_name, original_project_name,
                 imported_count, skipped_count) = _create_project_from_json_data(data, import_user_id)
            except ValueError as ve:
                flash(str(ve), 'error')
                return redirect(url_for('colrep.index'))

            if original_project_name != project_name:
                flash(f'プロジェクト「{project_name}」をインポートしました（元の名前: {original_project_name}、{imported_count}件のタスク、{skipped_count}件スキップ）\n※あなたがプロジェクト責任者およびすべてのタスクの担当者として設定されました', 'success')
            else:
                flash(f'プロジェクト「{project_name}」をインポートしました（{imported_count}件のタスク、{skipped_count}件スキップ）\n※あなたがプロジェクト責任者およびすべてのタスクの担当者として設定されました', 'success')

            return redirect(url_for('colrep.view_project', project_id=project_id))

        except Exception as e:
            logging.error(f"Error in import_project_json: {str(e)}")
            import traceback
            logging.error(traceback.format_exc())
            flash(f'インポート中にエラーが発生しました: {str(e)}', 'error')
            return redirect(url_for('colrep.index'))

    return redirect(url_for('colrep.index'))


@colrep_bp.route('/import_from_archive/<int:doc_id>', methods=['POST'])
@login_required
def import_from_archive(doc_id):
    """文書アーカイブに保存されたプロジェクトソース(JSON)を、
    新規CoRePoプロジェクト（責任者・全担当者=操作者本人）としてインポートする。
    権限: システム管理者(admin) または コレポ管理者グループ所属者。"""
    permissions = get_user_permissions()
    if not (permissions['is_colrep_admin'] or permissions['is_colrep_manager']):
        flash('この操作にはCoRePo総管理者以上の権限が必要です。', 'error')
        return redirect(url_for('document_archive.dashboard'))

    import_user_id = session.get('user_id')
    if not import_user_id:
        flash('ログインが必要です。', 'error')
        return redirect(url_for('document_archive.dashboard'))

    conn = None
    try:
        # アーカイブDBから当該ドキュメントのソースJSONを取得
        conn = mysql.connector.connect(**DatabaseConfig.fujinp())
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT id, title, corepo_source_json
            FROM public_documents
            WHERE id = %s
        """, (doc_id,))
        doc = cursor.fetchone()
        cursor.close()
        conn.close()
        conn = None

        if not doc:
            flash('文書が見つかりません。', 'error')
            return redirect(url_for('document_archive.dashboard'))

        if not doc.get('corepo_source_json'):
            flash('この文書にはCoRePoプロジェクトのソース(JSON)が含まれていません。', 'error')
            return redirect(url_for('document_archive.dashboard'))

        try:
            data = json.loads(doc['corepo_source_json'])
        except (json.JSONDecodeError, TypeError):
            flash('保存されているプロジェクトソースの形式が不正です。', 'error')
            return redirect(url_for('document_archive.dashboard'))

        try:
            (project_id, project_name, original_project_name,
             imported_count, skipped_count) = _create_project_from_json_data(data, import_user_id)
        except ValueError as ve:
            flash(str(ve), 'error')
            return redirect(url_for('document_archive.dashboard'))

        if original_project_name != project_name:
            flash(f'アーカイブ文書「{doc["title"]}」からプロジェクト「{project_name}」をインポートしました（元の名前: {original_project_name}、{imported_count}件のタスク、{skipped_count}件スキップ）\n※あなたがプロジェクト責任者およびすべてのタスクの担当者として設定されました', 'success')
        else:
            flash(f'アーカイブ文書「{doc["title"]}」からプロジェクト「{project_name}」をインポートしました（{imported_count}件のタスク、{skipped_count}件スキップ）\n※あなたがプロジェクト責任者およびすべてのタスクの担当者として設定されました', 'success')

        return redirect(url_for('colrep.view_project', project_id=project_id))

    except Exception as e:
        logging.error(f"Error in import_from_archive: {str(e)}")
        import traceback
        logging.error(traceback.format_exc())
        flash(f'インポート中にエラーが発生しました: {str(e)}', 'error')
        return redirect(url_for('document_archive.dashboard'))
    finally:
        if conn is not None and conn.is_connected():
            conn.close()

@colrep_bp.route('/upload_image', methods=['POST'])
@login_required
def upload_image():
    """画像アップロード"""
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'ファイルが選択されていません'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'error': 'ファイルが選択されていません'}), 400

    if not allowed_file(file.filename):
        return jsonify({'success': False, 'error': '許可されていないファイル形式です'}), 400

    try:
        # ディレクトリ作成
        os.makedirs(UPLOAD_FOLDER, exist_ok=True)

        # ファイル名生成（タイムスタンプ付き）
        timestamp = get_jst_now().strftime('%Y%m%d_%H%M%S')
        filename = secure_filename(file.filename)
        name, ext = os.path.splitext(filename)
        unique_filename = f"{name}_{timestamp}{ext}"

        filepath = os.path.join(UPLOAD_FOLDER, unique_filename)

        if ext.lower() == '.svg':
            # SVG はサニタイズしてから保存
            svg_data = _sanitize_svg(file.read())
            with open(filepath, 'wb') as f:
                f.write(svg_data)
        else:
            file.save(filepath)

        # URL生成
        url = f"/static/mdimgs/{unique_filename}"

        return jsonify({'success': True, 'filename': unique_filename, 'url': url})

    except Exception as e:
        logging.error(f"画像アップロードエラー: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@colrep_bp.route('/edit_description_external/<int:project_id>/<int:task_id>')
@login_required
def edit_description_external(project_id, task_id):
    """タスク説明の外部エディタ"""
    user_info = get_user_info()
    user_id = user_info.get('user_id')

    try:
        # conn = mysql.connector.connect(**base_db_config, database=TARGET_DATABASE)
        conn = mysql.connector.connect(**DatabaseConfig.fujinp())
        cursor = conn.cursor(dictionary=True)

        # プロジェクト情報を取得
        cursor.execute(f"SELECT テーブル名, 責任者 FROM {COLREP_PROJECTS_TABLE} WHERE id = %s", (project_id,))
        project = cursor.fetchone()

        if not project:
            flash('プロジェクトが見つかりません。', 'error')
            return redirect(url_for('colrep.index'))

        # 権限チェック：プロジェクト責任者のみ
        if project['責任者'] != user_id:
            flash('説明の編集はプロジェクト責任者のみ実行可能です。', 'error')
            return redirect(url_for('colrep.manage_project_table', project_id=project_id))

        return render_template('colrep_description_editor.html',
                              project_id=project_id,
                              task_id=task_id)

    except Exception as e:
        logging.error(f"Error in edit_description_external: {str(e)}")
        flash(f'エラーが発生しました: {str(e)}', 'error')
        return redirect(url_for('colrep.manage_project_table', project_id=project_id))
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close()
            conn.close()


@colrep_bp.route('/get_task_description/<int:project_id>/<int:task_id>')
@login_required
def get_task_description(project_id, task_id):
    """タスク説明を取得"""
    user_info = get_user_info()
    user_id = user_info.get('user_id')

    try:
        # conn = mysql.connector.connect(**base_db_config, database=TARGET_DATABASE)
        conn = mysql.connector.connect(**DatabaseConfig.fujinp())
        cursor = conn.cursor(dictionary=True)

        # プロジェクト情報を取得
        cursor.execute(f"SELECT p.テーブル名, p.プロジェクト名, p.責任者 FROM {COLREP_PROJECTS_TABLE} p WHERE p.id = %s", (project_id,))
        project = cursor.fetchone()

        if not project:
            return jsonify({'success': False, 'error': 'プロジェクトが見つかりません。'}), 404

        table_name = project['テーブル名']

        # タスク情報を取得
        cursor.execute(f"SELECT カラム名, 説明, 担当者アカウント FROM `{table_name}` WHERE id = %s",
                       (task_id,))
        task = cursor.fetchone()

        if not task:
            return jsonify({'success': False, 'error': 'タスクが見つかりません。'}), 404

        # [2026-07-25 改修2] 権限を拡張：責任者のみ → 責任者 / システム管理者 / 当該タスクの担当者
        # 作業ダッシュボードの「詳細表示」は担当者が使うため、
        # 従来の「責任者のみ」では担当者が自分への指示を読めなかった。
        # 説明の編集（save_description_external）は従来どおり責任者のみ。
        is_system_admin = (session.get('user_category') == 'admin')
        if not (is_system_admin
                or project['責任者'] == user_id
                or task.get('担当者アカウント') == user_id):
            return jsonify({'success': False, 'error': '権限がありません。'}), 403

        return jsonify({
            'success': True,
            'task_title': task['カラム名'],
            'project_name': project['プロジェクト名'],
            'description': task['説明'] or ''
        })

    except Exception as e:
        logging.error(f"Error in get_task_description: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close()
            conn.close()


@colrep_bp.route('/save_description_external/<int:project_id>/<int:task_id>', methods=['POST'])
@login_required
def save_description_external(project_id, task_id):
    """タスク説明を保存"""
    user_info = get_user_info()
    user_id = user_info.get('user_id')

    data = request.get_json()
    description = data.get('description', '').strip()

    try:
        # conn = mysql.connector.connect(**base_db_config, database=TARGET_DATABASE)
        conn = mysql.connector.connect(**DatabaseConfig.fujinp())
        cursor = conn.cursor(dictionary=True)

        # プロジェクト情報を取得
        cursor.execute(f"SELECT テーブル名, 責任者 FROM {COLREP_PROJECTS_TABLE} WHERE id = %s", (project_id,))
        project = cursor.fetchone()

        if not project:
            return jsonify({'success': False, 'error': 'プロジェクトが見つかりません。'}), 404

        # 権限チェック：プロジェクト責任者のみ
        if project['責任者'] != user_id:
            return jsonify({'success': False, 'error': '権限がありません。'}), 403

        table_name = project['テーブル名']

        # 説明を更新
        now = get_jst_now()
        update_query = f"UPDATE `{table_name}` SET 説明 = %s, 更新日時 = %s WHERE id = %s"
        cursor.execute(update_query, (description, now, task_id))
        conn.commit()

        return jsonify({'success': True, 'message': '説明を保存しました。'})

    except Exception as e:
        logging.error(f"Error in save_description_external: {str(e)}")
        if 'conn' in locals() and conn.is_connected():
            conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close()
            conn.close()

@colrep_bp.route('/get_task_data/<int:project_id>/<int:task_id>')
@login_required
def get_task_data(project_id, task_id):
    try:
        user_id = session.get('user_id')

        # conn = mysql.connector.connect(**base_db_config, database=TARGET_DATABASE)
        conn = mysql.connector.connect(**DatabaseConfig.fujinp())
        cursor = conn.cursor(dictionary=True)

        # プロジェクト存在確認
        cursor.execute(f"SELECT テーブル名, 責任者 FROM {COLREP_PROJECTS_TABLE} WHERE id = %s", (project_id,))
        project = cursor.fetchone()

        if not project:
            return jsonify({'success': False, 'error': 'プロジェクトが見つかりません'}), 404

        table_name = project['テーブル名']

        # タスクデータを取得
        query = f"SELECT * FROM `{table_name}` WHERE id = %s"
        cursor.execute(query, (task_id,))
        task = cursor.fetchone()

        if not task:
            return jsonify({'success': False, 'error': 'タスクが見つかりません'}), 404

        # [2026-07-25 改修] 権限チェックを追加（従来は誰でもタスク本文を取得できた）
        # admin / プロジェクト責任者 / そのタスクの担当者 のみ
        is_system_admin = (session.get('user_category') == 'admin')
        if not (is_system_admin
                or project['責任者'] == user_id
                or task.get('担当者アカウント') == user_id):
            return jsonify({'success': False,
                            'error': 'このタスクを参照する権限がありません'}), 403

        # ★★★ 修正箇所: serialize_for_json でラップする ★★★
        return jsonify({'success': True, 'task': serialize_for_json(task)})

    except Exception as e:
        logging.error(f"Error getting task data: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close()
            conn.close()

@colrep_bp.route('/return_to_fujin')
@login_required
def return_to_fujin():
    """FUJINダッシュボードに戻る"""
    return redirect_to_dashboard()